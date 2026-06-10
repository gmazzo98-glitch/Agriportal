import streamlit as st
import plotly.graph_objects as go
import pandas as pd
import sys
import os
import copy
import io
from datetime import date
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import PageBreak
from reportlab.lib.utils import ImageReader
import plotly.io as pio

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from core.roi_calculate import calculate
from core.greenhouse_calculate import calculate_greenhouse
from core.greenhouse_data_tables import GREENHOUSE_CROPS, POLYTUNNEL_CROPS, FISH_SPECIES, CROP_NUTRIENT_DEMAND, COUPLING_PARAMS
from core.aquaponics_calculate import calculate_aquaponics, calculate_fish, COUNTRY_AMBIENT_TEMP
from core.data_tables import COUNTRIES, CROPS, LIGHTS
from core._charts import style_fig
from core._tables import severity_cell, MATCH
from core.climate import fetch_climate_profile, compute_natural_dli_fraction
from core.energy_labour import get_rates_for_country_name, get_full_rates
from core._styles import inject_styles
from core.auth import require_login, current_user
from core.farm_context import render_farm_context_sidebar, load_farm, clear_farm, MODALITY_RADIO
import json
from supabase import create_client, Client
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

@st.cache_resource
def get_supabase() -> Client:
    return create_client(st.secrets["SUPABASE_URL"], st.secrets["SUPABASE_KEY"])

supabase = get_supabase()

st.set_page_config(page_title="ROI Calculator", page_icon="📊", layout="wide") # Keep page_icon emoji
inject_styles()
require_login()

# ── Sidebar Dropdown & Expander Readability Fix ───────────────────────────
st.markdown("""
<style>
  /* Selectboxes and Multiselects inside sidebar expanders */
  /* Since expanders have a light background, these must use the light-theme palette 
     to avoid dark-on-dark text from inheriting sidebar styles. */
  [data-testid="stSidebar"] [data-testid="stExpander"] [data-baseweb="select"] > div {
    background-color: var(--surface) !important;
    border-color: var(--rule) !important;
  }
  [data-testid="stSidebar"] [data-testid="stExpander"] [data-baseweb="select"] * {
    color: var(--ink) !important;
    fill: var(--ink) !important;
  }
  /* Sidebar selectboxes outside of expanders: ensure borders use style tokens */
  [data-testid="stSidebar"] [data-baseweb="select"] > div {
    border-color: var(--rule-soft) !important;
  }

  /* ── Modality Radio Button Visibility Fix ── */
  /* Unselected state circle: ensure it's visible against the background */
  [data-testid="stRadio"] [data-baseweb="radio"] div:first-child {
    border-color: var(--ink-2) !important;
    background-color: var(--surface-2) !important;
  }
  /* Selected state circle: use botanical sage */
  [data-testid="stRadio"] [data-checked="true"] div:first-child {
    background-color: var(--surface) !important;
    border-color: var(--sage) !important;
  }
  /* Selected inner dot: botanical green (matching slider) */
  [data-testid="stRadio"] [data-checked="true"] div:first-child div {
    background-color: var(--sage) !important;
  }
</style>
""", unsafe_allow_html=True)

def _render_farm_selector_sidebar():
    """
    Persistent farm selector rendered at the top of the sidebar on every modality.
    Sets st.session_state["active_farm"] and triggers st.rerun() on load.
    Returns the active farm dict or None.
    """
    with st.sidebar:
        st.markdown("### 🌱 Active Farm")
        try:
            _resp = supabase.table("farms").select(
                "id, name, modality, country, crop, footprint, levels, lights_tier, hvac, "
                "automation, price_scenario, price_override, packaging_cost, loss_rate, "
                "net_grow_factor, walkways_factor, water_price, rent_monthly, real_estate_capex, "
                "harvest_mode, depreciation_years, tax_rate, ltv, interest_rate, loan_term_years, "
                "lat, lon, crop_mix_json, ambient_temp_annual, mean_annual_dli, crop_source, discount_rate"
            ).order("created_at", desc=True).execute()
            _farms_list = _resp.data or []
        except Exception:
            _farms_list = []

        _active = st.session_state.get("active_farm")
        _active_name = _active["name"] if _active else None

        if _active:
            _mod_badge = { # Keep emojis in _mod_badge dictionary
                "vertical_farm": "🏭",
                "greenhouse": "🌿",
                "polytunnel": "🌿",
                "aquaponics_decoupled": "🐟",
                "aquaponics_coupled": "♻️",
            }.get(_active.get("modality", ""), "🌱")
            st.success(f"{_mod_badge} **{_active['name']}**")
            if _active.get("country"): # Keep emoji in success message
                st.caption(f"📍 {_active['country']}")
        else:
            st.info("No farm loaded. Select or create one below.")

        _farm_names = ["— select —"] + [f["name"] for f in _farms_list]
        _sel = st.selectbox(
            "Load farm",
            options=_farm_names,
            index=_farm_names.index(_active_name) if _active_name in _farm_names else 0,
            key="global_farm_selector",
        )

        if _sel != "— select —":
            _farm = next((f for f in _farms_list if f["name"] == _sel), None)
            if _farm and (not _active or _active.get("name") != _sel):
                if st.button("⬇️ Load", use_container_width=True, key="global_farm_load_btn"):
                    # Clear stale multi-crop row keys so _pending_farm_load handler writes cleanly
                    for _sli in range(6):
                        for _spfx in ("roi", "gh", "aq"):
                            st.session_state.pop(f"{_spfx}_mix_crop_{_sli}", None)
                            st.session_state.pop(f"{_spfx}_mix_pct_{_sli}", None)
                    for _smk in ("roi_multi_crop", "gh_multi_crop", "aq_multi_crop",
                                 "roi_crop_mix", "gh_crop_mix", "aq_crop_mix"):
                        st.session_state.pop(_smk, None)
                    st.session_state["active_farm"]        = _farm
                    st.session_state["_pending_farm_load"] = _farm
                    st.session_state["gh_country"]         = _farm.get("country", "Germany")
                    st.session_state["gh_footprint"]       = int(_farm.get("footprint") or 5000)
                    st.session_state["gh_automation"]      = _farm.get("automation", "Medium")
                    _sl_gh_src = ("Polytunnel" if (_farm.get("crop_source") or "greenhouse").lower() == "polytunnel" else "Greenhouse")
                    st.session_state["gh_crop_source"]     = _sl_gh_src
                    _sl_gh_dict = POLYTUNNEL_CROPS if _sl_gh_src == "Polytunnel" else GREENHOUSE_CROPS
                    _sl_gh_crop = _farm.get("crop", "")
                    # Migration: Sweet Pepper renamed
                    if _sl_gh_crop == "Sweet Pepper":
                        _sl_gh_crop = "Sweet Pepper (GH Substrate)"
                    st.session_state["gh_crop"]            = _sl_gh_crop if _sl_gh_crop in _sl_gh_dict else list(_sl_gh_dict.keys())[0]
                    st.session_state["aq_country"]         = _farm.get("country", "Germany")
                    st.session_state["aq_plant_crop"]      = _farm.get("crop", "Lettuce (Butterhead)")
                    if _farm.get("lat") and _farm.get("lon"):
                        st.session_state["shared_lat"] = _farm["lat"]
                        st.session_state["shared_lng"] = _farm["lon"]
                        st.session_state["fim_lat"]    = _farm["lat"]
                        st.session_state["fim_lng"]    = _farm["lon"]
                    _mod = _farm.get("modality", "vertical_farm")
                    _mod_map = { # Keep emojis in _mod_map dictionary
                        "vertical_farm":        "🏭 Indoor Vertical Farm",
                        "greenhouse":           "🌿 High-Tech Greenhouse",
                        "polytunnel":           "🌿 High-Tech Greenhouse",
                        "aquaponics_decoupled": "🐟 Decoupled Aquaponics",
                        "aquaponics_coupled":   "♻️ Coupled Aquaponics",
                    }
                    st.session_state["_pending_modality"] = _mod_map.get(_mod, "🏭 Indoor Vertical Farm")
                    st.rerun()

        if _active:
            if st.button("✖ Clear farm", use_container_width=True, key="global_farm_clear_btn"): # Keep emoji in button
                st.session_state.pop("active_farm", None)
                st.rerun()

        st.divider()
        return st.session_state.get("active_farm")


# ═══════════════════════════════════════════════════════════════
# UNIFIED PDF ENGINE — all modalities
# ═══════════════════════════════════════════════════════════════
# ─────────────────────────────────────────────────────────────────────────────
# UNIFIED PDF ENGINE — all modalities
# Design system: ink-on-paper, sage accent, JetBrains Mono numerics
# ─────────────────────────────────────────────────────────────────────────────

# ═══════════════════════════════════════════════════════════════
# UNIFIED PDF ENGINE — all modalities
# ═══════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════════════
# UNIFIED PDF ENGINE — Agricultural Intelligence Portal
# Consultant-grade feasibility report · all modalities
# Style: ink-on-paper, sage accent, mono numerics
# ═══════════════════════════════════════════════════════════════════════════════

def _build_feasibility_pdf(
    result_dict: dict,
    inputs_dict: dict,
    modality: str,
    farm_name: str = "",
    run_sens_fn=None,          # callable(kwh_m,lab_m,yld_m,prc_m)->dict for sensitivity
    aq_plant_sens_inputs=None, # needed for AQ plant-side sensitivity
) -> bytes:
    """
    Single entry point for all modalities.
    modality: "vf" | "gh" | "pt" | "aqd" | "aqc"
    Returns PDF bytes.
    """
    import io, hashlib, copy
    from datetime import date as _date
    import plotly.graph_objects as go
    import plotly.io as pio
    from reportlab.platypus import (
        BaseDocTemplate, PageTemplate, Frame, Paragraph, Spacer,
        Table, TableStyle, HRFlowable, PageBreak, KeepTogether,
        Image as RLImage,
    )
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
    from reportlab.graphics.shapes import Drawing, Rect, Line
    from reportlab.graphics import renderPDF

    # ── Design tokens ──────────────────────────────────────────────────────────
    INK       = colors.HexColor("#161a16")
    INK_2     = colors.HexColor("#4a524a")
    INK_3     = colors.HexColor("#7a807a")
    INK_4     = colors.HexColor("#aeb2a8")
    LINEN     = colors.HexColor("#f4f1ea")
    LINEN_2   = colors.HexColor("#fbf9f4")
    RULE      = colors.HexColor("#d6d2c4")
    RULE_SOFT = colors.HexColor("#ece8db")
    SAGE      = colors.HexColor("#2f5d3a")
    SAGE_TINT = colors.HexColor("#e6ede4")
    CLAY      = colors.HexColor("#b85c38")
    CLAY_TINT = colors.HexColor("#fdf0ea")
    AMBER_C   = colors.HexColor("#c08a2e")
    AMBER_TINT= colors.HexColor("#fdf6e3")

    # TODO: register Inter / JetBrains Mono from assets/fonts/ when available
    SANS   = "Helvetica"
    SANS_B = "Helvetica-Bold"
    MONO   = "Courier"
    MONO_B = "Courier-Bold"

    PW, PH = A4
    L_BAND = 7*mm
    LM, RM, TM, BM = 24*mm, 18*mm, 16*mm, 20*mm
    BW = PW - LM - RM   # body width

    # ── Document ID ────────────────────────────────────────────────────────────
    MC = {"vf":"VF","gh":"GH","pt":"PT","aqd":"AQD","aqc":"AQC"}.get(modality,"XX")
    _today = _date.today()
    _ds = _today.strftime("%Y%m%d")
    _rd = _today.strftime("%d %B %Y")
    _nnn = int(hashlib.md5((farm_name + _ds).encode()).hexdigest(),16) % 1000
    DOC_ID = f"DOC {MC}-{_ds}-{_nnn:03d}"

    IS_AQ = modality in ("aqd","aqc")
    ML = {"vf":"Vertical Farm","gh":"High-Tech Greenhouse","pt":"Polytunnel",
          "aqd":"Decoupled Aquaponics","aqc":"Coupled Aquaponics"}.get(modality,"—")

    # ── Normalise result dicts ────────────────────────────────────────────────
    if IS_AQ:
        _pr = result_dict["plant"]
        _fr = result_dict["fish"]
        _cr = result_dict   # combined
    else:
        _pr = result_dict
        _fr = None
        _cr = result_dict

    # Plant scalars
    P_REV    = _pr.get("annual_revenue",0)
    P_EBITDA = _pr.get("ebitda",0)
    P_CAPEX  = _pr.get("total_capex",0)
    P_COSTS  = _pr.get("total_annual_costs",0)
    P_KG     = _pr.get("total_annual_kg",0)
    P_PRICE  = _pr.get("effective_price",0)
    P_ENERGY = _pr.get("annual_energy_cost",0)
    P_LABOUR = _pr.get("annual_labour_cost",0)
    P_VAR    = _pr.get("annual_variable_cost",0)
    P_WATER  = _pr.get("annual_water_cost",0)
    P_MAINT  = _pr.get("annual_maintenance",0)
    P_RENT   = _pr.get("annual_rent",0)
    P_KWH    = _pr.get("total_annual_kwh", _pr.get("annual_kwh",0))
    P_MARGIN = _pr.get("ebitda_margin",0)
    P_DCF    = _pr.get("dcf_cashflows",[])
    P_NPV    = _pr.get("npv",0)
    P_DSCR   = _pr.get("dscr")
    P_PBK    = _pr.get("payback_years")
    P_EGA    = _pr.get("effective_grow_area", inputs_dict.get("footprint",0))
    P_GA     = _pr.get("gross_area", inputs_dict.get("footprint",0))
    P_CY     = _pr.get("cycles_per_year",0)
    P_ECD    = _pr.get("effective_cycle_days",0)
    P_LH     = _pr.get("annual_labour_hours",0)
    P_DEPR   = _pr.get("annual_depreciation",0)
    P_DEBT   = _pr.get("debt_amount",0)
    P_DS     = _pr.get("annual_debt_service",0)
    P_EBIT   = _pr.get("ebit",0)
    P_NI     = _pr.get("net_income",0)
    P_TAX    = _pr.get("tax_charge",0)
    P_EQ     = _pr.get("equity_invested", P_CAPEX*(1-inputs_dict.get("ltv",0)/100))
    P_EPCT   = P_ENERGY/P_REV*100 if P_REV else 0

    # Derived plant metrics
    P_BE_PRICE = P_COSTS/P_KG if P_KG else None
    P_HDROOM   = (P_PRICE-P_BE_PRICE)/P_PRICE*100 if P_BE_PRICE and P_PRICE else None
    _loss_r    = inputs_dict.get("loss_rate",5)/100
    _denom     = P_PRICE*(1-_loss_r)*P_CY*P_EGA if P_EGA else 0
    P_BE_YIELD = P_COSTS/_denom if _denom else None
    P_REV_M2   = P_REV/P_EGA if P_EGA else 0
    P_KG_M2    = P_KG/P_EGA if P_EGA else 0
    P_KWH_KG   = P_KWH/P_KG if P_KG else 0
    P_ENERGY_KG= P_ENERGY/P_KG if P_KG else 0
    P_LABOUR_KG= P_LABOUR/P_KG if P_KG else 0

    # Combined/fish scalars
    if IS_AQ:
        F_REV    = _fr.get("annual_fish_revenue",0)
        F_EBITDA = _fr.get("fish_ebitda",0)
        F_CAPEX  = _fr.get("total_fish_capex",0)
        F_COSTS  = _fr.get("total_fish_costs",0)
        F_KG     = _fr.get("annual_kg_fish",0)
        F_PRICE  = _fr.get("effective_fish_price",0)
        F_ENERGY = _fr.get("annual_fish_energy_cost",0)
        F_FEED   = _fr.get("annual_feed_cost",0)
        F_FING   = _fr.get("annual_fingerling_cost",0)
        F_WATER  = _fr.get("annual_water_cost",0)
        F_LABOUR = _fr.get("annual_fish_labour_cost",0)
        F_MAINT  = _fr.get("annual_fish_maintenance",0)
        F_MARGIN = _fr.get("fish_ebitda_margin",0)
        F_DSCR   = _fr.get("dscr")
        F_DCF    = _fr.get("dcf_cashflows",[])
        F_NPV    = _fr.get("npv",0)
        F_SPECIES= _fr.get("species","—")
        F_TVOL   = _fr.get("tank_volume_m3",0)
        F_KPC    = _fr.get("kg_per_cycle",0)
        F_CY     = _fr.get("cycles_per_year",0)
        F_SCALE  = _fr.get("system_scale","—")
        F_DELTA_T= _fr.get("delta_t",0)
        F_HHEAT  = _fr.get("heating_kwh",0)
        F_HAER   = _fr.get("aeration_kwh",0)
        F_HPUMP  = _fr.get("pump_kwh",0)
        F_TKAP   = _fr.get("tank_capex",0)
        F_FKAP   = _fr.get("filtration_capex",0)
        F_AKAP   = _fr.get("aeration_capex",0)
        F_MKAP   = _fr.get("monitoring_capex",0)
        F_PKAP   = _fr.get("plumbing_capex",0)
        NUTR_SAV = _cr.get("nutrient_offset_saving",0)
        INT_CAP  = _cr.get("integration_capex",0)
        C_REV    = _cr.get("combined_revenue", P_REV+F_REV)
        C_EBITDA = _cr.get("combined_ebitda",  P_EBITDA+F_EBITDA)
        C_CAPEX  = _cr.get("combined_capex",   P_CAPEX+F_CAPEX+INT_CAP)
        C_COSTS  = _cr.get("combined_costs",   P_COSTS+F_COSTS)
        C_MARGIN = _cr.get("combined_ebitda_margin", C_EBITDA/C_REV if C_REV else 0)
        C_DSCR   = _cr.get("combined_dscr", P_DSCR)
        F_PCT_REV= F_REV/C_REV*100 if C_REV else 0
        C_PBK    = P_PBK  # plant payback
    else:
        C_REV=P_REV; C_EBITDA=P_EBITDA; C_CAPEX=P_CAPEX; C_MARGIN=P_MARGIN
        C_DSCR=P_DSCR; C_PBK=P_PBK; C_COSTS=P_COSTS

    # Viability
    if P_EPCT < 30:   _viab, _vc = "VIABLE",    SAGE
    elif P_EPCT < 60: _viab, _vc = "MARGINAL",  AMBER_C
    else:             _viab, _vc = "NOT VIABLE", CLAY

    _vc_hex = {SAGE:"#2f5d3a", AMBER_C:"#c08a2e", CLAY:"#b85c38"}[_vc]

    # ── Sensitivity runs ───────────────────────────────────────────────────────
    # Accept None if no helper provided (PDF still works, just skips tornado)
    def _sens(fn, **kw):
        if fn is None: return None
        try:
            return fn(**kw)
        except Exception:
            return None

    _tvars = []
    if run_sens_fn:
        _sens_defs = [
            ("Selling Price", dict(prc_m=0.80), dict(prc_m=1.20), "Price −20%","Price +20%"),
            ("Energy Cost",   dict(kwh_m=1.50), dict(kwh_m=0.70), "Energy +50%","Energy −30%"),
            ("Yield",         dict(yld_m=0.80), dict(yld_m=1.20), "Yield −20%","Yield +20%"),
            ("Labour Cost",   dict(lab_m=1.30), dict(lab_m=0.80), "Labour +30%","Labour −20%"),
        ]
        for lbl, pk, ok, pl, ol in _sens_defs:
            _p = _sens(run_sens_fn, **pk)
            _o = _sens(run_sens_fn, **ok)
            if _p and _o:
                _dp = _p["ebitda"] - P_EBITDA
                _do = _o["ebitda"] - P_EBITDA
                _tvars.append({"label":lbl,"dp":_dp,"do":_do,
                               "pl":pl,"ol":ol,"swing":abs(_do-_dp)})
        _tvars.sort(key=lambda x: x["swing"], reverse=True)

        # Three scenarios (price)
        _s_lo = _sens(run_sens_fn, prc_m=0.80)
        _s_hi = _sens(run_sens_fn, prc_m=1.20)
        _scen_names  = ["Low (−20%)", "Base", "High (+20%)"]
        _scen_ebitda = [_s_lo["ebitda"] if _s_lo else 0, P_EBITDA, _s_hi["ebitda"] if _s_hi else 0]
        _scen_rev    = [_s_lo["annual_revenue"] if _s_lo else 0, P_REV, _s_hi["annual_revenue"] if _s_hi else 0]
        _scen_margin = [_s_lo["ebitda_margin"]*100 if _s_lo else 0, P_MARGIN*100, _s_hi["ebitda_margin"]*100 if _s_hi else 0]
    else:
        _scen_names = _scen_ebitda = _scen_rev = _scen_margin = None

    # ── Paragraph style helper ─────────────────────────────────────────────────
    _ps_cache = {}
    def ps(name, sz, font=SANS, col=INK, align=TA_LEFT, sb=0, sa=3, lm=1.38):
        k = (name, sz, font, str(col), align, sb, sa, lm)
        if k not in _ps_cache:
            _ps_cache[k] = ParagraphStyle(
                name, fontName=font, fontSize=sz, textColor=col,
                alignment=align, spaceBefore=sb, spaceAfter=sa, leading=sz*lm)
        return _ps_cache[k]

    def P(txt, style): return Paragraph(txt, style)

    # Shorthand styles
    Seyebrow = ps("ey",  7.5,MONO_B, INK_3, sa=2)
    Stitle   = ps("ti",  26, SANS_B, INK,   sa=3, lm=1.05)
    Ssub     = ps("sb",  10, SANS,   INK_2, sa=6)
    Ssect    = ps("sc",  11, SANS_B, INK,   sb=8, sa=4)
    Sbody    = ps("bo",   9.5,SANS,  INK,   sa=4, lm=1.5)
    Sbody2   = ps("b2",   9, SANS,   INK_2, sa=3, lm=1.5)
    Scap     = ps("ca",   8, SANS,   INK_3, sa=3, lm=1.5)
    Scap_i   = ps("ci",   8, SANS,   INK_3, sa=3, lm=1.5)
    Smethlab = ps("ml",   7.5,MONO_B,SAGE,  sa=3)
    Smethbod = ps("mb",   9, SANS,   INK_2, sa=3, lm=1.55)
    Skpilbl  = ps("kl",   7.5,MONO_B,INK_2, sa=1)
    Skpival  = ps("kv",  22, MONO_B, INK,   sa=0)
    Skpival_s= ps("kvs", 22, MONO_B, SAGE,  sa=0)
    Skpival_c= ps("kvc", 22, MONO_B, CLAY,  sa=0)
    Skpisub  = ps("ks",   8, SANS,   INK_3, sa=0)
    Sthlbl   = ps("th",   7.5,SANS_B,INK_2, sa=0)
    Sthlbl_r = ps("thr",  7.5,SANS_B,INK_2, align=TA_RIGHT, sa=0)
    Stbdy    = ps("tb",   9.5,SANS,  INK,   sa=0)
    Stnum    = ps("tn",   9.5,MONO,  INK,   align=TA_RIGHT, sa=0)
    Stnum_s  = ps("tns",  9.5,MONO_B,SAGE,  align=TA_RIGHT, sa=0)
    Stnum_c  = ps("tnc",  9.5,MONO_B,CLAY,  align=TA_RIGHT, sa=0)
    Stnum_3  = ps("tn3",  8.5,MONO,  INK_3, align=TA_RIGHT, sa=0)
    Stnote   = ps("tno",  8.5,SANS,  INK_2, sa=0, lm=1.4)
    Scfgk    = ps("ck",   9, SANS,   INK_2, sa=0)
    Scfgv    = ps("cv",   9.5,MONO,  INK,   align=TA_RIGHT, sa=0)
    Scfgvb   = ps("cvb",  9.5,MONO_B,INK,   align=TA_RIGHT, sa=0)
    Sfooter  = ps("ft",   7, SANS,   INK_3, align=TA_CENTER, sa=0)

    def _mk(v, prefix="$", suf="", dec=0, s=None):
        neg = v < 0
        txt = f"{prefix}{abs(v):,.{dec}f}{suf}"
        if neg: txt = f"−{txt}"
        st = s or (Stnum_c if neg else Stnum)
        return P(txt, st)

    def _dash(): return P("—", Stnum_3)

    # ── Running chrome (every page) ────────────────────────────────────────────
    def _chrome(canvas, doc):
        canvas.saveState()
        # Left band
        canvas.setFillColor(SAGE)
        canvas.rect(0, PH*0.72, L_BAND, PH*0.28, fill=1, stroke=0)
        canvas.setFillColor(LINEN)
        canvas.rect(0, 0, L_BAND, PH*0.72, fill=1, stroke=0)
        # Rotated label
        canvas.saveState()
        canvas.setFont(MONO, 6)
        canvas.setFillColor(INK_3)
        canvas.translate(L_BAND/2, PH*0.42)
        canvas.rotate(90)
        canvas.drawCentredString(0, 0, "CEA FEASIBILITY  ·  VOL. II")
        canvas.restoreState()
        # Header
        canvas.setStrokeColor(RULE); canvas.setLineWidth(0.5)
        canvas.line(LM, PH-TM+2*mm, PW-RM, PH-TM+2*mm)
        canvas.setFont(MONO, 7); canvas.setFillColor(INK_3)
        canvas.drawString(LM, PH-TM+4*mm, "AGRIPORTAL  ·  AGRICULTURAL INTELLIGENCE")
        canvas.drawRightString(PW-RM, PH-TM+4*mm, DOC_ID)
        # Footer
        canvas.line(LM, BM-3*mm, PW-RM, BM-3*mm)
        canvas.setFont(SANS_B, 7); canvas.setFillColor(INK_3)
        canvas.drawString(LM, BM-7*mm, f"AGRIPORTAL V2  ·  {DOC_ID}")
        canvas.setFont(SANS, 7)
        canvas.drawCentredString(PW/2, BM-7*mm, "Indicative model output — not investment advice.")
        canvas.drawRightString(PW-RM, BM-7*mm, f"PAGE {doc.page:02d} OF 06")
        canvas.restoreState()

    # ── Chart helpers ──────────────────────────────────────────────────────────
    def _theme(fig):
        fig.update_layout(
            font=dict(family="Helvetica,sans-serif", color="#161a16", size=10),
            paper_bgcolor="#fbf9f4", plot_bgcolor="#fbf9f4",
            margin=dict(l=50,r=16,t=8,b=36),
            title=None, showlegend=False,
            xaxis=dict(showgrid=False, zeroline=False, showline=True,
                       linecolor="#161a16", linewidth=1,
                       tickfont=dict(family="Courier",color="#7a807a",size=9)),
            yaxis=dict(showgrid=True, gridcolor="#ece8db", gridwidth=0.8,
                       zeroline=True, zerolinecolor="#161a16", zerolinewidth=1,
                       tickfont=dict(family="Courier",color="#7a807a",size=9)),
        )
        return fig

    def _img(fig, w=BW, h=66*mm, pw=1600, ph=560):
        _theme(fig)
        png = fig.to_image(format="png", width=pw, height=ph, scale=2)
        box = Table([[RLImage(io.BytesIO(png), width=w, height=h)]],
                    colWidths=[w])
        box.setStyle(TableStyle([
            ("BOX",(0,0),(-1,-1),0.5,RULE),
            ("BACKGROUND",(0,0),(-1,-1),LINEN_2),
            ("TOPPADDING",(0,0),(-1,-1),2),("BOTTOMPADDING",(0,0),(-1,-1),2),
            ("LEFTPADDING",(0,0),(-1,-1),2),("RIGHTPADDING",(0,0),(-1,-1),2),
        ]))
        return box

    # ── Section header ─────────────────────────────────────────────────────────
    def _sh(num, title, hint=""):
        n = P(f'<font color="#2f5d3a"><b>{num}</b></font>', ps("sn",11,MONO_B,INK_3,sa=0))
        t = P(f"<b>{title.upper()}</b>", ps("st",11,SANS_B,INK,sa=0))
        h = P(hint, ps("sh",7.5,MONO,INK_3,align=TA_RIGHT,sa=0))
        row = Table([[n,t,h]], colWidths=[10*mm, BW-62*mm, 52*mm])
        row.setStyle(TableStyle([
            ("VALIGN",(0,0),(-1,-1),"BOTTOM"),
            ("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0),
            ("BOTTOMPADDING",(0,0),(-1,-1),2),("TOPPADDING",(0,0),(-1,-1),0),
            ("LINEBELOW",(0,0),(-1,0),0.5,INK),
        ]))
        return [row, Spacer(1,4*mm)]

    def _ch(num, title, hint=""):
        n = P(f'<font color="#2f5d3a"><b>{num}</b></font>', ps("cn",8.5,MONO_B,INK_3,sa=0))
        t = P(f"<b>{title.upper()}</b>", ps("ct",9,SANS_B,INK,sa=0))
        h = P(hint, ps("ch",7.5,MONO,INK_3,align=TA_RIGHT,sa=0))
        row = Table([[n,t,h]], colWidths=[9*mm, BW-62*mm, 53*mm])
        row.setStyle(TableStyle([
            ("VALIGN",(0,0),(-1,-1),"BOTTOM"),
            ("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0),
            ("BOTTOMPADDING",(0,0),(-1,-1),2),("TOPPADDING",(0,0),(-1,-1),0),
            ("LINEBELOW",(0,0),(-1,0),0.3,RULE),
        ]))
        return [row, Spacer(1,2*mm)]

    def _subh(txt):
        return P(f'<font color="#2f5d3a">▮</font>  <b>{txt}</b>',
                 ps("sbh",10,SANS_B,INK,sb=5,sa=3))

    def _mbox(label, *paras):
        rows = [[P(label, Smethlab)]] + [[P(p, Smethbod)] for p in paras]
        inner = Table(rows, colWidths=[BW-14*mm])
        inner.setStyle(TableStyle([("TOPPADDING",(0,0),(-1,-1),2),
                                    ("BOTTOMPADDING",(0,0),(-1,-1),2),
                                    ("LEFTPADDING",(0,0),(-1,-1),4),
                                    ("RIGHTPADDING",(0,0),(-1,-1),4)]))
        outer = Table([[inner]], colWidths=[BW])
        outer.setStyle(TableStyle([
            ("BOX",(0,0),(-1,-1),0.5,RULE),
            ("BACKGROUND",(0,0),(-1,-1),LINEN_2),
            ("LINEAFTER",(0,0),(0,-1),2.0,SAGE),
            ("LEFTPADDING",(0,0),(-1,-1),6),
            ("RIGHTPADDING",(0,0),(-1,-1),4),
            ("TOPPADDING",(0,0),(-1,-1),4),
            ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ]))
        return outer

    # ── Financial table ────────────────────────────────────────────────────────
    def _ftbl(data, cw, erow=None, totrow=None):
        t = Table(data, colWidths=cw, repeatRows=1)
        ts = [
            ("FONTNAME",(0,0),(-1,0),SANS_B),("FONTSIZE",(0,0),(-1,-1),8.5),
            ("BACKGROUND",(0,0),(-1,0),LINEN_2),("TEXTCOLOR",(0,0),(-1,0),INK_2),
            ("LINEBELOW",(0,0),(-1,0),0.5,INK_3),
            ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
            ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ]
        for i in range(1,len(data)):
            ts.append(("LINEBELOW",(0,i),(-1,i),0.3,RULE_SOFT))
        if erow is not None:
            ts += [("LINEABOVE",(0,erow),(-1,erow),0.5,INK),
                   ("LINEBELOW",(0,erow),(-1,erow),0.5,INK),
                   ("BACKGROUND",(0,erow),(-1,erow),LINEN),
                   ("FONTNAME",(0,erow),(0,erow),SANS_B)]
        if totrow is not None:
            ts += [("LINEABOVE",(0,totrow),(-1,totrow),0.5,INK),
                   ("BACKGROUND",(0,totrow),(-1,totrow),LINEN),
                   ("FONTNAME",(0,totrow),(-1,totrow),SANS_B)]
        t.setStyle(TableStyle(ts))
        return t

    # ── Trajectory spark bar ───────────────────────────────────────────────────
    def _spark(v, max_v, w=34*mm, h=4*mm):
        d = Drawing(w, h)
        d.add(Rect(0,0,w,h, fillColor=LINEN, strokeColor=None))
        d.add(Line(w/2,0,w/2,h, strokeColor=INK_3, strokeWidth=0.4))
        if max_v:
            fw = abs(v)/max_v * (w/2)
            fc = colors.HexColor("#2f5d3a55") if v >= 0 else colors.HexColor("#b85c3855")
            x  = w/2 if v >= 0 else w/2 - fw
            d.add(Rect(x, 0.5, fw, h-1, fillColor=fc, strokeColor=None))
        return d

    # ══════════════════════════════════════════════════════════════════════════
    # PAGE 1 — Cover, Viability, KPIs, EBITDA Bridge
    # ══════════════════════════════════════════════════════════════════════════
    def _p1():
        E = []
        # Eyebrow
        E.append(P(f"CEA FEASIBILITY ASSESSMENT  ·  {ML.upper()}", Seyebrow))
        E.append(Spacer(1,2*mm))
        # Title
        _crop_name_from_inputs = inputs_dict.get("crop") or inputs_dict.get("plant_crop","—")
        _crop_mix_json_raw = inputs_dict.get("crop_mix_json")
        _primary_crop_for_display = None
        if _crop_mix_json_raw:
            try:
                _parsed_mix = json.loads(_crop_mix_json_raw) if isinstance(_crop_mix_json_raw, str) else _crop_mix_json_raw
                if isinstance(_parsed_mix, list) and _parsed_mix:
                    _primary_crop_for_display = _parsed_mix[0]["crop"]
            except Exception:
                pass

        _crop_display_name = _primary_crop_for_display or _crop_name_from_inputs

        if IS_AQ:
            E.append(P(f'<b>{_crop_display_name}</b>  <font color="#2f5d3a"><i>×</i></font>  <b>{F_SPECIES}</b>',
                       ps("ti2",24,SANS_B,INK,sa=3,lm=1.05)))
        else:
            E.append(P(f"<b>{_crop_display_name}</b>", Stitle))
        # Subline
        _fp = inputs_dict.get("footprint") or inputs_dict.get("plant_footprint",0)
        _parts = [inputs_dict.get("country","—"), f"{int(_fp):,} m² plant area"]
        if IS_AQ: _parts.append(f"{F_TVOL:.0f} m³ fish tank")
        _parts += [inputs_dict.get("automation","—"), _rd]
        E.append(P("   ·   ".join(str(x) for x in _parts), Ssub))
        E.append(Spacer(1,4*mm))

        # Viability strip
        _vibe = {
            "VIABLE":    f"Plant energy is <b>{P_EPCT:.1f}%</b> of {'combined revenue — well below the 30% caution threshold' if IS_AQ else 'revenue — within viable range'}."
                         + (f" {'Combined' if IS_AQ else 'Farm'} EBITDA is {'positive' if C_EBITDA>=0 else 'negative'} at ${abs(C_EBITDA/1e3):.0f}K." ),
            "MARGINAL":  f"Energy intensity of <b>{P_EPCT:.1f}%</b> of revenue is above the 30% caution threshold. System is marginally viable; energy cost is a key risk.",
            "NOT VIABLE":f"Energy intensity of <b>{P_EPCT:.1f}%</b> of revenue exceeds the 60% non-viability threshold. Structurally unviable at current electricity prices.",
        }[_viab]
        if C_DSCR and C_DSCR < 1.0:
            _vibe += f" DSCR {C_DSCR:.2f}× — EBITDA insufficient to service debt."

        _viab_lhs = Table([
            [P(f'<font color="{_vc_hex}">●</font>  <b><font color="{_vc_hex}">{_viab}</font></b>  <font color="#7a807a">STRUCTURAL VIABILITY SIGNAL</font>',
               ps("vl1",7.5,MONO_B,INK_3,sa=3))],
            [P(_vibe, ps("vl2",8.5,SANS,INK_2,sa=0,lm=1.5))],
        ], colWidths=[BW-32*mm])
        _viab_lhs.setStyle(TableStyle([("LEFTPADDING",(0,0),(-1,-1),0),
                                        ("TOPPADDING",(0,0),(-1,-1),1),
                                        ("BOTTOMPADDING",(0,0),(-1,-1),1)]))
        _stamp = Table([
            [P("ENERGY RATIO", ps("s1",7,MONO_B,INK_3,align=TA_CENTER,sa=1))],
            [P(f"{P_EPCT:.1f}%", ps("s2",14,MONO_B,INK,align=TA_CENTER,sa=0))],
        ], colWidths=[28*mm])
        _stamp.setStyle(TableStyle([("TOPPADDING",(0,0),(-1,-1),2),
                                     ("BOTTOMPADDING",(0,0),(-1,-1),2)]))
        viab = Table([[_viab_lhs, _stamp]], colWidths=[BW-32*mm, 32*mm])
        viab.setStyle(TableStyle([
            ("BOX",(0,0),(-1,-1),0.5,RULE),("BACKGROUND",(0,0),(-1,-1),LINEN_2),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
            ("LEFTPADDING",(0,0),(0,-1),5),("LINEAFTER",(0,0),(0,-1),0.4,RULE),
            ("LEFTPADDING",(1,0),(1,-1),6),
        ]))
        E.append(viab); E.append(Spacer(1,4*mm))

        # KPI grid (3×2)
        def _kc(lbl, val_p, sub="", primary=False):
            bg = SAGE_TINT if primary else LINEN_2
            rows = [[P(lbl.upper(), Skpilbl)],[val_p]]
            if sub: rows.append([P(sub,Skpisub)])
            t = Table(rows, colWidths=[BW/3-1*mm])
            ts2 = [("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
                   ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
                   ("BACKGROUND",(0,0),(-1,-1),bg)]
            if primary: ts2.append(("LINEABOVE",(0,0),(-1,0),2.0,SAGE))
            t.setStyle(TableStyle(ts2)); return t

        def _krow(cells):
            t = Table([cells], colWidths=[BW/3]*3)
            t.setStyle(TableStyle([("BOX",(0,0),(-1,-1),0.5,RULE),
                                    ("INNERGRID",(0,0),(-1,-1),0.5,RULE),
                                    ("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0),
                                    ("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
            return t

        def _vp(v, prefix="$", suf="K", divk=True):
            val = v/1e3 if divk else v
            neg = val < 0
            txt = f"{prefix}{abs(val):,.0f}{suf}"
            if neg: txt = f"−{txt}"
            return P(txt, Skpival_c if neg else (Skpival_s if not neg and v == C_REV else Skpival))

        if IS_AQ:
            r1 = [_kc("Combined Revenue",  P(f"${C_REV/1e3:.0f}<font size='11'>K</font>", Skpival_s),
                       sub=f"Plant ${P_REV/1e3:.0f}K  ·  Fish ${F_REV/1e3:.0f}K", primary=True),
                  _kc("Combined EBITDA",
                       P(f"${abs(C_EBITDA/1e3):.0f}<font size='11'>K</font>",
                         Skpival_s if C_EBITDA>=0 else Skpival_c),
                       sub=f"Margin {C_MARGIN*100:.1f}%"),
                  _kc("Combined CAPEX",
                       P(f"${C_CAPEX/1e3:.0f}<font size='11'>K</font>",Skpival),
                       sub="Plant + fish + shared")]
            _pb_s = f"{C_PBK:.1f}<font size='11'>yr</font>" if C_PBK else "N/A"
            _ds_s = f"{C_DSCR:.2f}<font size='11'>×</font>" if C_DSCR else "N/A"
            r2 = [_kc("Plant Payback",
                       P(_pb_s, Skpival_c if not C_PBK else Skpival),
                       sub="Not reached < 10y" if not C_PBK else "yrs to equity return"),
                  _kc("Combined DSCR",
                       P(_ds_s, Skpival_c if (C_DSCR and C_DSCR<1.0) else Skpival),
                       sub="Coverage below 1.0×" if (C_DSCR and C_DSCR<1.0) else "Debt coverage ratio"),
                  _kc("Fish % Revenue",
                       P(f"{F_PCT_REV:.0f}<font size='11'>%</font>",
                         Skpival if F_PCT_REV>20 else Skpival_c),
                       sub=f"{F_KG:,.0f} kg  @  ${F_PRICE:.2f}/kg")]
        else:
            r1 = [_kc("Annual Revenue",
                       P(f"${P_REV/1e3:.0f}<font size='11'>K</font>",Skpival_s), primary=True),
                  _kc("EBITDA",
                       P(f"${abs(P_EBITDA/1e3):.0f}<font size='11'>K</font>",
                         Skpival_s if P_EBITDA>=0 else Skpival_c),
                       sub=f"Margin {P_MARGIN*100:.1f}%"),
                  _kc("Total CAPEX",
                       P(f"${P_CAPEX/1e3:.0f}<font size='11'>K</font>",Skpival))]
            _pb_s = f"{P_PBK:.1f}<font size='11'>yr</font>" if P_PBK else "N/A"
            _ds_s = f"{P_DSCR:.2f}<font size='11'>×</font>" if P_DSCR else "N/A"
            r2 = [_kc("Payback Period",
                       P(_pb_s, Skpival_c if not P_PBK else Skpival)),
                  _kc("DSCR",
                       P(_ds_s, Skpival_c if (P_DSCR and P_DSCR<1.0) else Skpival),
                       sub="Debt coverage ratio"),
                  _kc("NPV @ Year 10",
                       P(f"${abs(P_NPV/1e3):.0f}<font size='11'>K</font>",
                         Skpival_s if P_NPV>=0 else Skpival_c))]

        kgrid = Table([[_krow(r1)],[_krow(r2)]], colWidths=[BW])
        kgrid.setStyle(TableStyle([("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0),
                                    ("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
        E.append(kgrid); E.append(Spacer(1,5*mm))

        # EBITDA Bridge
        E += _ch("01","Combined EBITDA Bridge" if IS_AQ else "EBITDA Bridge","USD  ·  ANNUAL")
        if IS_AQ:
            _bl = ["Plant Rev","Fish Rev","Plant Costs","Fish Costs","Nutrient ↔","EBITDA"]
            _bv = [P_REV, F_REV, -P_COSTS, -F_COSTS, NUTR_SAV, C_EBITDA]
            _bc = ["#2f5d3a","#3e7448","rgba(184,92,56,0.80)","rgba(184,92,56,0.65)",
                   "#2c5a78","#2f5d3a" if C_EBITDA>=0 else "#b85c38"]
        else:
            _bl = ["Revenue","Energy","Labour","Variable","Water","Rent","Maint.","EBITDA"]
            _bv = [P_REV,-P_ENERGY,-P_LABOUR,-P_VAR,-P_WATER,-P_RENT,-P_MAINT,P_EBITDA]
            _bc = ["#2f5d3a","rgba(184,92,56,0.8)","rgba(184,92,56,0.7)","rgba(184,92,56,0.6)",
                   "rgba(184,92,56,0.5)","rgba(184,92,56,0.5)","rgba(184,92,56,0.6)",
                   "#2f5d3a" if P_EBITDA>=0 else "#b85c38"]
        fig_br = go.Figure(go.Bar(
            x=_bl, y=_bv, marker_color=_bc,
            text=[f"${v/1e3:+.0f}K" for v in _bv], textposition="outside",
        ))
        fig_br.update_layout(yaxis=dict(tickprefix="$",tickformat=",.0f"))
        E.append(_img(fig_br))
        E.append(P(
            f"Revenue and costs shown as annual contributions to combined EBITDA. "
            + (f"Fish EBITDA: ${F_EBITDA/1e3:+.0f}K ({F_MARGIN*100:.1f}% margin). "
               f"Plant EBITDA: ${P_EBITDA/1e3:+.0f}K ({P_MARGIN*100:.1f}% margin). "
               if IS_AQ else "")
            + f"Combined margin: {C_MARGIN*100:.1f}%.", Scap_i))
        return E

    # ══════════════════════════════════════════════════════════════════════════
    # PAGE 2 — Full P&L (Plant + Fish)
    # ══════════════════════════════════════════════════════════════════════════
    def _p2():
        E = []
        E += _sh("II","Cost Structure & Profit / Loss","Annual basis  ·  USD")

        # ── Plant P&L ─────────────────────────────────────────────────────────
        E.append(_subh("Plant Side — Annual P&L"))
        def _pct(v, d): return P(f"{v/d*100:.1f}%", Stnum) if d else _dash()
        pl_hdr = [P("ITEM",Sthlbl), P("$/YEAR",Sthlbl_r),
                  P("% COSTS",Sthlbl_r), P("% REV",Sthlbl_r)]
        pl_rows = [pl_hdr,
            [P("Revenue",Stbdy),       _mk(P_REV,s=Stnum),  _dash(),         P("100%",Stnum)],
            [P("Energy",Stbdy),        _mk(P_ENERGY,s=Stnum),_pct(P_ENERGY,P_COSTS),_pct(P_ENERGY,P_REV)],
            [P("Labour",Stbdy),        _mk(P_LABOUR,s=Stnum),_pct(P_LABOUR,P_COSTS),_pct(P_LABOUR,P_REV)],
            [P("Variable (nutrients/seeds)",Stbdy), _mk(P_VAR,s=Stnum),_pct(P_VAR,P_COSTS),P("—",Stnum_3)],
            [P("Water",Stbdy),         _mk(P_WATER,s=Stnum), _pct(P_WATER,P_COSTS),P("—",Stnum_3)],
            [P("Maintenance",Stbdy),   _mk(P_MAINT,s=Stnum), _pct(P_MAINT,P_COSTS),P("—",Stnum_3)],
            [P("Rent",Stbdy),          _mk(P_RENT,s=Stnum),  _pct(P_RENT,P_COSTS), P("—",Stnum_3)],
            [P("Total Costs",ps("tc",9.5,SANS_B,INK,sa=0)), _mk(P_COSTS,s=Stnum), P("100%",Stnum), _pct(P_COSTS,P_REV)],
            [P("EBITDA",ps("eb",9.5,SANS_B,INK,sa=0)),
             _mk(P_EBITDA,s=Stnum_s if P_EBITDA>=0 else Stnum_c),
             _dash(), P(f"{P_MARGIN*100:.1f}%", Stnum_s if P_EBITDA>=0 else Stnum_c)],
        ]
        # Below-EBITDA items
        pl_rows += [
            [P("— Depreciation",Sbody2), _mk(P_DEPR,s=Stnum), _dash(), _dash()],
            [P("EBIT",ps("ei",9.5,SANS_B,INK,sa=0)), _mk(P_EBIT,s=Stnum_s if P_EBIT>=0 else Stnum_c), _dash(), _dash()],
            [P("— Interest & tax",Sbody2), _mk(P_DS+P_TAX,s=Stnum), _dash(), _dash()],
            [P("Net Income",ps("ni",9.5,SANS_B,INK,sa=0)), _mk(P_NI,s=Stnum_s if P_NI>=0 else Stnum_c), _dash(), _dash()],
        ]
        cw = [BW*0.40, BW*0.20, BW*0.20, BW*0.20]
        E.append(_ftbl(pl_rows, cw, erow=9, totrow=8))
        E.append(Spacer(1,3*mm))

        # Key per-unit metrics inline
        metrics_data = [
            [P("Revenue / m²",Stbdy), P(f"${P_REV_M2:,.0f}/yr",Stnum),
             P("kWh / kg",Stbdy), P(f"{P_KWH_KG:.1f}",Stnum),
             P("Energy / kg",Stbdy), P(f"${P_ENERGY_KG:.2f}",Stnum),
             P("Labour hrs/yr",Stbdy), P(f"{P_LH:,.0f}",Stnum)],
        ]
        mt = Table(metrics_data, colWidths=[BW/8]*8)
        mt.setStyle(TableStyle([
            ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
            ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
            ("BACKGROUND",(0,0),(-1,-1),LINEN_2),
            ("BOX",(0,0),(-1,-1),0.5,RULE),("INNERGRID",(0,0),(-1,-1),0.3,RULE_SOFT),
        ]))
        E.append(mt); E.append(Spacer(1,4*mm))

        # ── Fish P&L (AQ only) ────────────────────────────────────────────────
        if IS_AQ:
            E.append(_subh("Fish Side — Annual P&L"))
            fh = [P("ITEM",Sthlbl), P("$/YEAR",Sthlbl_r), P("NOTES",Sthlbl)]
            fr_rows = [fh,
                [P("Revenue",Stbdy), _mk(F_REV,s=Stnum),
                 P(f"{F_KG:,.0f} kg  @  ${F_PRICE:.2f}/kg", Stnote)],
                [P("Feed",Stbdy), _mk(F_FEED,s=Stnum),
                 P(f"FCR {_fr.get('fcr',1.5):.1f}  ·  {F_KG*_fr.get('fcr',1.5):,.0f} kg feed/yr", Stnote)],
                [P("Fingerlings",Stbdy), _mk(F_FING,s=Stnum), P("—",Stnote)],
                [P("Energy (heating+aeration)",Stbdy), _mk(F_ENERGY,s=Stnum),
                 P(f"ΔT={F_DELTA_T:.0f}°C  ·  heating {F_HHEAT:,.0f} kWh  ·  aeration {F_HAER:,.0f} kWh", Stnote)],
                [P("Water & other",Stbdy), _mk(F_WATER,s=Stnum), P("—",Stnote)],
                [P("Labour",Stbdy), _mk(F_LABOUR,s=Stnum),
                 P(f"{_fr.get('annual_fish_labour_hours',0):,.0f} hrs/yr", Stnote)],
                [P("Maintenance",Stbdy), _mk(F_MAINT,s=Stnum), P("—",Stnote)],
                [P("Total Fish Costs",ps("ftc",9.5,SANS_B,INK,sa=0)), _mk(F_COSTS,s=Stnum), P("—",Stnote)],
                [P("Fish EBITDA",ps("feb",9.5,SANS_B,INK,sa=0)),
                 _mk(F_EBITDA,s=Stnum_s if F_EBITDA>=0 else Stnum_c),
                 P(f"Margin {F_MARGIN*100:.1f}%", Stnote)],
            ]
            cw2 = [BW*0.36, BW*0.19, BW*0.45]
            E.append(_ftbl(fr_rows, cw2, erow=9, totrow=8))
            E.append(Spacer(1,3*mm))
            if NUTR_SAV > 0:
                E.append(P(f"Nutrient offset saving (fish effluent to plant side): ${NUTR_SAV:,.0f}/yr — "
                           f"reduces plant variable costs by {NUTR_SAV/P_VAR*100:.1f}% where applicable.", Scap))
            E.append(Spacer(1,2*mm))

        E.append(_mbox("READING THIS SECTION",
            "All P&L figures are annual and in USD. EBITDA excludes depreciation, interest, and tax.",
            "Break-even price for plant side: "
            + (f"${P_BE_PRICE:.2f}/kg (current: ${P_PRICE:.2f}/kg — "
               f"{'headroom' if P_HDROOM and P_HDROOM>0 else 'deficit'}: "
               f"{abs(P_HDROOM):.0f}%)." if P_BE_PRICE else "N/A."),
            "Net income is after depreciation, interest on debt, and tax at the specified rate."
        ))
        return E

    # ══════════════════════════════════════════════════════════════════════════
    # PAGE 3 — CAPEX & Funding Structure
    # ══════════════════════════════════════════════════════════════════════════
    def _p3():
        E = []
        E += _sh("III","Capital Expenditure & Funding","One-time investment  ·  USD")

        # ── CAPEX breakdown ───────────────────────────────────────────────────
        E.append(_subh("Plant CAPEX Breakdown"))
        if "led_capex" in _pr:
            _plant_capex_items = [
                ("LED Lighting",        _pr.get("led_capex",0)),
                ("HVAC",                _pr.get("hvac_capex",0)),
                ("Racking & Structures",_pr.get("racks_capex",0)),
                ("Building & Enclosure",_pr.get("building_capex",0)),
                ("Automation Controls", _pr.get("automation_capex",0)),
                ("Robotics",            _pr.get("robotics_capex",0)),
                ("Electrical",          _pr.get("electrical_capex",0)),
                ("Water & Irrigation",  _pr.get("water_capex",0)),
                ("Installation",        _pr.get("installation_capex",0)),
            ]
        else:
            _plant_capex_items = [
                ("Structure & Cladding",_pr.get("structure_capex",0)),
                ("Climate Control",     _pr.get("climate_capex",0)),
                ("Irrigation & Hydro",  _pr.get("irrigation_capex",0)),
                ("Supplemental Lighting",_pr.get("lighting_capex",0)),
                ("Automation",          _pr.get("automation_capex",0)),
                ("Real Estate",         _pr.get("real_estate_capex",0)),
            ]
        _plant_capex_items = [(k,v) for k,v in _plant_capex_items if v>0]

        capex_hdr = [P("COMPONENT",Sthlbl), P("$ AMOUNT",Sthlbl_r), P("% OF TOTAL",Sthlbl_r), P("$/M² GROW",Sthlbl_r)]
        c_rows = [capex_hdr]
        for k,v in _plant_capex_items:
            c_rows.append([P(k,Stbdy), _mk(v,s=Stnum),
                           P(f"{v/P_CAPEX*100:.1f}%" if P_CAPEX else "—",Stnum),
                           P(f"${v/P_EGA:.0f}" if P_EGA else "—",Stnum)])
        c_rows.append([P("TOTAL PLANT CAPEX",ps("tpc",9.5,SANS_B,INK,sa=0)),
                        _mk(P_CAPEX,s=Stnum),
                        P("100%",Stnum),
                        P(f"${P_CAPEX/P_EGA:.0f}" if P_EGA else "—",Stnum)])
        cw3 = [BW*0.40, BW*0.20, BW*0.20, BW*0.20]
        E.append(_ftbl(c_rows, cw3, totrow=len(c_rows)-1))

        if IS_AQ:
            E.append(Spacer(1,4*mm))
            E.append(_subh("Fish System CAPEX"))
            fish_capex_items = [
                ("Fish Tanks",          F_TKAP),
                ("Filtration (RAS)",    F_FKAP),
                ("Aeration",            F_AKAP),
                ("Monitoring & Control",F_MKAP),
                ("Plumbing & Pipework", F_PKAP),
            ]
            fish_capex_items = [(k,v) for k,v in fish_capex_items if v>0]
            fc_rows = [capex_hdr]
            for k,v in fish_capex_items:
                fc_rows.append([P(k,Stbdy), _mk(v,s=Stnum),
                                P(f"{v/F_CAPEX*100:.1f}%" if F_CAPEX else "—",Stnum),
                                P(f"${v/F_TVOL:.0f}/m³" if F_TVOL else "—",Stnum)])
            fc_rows.append([P("TOTAL FISH CAPEX",ps("tfc",9.5,SANS_B,INK,sa=0)),
                             _mk(F_CAPEX,s=Stnum), P("100%",Stnum),
                             P(f"${F_CAPEX/F_TVOL:.0f}/m³" if F_TVOL else "—",Stnum)])
            if INT_CAP > 0:
                fc_rows.append([P("Integration (shared infra)",Stbdy),
                                 _mk(INT_CAP,s=Stnum), _dash(), _dash()])
            E.append(_ftbl(fc_rows, cw3, totrow=len(fc_rows)-1))

            E.append(Spacer(1,3*mm))
            comb_capex_tbl = Table([
                [P("COMBINED TOTAL CAPEX",ps("ctc",10,SANS_B,INK,sa=0)),
                 P(f"${C_CAPEX:,.0f}",ps("ctv",10,MONO_B,INK,align=TA_RIGHT,sa=0)),
                 P(f"${C_CAPEX/P_EGA:.0f}/m² grow" if P_EGA else "—",
                   ps("ctu",9,MONO,INK_2,align=TA_RIGHT,sa=0))]
            ], colWidths=[BW*0.45, BW*0.28, BW*0.27])
            comb_capex_tbl.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,-1),SAGE_TINT),
                ("LINEABOVE",(0,0),(-1,0),1.0,SAGE),
                ("LINEBELOW",(0,0),(-1,-1),0.5,RULE),
                ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
                ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
            ]))
            E.append(comb_capex_tbl)

        E.append(Spacer(1,5*mm))

        # ── Funding structure ─────────────────────────────────────────────────
        E.append(_subh("Funding Structure"))
        _ltv    = inputs_dict.get("ltv",0)
        _ir     = inputs_dict.get("interest_rate",0)
        _lt     = inputs_dict.get("loan_term_years",0)
        _eq     = C_CAPEX*(1-_ltv/100)
        _dbt    = C_CAPEX*_ltv/100
        fund_rows = [
            [P("ITEM",Sthlbl), P("AMOUNT",Sthlbl_r), P("SHARE",Sthlbl_r), P("NOTES",Sthlbl)],
            [P("Equity (own funds)",Stbdy), _mk(_eq,s=Stnum),
             P(f"{100-_ltv:.0f}%",Stnum), P("Investor / owner equity",Stnote)],
            [P("Debt (bank loan)",Stbdy), _mk(_dbt,s=Stnum),
             P(f"{_ltv:.0f}%",Stnum), P(f"{_ir:.1f}% interest  ·  {_lt} yr term",Stnote)],
            [P("TOTAL CAPEX",ps("tcf",9.5,SANS_B,INK,sa=0)),
             _mk(C_CAPEX,s=Stnum), P("100%",Stnum), P("—",Stnote)],
        ]
        fund_rows2 = [
            [P("Annual Debt Service",Stbdy), _mk(P_DS,s=Stnum), _dash(),
             P("Principal + interest",Stnote)],
            [P("DSCR (Plant)",Stbdy),
             P(f"{P_DSCR:.2f}×" if P_DSCR else "N/A",
               Stnum_c if (P_DSCR and P_DSCR<1.0) else Stnum),
             _dash(), P("EBITDA ÷ debt service (>1.25× preferred)",Stnote)],
        ]
        if IS_AQ:
            fund_rows2 += [
                [P("DSCR (Fish)",Stbdy),
                 P(f"{F_DSCR:.2f}×" if F_DSCR else "N/A",
                   Stnum_c if (F_DSCR and F_DSCR<1.0) else Stnum),
                 _dash(), P("Fish EBITDA ÷ fish debt service",Stnote)],
                [P("DSCR (Combined)",Stbdy),
                 P(f"{C_DSCR:.2f}×" if C_DSCR else "N/A",
                   Stnum_c if (C_DSCR and C_DSCR<1.0) else Stnum),
                 _dash(), P("Combined EBITDA ÷ total debt service",Stnote)],
            ]
        all_fund = fund_rows + fund_rows2[1:]
        cw4 = [BW*0.33, BW*0.20, BW*0.14, BW*0.33]
        E.append(_ftbl(all_fund, cw4, totrow=3))

        if C_DSCR and C_DSCR < 1.0:
            E.append(Spacer(1,2*mm))
            E.append(P(f"⚠  DSCR of {C_DSCR:.2f}× indicates combined EBITDA is insufficient to service total debt. "
                       "Consider reducing LTV, extending loan term, or improving margins before debt financing.",
                       ps("dw",8.5,SANS_B,CLAY,sa=3)))

        E.append(Spacer(1,4*mm))
        E.append(_mbox("CAPEX NOTES",
            f"Plant CAPEX: ${P_CAPEX:,.0f} (${P_CAPEX/P_EGA:.0f}/m² effective grow area). "
            + (f"Fish CAPEX: ${F_CAPEX:,.0f} (${F_CAPEX/F_TVOL:.0f}/m³ tank volume). " if IS_AQ else "")
            + "CAPEX excludes working capital and pre-opening costs.",
            "Funding structure assumes a single tranche senior debt facility. "
            "Mezzanine or grant financing not modelled."
        ))
        return E

    # ══════════════════════════════════════════════════════════════════════════
    # PAGE 4 — Break-Even & Scenario Analysis
    # ══════════════════════════════════════════════════════════════════════════
    def _p4():
        E = []
        E += _sh("IV","Break-Even & Scenario Analysis","Price sensitivity  ·  USD")

        # ── Break-even table ──────────────────────────────────────────────────
        E.append(_subh("Plant Side — Break-Even Analysis"))
        be_rows = [
            [P("METRIC",Sthlbl), P("VALUE",Sthlbl_r), P("NOTES",Sthlbl)],
            [P("Annual production",Stbdy), P(f"{P_KG:,.0f} kg",Stnum),
             P(f"{P_KG_M2:.1f} kg/m² eff. grow · {P_CY} cycles/yr · {P_ECD:.0f} days/cycle",Stnote)],
            [P("Current realised price",Stbdy), P(f"${P_PRICE:.2f}/kg",Stnum),
             P(inputs_dict.get("price_scenario","base").title()+" price scenario",Stnote)],
            [P("Break-even price",Stbdy),
             P(f"${P_BE_PRICE:.2f}/kg" if P_BE_PRICE else "N/A",
               Stnum_c if (P_BE_PRICE and P_BE_PRICE>P_PRICE) else Stnum),
             P("Total annual costs ÷ annual production",Stnote)],
            [P("Price headroom / gap",Stbdy),
             P((f"+{P_HDROOM:.0f}%" if P_HDROOM and P_HDROOM>0 else f"−{abs(P_HDROOM):.0f}%") if P_HDROOM else "N/A",
               Stnum_s if (P_HDROOM and P_HDROOM>0) else Stnum_c),
             P("Above break-even = viable at current price",Stnote)],
            [P("Break-even yield",Stbdy),
             P(f"{P_BE_YIELD:.2f} kg/m²/cycle" if P_BE_YIELD else "N/A", Stnum),
             P("Min yield to cover all costs at current price",Stnote)],
            [P("Revenue / m² (eff. grow)",Stbdy), P(f"${P_REV_M2:,.0f}/yr",Stnum),
             P("Effective growing area basis",Stnote)],
            [P("Energy cost / kg",Stbdy), P(f"${P_ENERGY_KG:.2f}",Stnum),
             P(f"{P_KWH_KG:.1f} kWh/kg  ·  ${inputs_dict.get('country','—')} energy rate",Stnote)],
            [P("Labour cost / kg",Stbdy), P(f"${P_LABOUR_KG:.2f}",Stnum),
             P(f"{P_LH:,.0f} hrs/yr total labour",Stnote)],
            [P("EBITDA / m² (eff. grow)",Stbdy), P(f"${P_EBITDA/P_EGA:,.0f}/yr" if P_EGA else "N/A",
               Stnum_s if P_EBITDA>=0 else Stnum_c), P("EBITDA density",Stnote)],
        ]
        if IS_AQ:
            be_rows += [
                [P("Fish break-even price",Stbdy),
                 P(f"${F_COSTS/F_KG:.2f}/kg" if F_KG else "N/A",
                   Stnum_c if F_KG and F_COSTS/F_KG>F_PRICE else Stnum),
                 P(f"Current fish price ${F_PRICE:.2f}/kg",Stnote)],
                [P("Nutrient offset saving",Stbdy), P(f"${NUTR_SAV:,.0f}",Stnum),
                 P("Fertiliser cost saving from fish effluent",Stnote)],
            ]
        cw5 = [BW*0.34, BW*0.20, BW*0.46]
        E.append(_ftbl(be_rows, cw5, highlight_rows_dict={3:CLAY_TINT if (P_BE_PRICE and P_BE_PRICE>P_PRICE) else SAGE_TINT,4:None}))
        E.append(Spacer(1,4*mm))

        # ── Scenarios ─────────────────────────────────────────────────────────
        if _scen_names and _scen_ebitda:
            E += _ch("02","Plant Price Scenario Comparison","Low / Base / High")
            fig_sc = go.Figure()
            _sc_col = ["#b85c38" if v<0 else ("#c08a2e" if v<P_REV*0.05 else "#2f5d3a") for v in _scen_ebitda]
            fig_sc.add_trace(go.Bar(name="Revenue", x=_scen_names, y=_scen_rev,
                marker_color=["rgba(47,93,58,0.25)"]*3,
                text=[f"${v/1e3:.0f}K" for v in _scen_rev], textposition="outside"))
            fig_sc.add_trace(go.Bar(name="EBITDA", x=_scen_names, y=_scen_ebitda,
                marker_color=_sc_col,
                text=[f"${v/1e3:.0f}K" for v in _scen_ebitda], textposition="outside"))
            fig_sc.update_layout(barmode="group", showlegend=True,
                legend=dict(orientation="h",yanchor="bottom",y=1.02,xanchor="right",x=1),
                yaxis=dict(tickprefix="$",tickformat=",.0f"))
            E.append(_img(fig_sc, h=60*mm, ph=480))
            sc_tbl_rows = [
                [P("SCENARIO",Sthlbl), P("REVENUE",Sthlbl_r), P("EBITDA",Sthlbl_r),
                 P("MARGIN",Sthlbl_r), P("vs BASE",Sthlbl_r)],
            ]
            for i,(sn,se,sr,sm) in enumerate(zip(_scen_names,_scen_ebitda,_scen_rev,_scen_margin)):
                delta = se-P_EBITDA
                sc_tbl_rows.append([
                    P(sn,Stbdy), _mk(sr,s=Stnum), _mk(se,s=Stnum_s if se>=0 else Stnum_c),
                    P(f"{sm:.1f}%",Stnum),
                    P("— (base)" if i==1 else (f"+${delta/1e3:.0f}K" if delta>=0 else f"−${abs(delta/1e3):.0f}K"),
                      Stnum_s if delta>=0 else Stnum_c)])
            E.append(_ftbl(sc_tbl_rows, [BW*0.26,BW*0.18,BW*0.18,BW*0.18,BW*0.20],
                           highlight_rows_dict={2:SAGE_TINT}))
            E.append(P("Low/High scenarios stress the plant selling price ±20% with all other inputs held constant.",Scap))
        else:
            E.append(P("Scenario analysis not available (sensitivity helper not provided).", Scap))

        E.append(Spacer(1,4*mm))
        E.append(_mbox("BREAK-EVEN INTERPRETATION",
            f"Break-even price of ${P_BE_PRICE:.2f}/kg represents the minimum realised price to cover all operating costs. "
            + (f"Current scenario (${P_PRICE:.2f}/kg) {'exceeds' if P_HDROOM and P_HDROOM>0 else 'is below'} this threshold by {abs(P_HDROOM):.0f}%." if P_HDROOM else ""),
            "Break-even yield represents the minimum crop output per m² per cycle at the current price. "
            "Both metrics are pre-tax and pre-debt-service."
        ))
        return E

    # ══════════════════════════════════════════════════════════════════════════
    # PAGE 5 — Sensitivity (Tornado) + Investment Returns (DCF)
    # ══════════════════════════════════════════════════════════════════════════
    def _p5():
        E = []
        E += _sh("V","Sensitivity & Investment Returns","EBITDA tornado  ·  10-yr DCF")

        # ── Tornado ───────────────────────────────────────────────────────────
        if _tvars:
            E += _ch("03","EBITDA Sensitivity — Key Drivers","Single-variable stress  ·  base = ${:.0f}K".format(P_EBITDA/1e3))
            fig_t = go.Figure()
            for tv in _tvars:
                fig_t.add_trace(go.Bar(
                    name="Pessimistic", y=[tv["label"]], x=[tv["dp"]],
                    orientation="h", marker_color="rgba(184,92,56,0.80)",
                    showlegend=(tv==_tvars[0]),
                    text=f"${tv['dp']/1e3:+.0f}K", textposition="outside"))
                fig_t.add_trace(go.Bar(
                    name="Optimistic", y=[tv["label"]], x=[tv["do"]],
                    orientation="h", marker_color="rgba(47,93,58,0.80)",
                    showlegend=(tv==_tvars[0]),
                    text=f"${tv['do']/1e3:+.0f}K", textposition="outside"))
            fig_t.add_vline(x=0, line_dash="solid", line_color="#161a16", line_width=0.8)
            fig_t.update_layout(barmode="overlay", showlegend=True,
                legend=dict(orientation="h",yanchor="bottom",y=1.02,xanchor="right",x=1),
                xaxis=dict(title="EBITDA delta ($)", tickprefix="$", tickformat=",.0f"),
                yaxis=dict(showgrid=False),
                margin=dict(l=90,r=80,t=8,b=36))
            E.append(_img(fig_t, h=58*mm, ph=440))
            E.append(Spacer(1,2*mm))
            torn_rows = [[P("DRIVER",Sthlbl), P("PESS. SCENARIO",Sthlbl),
                          P("EBITDA Δ",Sthlbl_r), P("OPT. SCENARIO",Sthlbl),
                          P("EBITDA Δ",Sthlbl_r)]]
            for tv in _tvars:
                torn_rows.append([
                    P(tv["label"],Stbdy),
                    P(tv["pl"],Stnote),
                    _mk(tv["dp"],s=Stnum_c if tv["dp"]<0 else Stnum_s),
                    P(tv["ol"],Stnote),
                    _mk(tv["do"],s=Stnum_s if tv["do"]>0 else Stnum_c),
                ])
            E.append(_ftbl(torn_rows, [BW*0.17,BW*0.22,BW*0.14,BW*0.22,BW*0.25]))
            E.append(P("Each variable stressed independently. All others held at base. "
                       "Sorted by total EBITDA swing (largest first).", Scap))
        else:
            E.append(P("Tornado chart not available (sensitivity helper not provided).", Scap))

        E.append(Spacer(1,5*mm))

        # ── DCF ───────────────────────────────────────────────────────────────
        E += _ch("04","10-Year DCF — Plant Side","USD  ·  Equity FCFE  ·  Discounted")
        _end_v = P_DCF[-1]["cumulative_npv"] if P_DCF else 0
        _lc = "#2f5d3a" if _end_v >= 0 else "#b85c38"
        _fc = "rgba(47,93,58,0.12)" if _end_v >= 0 else "rgba(184,92,56,0.12)"
        fig_d = go.Figure()
        fig_d.add_trace(go.Scatter(
            x=["Y0"] + [f"Y{d['year']}" for d in P_DCF],
            y=[-P_EQ] + [d["cumulative_npv"] for d in P_DCF],
            mode="lines+markers",
            line=dict(color=_lc, width=2),
            marker=dict(size=4, color=_lc),
            fill="tozeroy", fillcolor=_fc,
        ))
        fig_d.add_hline(y=0, line_color="#161a16", line_width=0.8)
        if P_DCF:
            fig_d.add_annotation(x=f"Y{len(P_DCF)}", y=_end_v,
                text=f"  ${_end_v/1e3:.0f}K @ Y{len(P_DCF)}",
                showarrow=False, font=dict(family="Courier",size=9,color=_lc), xanchor="left")
        fig_d.update_layout(
            xaxis=dict(title=None),
            yaxis=dict(tickprefix="$", tickformat=",.0f"))
        E.append(_img(fig_d))

        _max_abs = max((abs(d["cumulative_npv"]) for d in P_DCF), default=1)
        dcf_rows = [[P("YR",Sthlbl), P("FCFE ($)",Sthlbl_r), P("PV ($)",Sthlbl_r),
                     P("CUM. NPV ($)",Sthlbl_r), P("TRAJECTORY",Sthlbl)]]
        for d in P_DCF:
            cv = d["cumulative_npv"]
            dcf_rows.append([
                P(f"Y{d['year']}",Stbdy),
                _mk(d["fcfe"],s=Stnum_c if d["fcfe"]<0 else Stnum),
                _mk(d["pv"],   s=Stnum_c if d["pv"]<0 else Stnum),
                _mk(cv,        s=Stnum_c if cv<0 else Stnum),
                _spark(cv, _max_abs),
            ])
        E.append(_ftbl(dcf_rows, [8*mm, BW*0.19, BW*0.19, BW*0.22, 34*mm]))
        E.append(P(
            f"Discounted at {inputs_dict.get('discount_rate',8):.1f}% (equity cost). "
            f"Year-0 equity outlay: −${P_EQ:,.0f} (CAPEX × (1 − LTV)). "
            + ("Plant NPV turns positive before Y10." if _end_v>=0 else
               f"Plant side does not recover equity within 10 years at current assumptions. "
               f"Final NPV: ${_end_v:,.0f}."),
            Scap_i))

        if IS_AQ and F_DCF:
            E.append(Spacer(1,4*mm))
            E += _ch("05","10-Year DCF — Fish Side","USD  ·  Equity FCFE  ·  Discounted")
            _fe = F_DCF[-1]["cumulative_npv"] if F_DCF else 0
            _flc = "#2f5d3a" if _fe>=0 else "#b85c38"
            fig_fd = go.Figure()
            fig_fd.add_trace(go.Scatter(
                x=["Y0"] + [f"Y{d['year']}" for d in F_DCF],
                y=[-_fr.get("equity_invested",0)] + [d["cumulative_npv"] for d in F_DCF],
                mode="lines+markers",
                line=dict(color=_flc,width=2),
                marker=dict(size=4,color=_flc),
                fill="tozeroy", fillcolor=("rgba(47,93,58,0.12)" if _fe>=0 else "rgba(184,92,56,0.12)"),
            ))
            fig_fd.add_hline(y=0, line_color="#161a16", line_width=0.8)
            fig_fd.update_layout(yaxis=dict(tickprefix="$",tickformat=",.0f"))
            E.append(_img(fig_fd, h=52*mm, ph=400))
            E.append(P(f"Fish side DCF (equity basis). "
                       f"{'Fish NPV positive within horizon.' if _fe>=0 else 'Fish side does not recover equity within 10 years.'} "
                       f"Final NPV: ${_fe:,.0f}.",Scap_i))

        return E

    # ══════════════════════════════════════════════════════════════════════════
    # PAGE 6 — System Configuration, Risk Matrix, Methodology
    # ══════════════════════════════════════════════════════════════════════════
    def _p6():
        E = []
        E += _sh("VI","Configuration, Risk & Methodology",f"As modelled  ·  {_rd}")

        CW = BW/2 - 3*mm
        def _row(k,v,bold=False):
            return [P(k, ps("ck2",9,SANS,INK_2,sa=0)),
                    P(str(v), ps("cv2",9.5,MONO_B if bold else MONO,INK,align=TA_RIGHT,sa=0))]
        def _grph(lbl):
            return [P(lbl.upper(), ps("gh",7.5,SANS_B,INK,sa=0)),
                    P("",ps("gh2",7.5,SANS,INK_2,sa=0))]
        def _cfmt(rows):
            t = Table(rows, colWidths=[CW*0.58, CW*0.42])
            t.setStyle(TableStyle([
                ("TOPPADDING",(0,0),(-1,-1),2.5),("BOTTOMPADDING",(0,0),(-1,-1),2.5),
                ("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0),
                ("LINEBELOW",(0,0),(-1,-1),0.35,RULE_SOFT),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ]))
            return t

        _fp = inputs_dict.get("footprint") or inputs_dict.get("plant_footprint",0)
        _lvl= inputs_dict.get("levels",1)
        left_r = (
            [_grph("System & Site")] +
            [_row("Country",            inputs_dict.get("country","—")),
             _row("Modality",           ML),
             _row("Footprint",          f"{int(_fp):,} m²"),
             _row("Levels",             str(int(_lvl)) if _lvl else "1"),
             _row("Eff. grow area",     f"{P_EGA:,.0f} m²"),
             _row("Net grow factor",    f"{inputs_dict.get('net_grow_factor',85):.0f}%"),
             _row("Walkways factor",    f"{inputs_dict.get('walkways_factor',85):.0f}%"),
             _row("Automation",         inputs_dict.get("automation","—")),
             _row("Harvest mode",       inputs_dict.get("harvest_mode","—")),
             ("lights_tier" in inputs_dict and _row("Lights tier", inputs_dict["lights_tier"])
              or _row("Structure", _pr.get("structure_type","—"))),
            ] +
            [_grph("Crop & Pricing")] +
            [_row("Crop",               inputs_dict.get("crop") or inputs_dict.get("plant_crop","—")),
             _row("Price scenario",     inputs_dict.get("price_scenario","—")),
             _row("Selling price",      f"${P_PRICE:.2f}/kg"),
             _row("Cycles / yr",        str(P_CY)),
             _row("Cycle days",         str(int(P_ECD))),
             _row("Annual output",      f"{P_KG:,.0f} kg"),
             _row("Loss rate",          f"{inputs_dict.get('loss_rate',5):.1f}%"),
             _row("Packaging cost",     f"${inputs_dict.get('packaging_cost',0.25):.2f}/kg"),
             _row("Water price",        f"${inputs_dict.get('water_price',0):.2f}/m³"),
             _row("Rent / month",       f"${inputs_dict.get('rent_monthly',0):,.0f}"),
            ]
        )
        if IS_AQ:
            left_r += (
                [_grph("Fish System")] +
                [_row("Species",        F_SPECIES),
                 _row("Tank volume",    f"{F_TVOL:.0f} m³"),
                 _row("System scale",   F_SCALE),
                 _row("Annual output",  f"{F_KG:,.0f} kg/yr"),
                 _row("Cycles / yr",    str(F_CY)),
                 _row("Selling price",  f"${F_PRICE:.2f}/kg"),
                 _row("Target temp",    f"{inputs_dict.get('target_temp_c','—')}°C"),
                 _row("ΔT (heat req.)", f"{F_DELTA_T:.0f}°C"),
                ]
            )

        right_r = (
            [_grph("Financial Structure")] +
            [_row("LTV",               f"{inputs_dict.get('ltv',0):.0f}%"),
             _row("Interest rate",     f"{inputs_dict.get('interest_rate',0):.1f}%"),
             _row("Loan term",         f"{inputs_dict.get('loan_term_years',0)} yrs"),
             _row("Discount rate",     f"{inputs_dict.get('discount_rate',0):.1f}%"),
             _row("Depreciation",      f"{inputs_dict.get('depreciation_years',0)} yrs"),
             _row("Tax rate",          f"{inputs_dict.get('tax_rate',0):.1f}%"),
             _row("Equity invested",   f"${P_EQ:,.0f}"),
             _row("Debt",              f"${P_DEBT:,.0f}"),
             _row("Annual debt svc",   f"${P_DS:,.0f}" if P_DS else "N/A"),
             _row("DSCR (plant)",      f"{P_DSCR:.2f}×" if P_DSCR else "N/A"),
            ] +
            [_grph("Key Results")] +
            [_row("Annual revenue",    f"${P_REV:,.0f}"),
             _row("EBITDA",            f"${P_EBITDA:,.0f} ({P_MARGIN*100:.1f}%)"),
             _row("EBITDA / m²",       f"${P_EBITDA/P_EGA:,.0f}/yr" if P_EGA else "—"),
             _row("Revenue / m²",      f"${P_REV_M2:,.0f}/yr"),
             _row("kWh / kg",          f"{P_KWH_KG:.1f}"),
             _row("Energy / kg",       f"${P_ENERGY_KG:.2f}"),
             _row("Labour cost / kg",  f"${P_LABOUR_KG:.2f}"),
             _row("Break-even price",  f"${P_BE_PRICE:.2f}/kg" if P_BE_PRICE else "N/A"),
             _row("NPV (plant)",       f"${P_NPV:,.0f}"),
             _row("Payback",           f"{P_PBK:.1f} yrs" if P_PBK else "Not reached"),
            ]
        )
        if IS_AQ:
            right_r += (
                [_grph("Fish Results")] +
                [_row("Fish revenue",  f"${F_REV:,.0f}"),
                 _row("Fish EBITDA",   f"${F_EBITDA:,.0f} ({F_MARGIN*100:.1f}%)"),
                 _row("Fish NPV",      f"${F_NPV:,.0f}"),
                 _row("Combined rev",  f"${C_REV:,.0f}"),
                 _row("Combined EBITDA",f"${C_EBITDA:,.0f} ({C_MARGIN*100:.1f}%)"),
                ]
            )

        two_col = Table([[_cfmt(left_r), Spacer(6*mm,1), _cfmt(right_r)]],
                        colWidths=[CW, 6*mm, CW])
        two_col.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),
                                      ("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0),
                                      ("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
        E.append(two_col)
        E.append(Spacer(1,5*mm))

        # ── Risk matrix ───────────────────────────────────────────────────────
        E.append(_subh("Key Risk Factors"))
        _risks = []
        if C_DSCR and C_DSCR < 1.0:
            _risks.append(("HIGH","Debt Coverage",
                f"DSCR {C_DSCR:.2f}× — combined EBITDA cannot service debt at {inputs_dict.get('ltv',0):.0f}% LTV. "
                "Reduce leverage or improve margins before debt financing."))
        elif C_DSCR and C_DSCR < 1.25:
            _risks.append(("MED","Debt Coverage",f"DSCR {C_DSCR:.2f}× below standard 1.25× lender minimum."))
        if P_EPCT > 60:
            _risks.append(("HIGH","Energy Viability",
                f"Energy = {P_EPCT:.1f}% of plant revenue. Structurally unviable at current electricity tariff."))
        elif P_EPCT > 30:
            _risks.append(("MED","Energy Exposure",
                f"Energy = {P_EPCT:.1f}% of plant revenue. High sensitivity to electricity price volatility."))
        if P_HDROOM is not None and P_HDROOM < 10:
            _risks.append(("MED","Price Sensitivity",
                f"Only {abs(P_HDROOM):.0f}% headroom above plant break-even. "
                "Small market price drop flips EBITDA negative."))
        if IS_AQ and F_EBITDA < 0:
            _risks.append(("MED","Fish Operating Loss",
                f"Fish side EBITDA is −${abs(F_EBITDA/1e3):.0f}K. "
                f"Energy (${F_ENERGY/1e3:.0f}K) + feed (${F_FEED/1e3:.0f}K) exceed fish revenue. "
                "Review species, tank sizing, or energy source."))
        if IS_AQ and _fr.get("salmon_warning"):
            _risks.append(("HIGH","Species Thermal Risk",
                "Salmon warning: ambient temperature incompatible with target water temperature without "
                "prohibitive heating cost. Consider tropical species for this climate."))
        _risks += [
            ("LOW","Crop Biology","Controlled environment reduces but does not eliminate pest/disease risk."),
            ("LOW","Market Volatility","Wholesale fresh produce prices can deviate from scenario assumptions."),
            ("LOW","Labour Availability","Skilled CEA operators may be scarce in some markets."),
            ("LOW","Regulatory","Zoning, water rights, and discharge permits not modelled."),
        ]
        rcols = [BW*0.10, BW*0.22, BW*0.68]
        risk_rows = [[P("SEV.",Sthlbl), P("RISK FACTOR",Sthlbl), P("DESCRIPTION",Sthlbl)]]
        for sev, rf, desc in _risks:
            sc = CLAY if sev=="HIGH" else AMBER_C if sev=="MED" else INK_4
            risk_rows.append([
                P(f"<b>{sev}</b>", ps("rs",8,SANS_B,sc,sa=0)),
                P(rf, Stbdy),
                P(desc, Stnote),
            ])
        rt = Table(risk_rows, colWidths=rcols, repeatRows=1)
        rts = [
            ("FONTNAME",(0,0),(-1,0),SANS_B),("FONTSIZE",(0,0),(-1,-1),8.5),
            ("BACKGROUND",(0,0),(-1,0),LINEN_2),("TEXTCOLOR",(0,0),(-1,0),INK_2),
            ("LINEBELOW",(0,0),(-1,0),0.5,INK_3),
            ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
            ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
            ("VALIGN",(0,0),(-1,-1),"TOP"),("FONTNAME",(0,1),(-1,-1),SANS),
            ("TEXTCOLOR",(0,1),(-1,-1),INK),
        ]
        for i,(s,_,_) in enumerate(_risks,1):
            if i%2==1: rts.append(("BACKGROUND",(1,i),(-1,i),LINEN_2))
        rt.setStyle(TableStyle(rts))
        E.append(rt)
        E.append(Spacer(1,4*mm))

        E.append(_mbox("METHODOLOGY  ·  SOURCES  ·  LIMITATIONS",
            "Yield and cost benchmarks from the Agricultural Intelligence Portal crop and fish databases, "
            "calibrated against published CEA and aquaculture operational data. Country-specific kWh and "
            f"labour rates applied ({inputs_dict.get('country','—')} used here).",
            ("In decoupled aquaponics the plant and fish sub-systems are financially independent. "
             "Nutrient offset represents fertiliser savings credited to the plant side from fish effluent. "
             "Combined CAPEX includes integration infrastructure."
             if IS_AQ else
             f"This is a {ML} model. Energy demand scales with DLI requirements and HVAC load. "
             "For greenhouse/polytunnel, natural DLI fraction reduces supplemental lighting cost."),
            f"DCF is equity FCFE discounted at {inputs_dict.get('discount_rate',8):.1f}%. "
            "Results are indicative. Not investment advice. "
            "Validate with site-specific engineering and market studies before any investment decision."
        ))
        return E

    # ── Helper used in _ftbl ──────────────────────────────────────────────────
    # Patch _ftbl to accept highlight_rows_dict kwarg (optional)
    _orig_ftbl = _ftbl
    def _ftbl(data, cw, erow=None, totrow=None, highlight_rows_dict=None):
        t = Table(data, colWidths=cw, repeatRows=1)
        ts = [
            ("FONTNAME",(0,0),(-1,0),SANS_B),("FONTSIZE",(0,0),(-1,-1),8.5),
            ("BACKGROUND",(0,0),(-1,0),LINEN_2),("TEXTCOLOR",(0,0),(-1,0),INK_2),
            ("LINEBELOW",(0,0),(-1,0),0.5,INK_3),
            ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
            ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ]
        for i in range(1,len(data)):
            ts.append(("LINEBELOW",(0,i),(-1,i),0.3,RULE_SOFT))
        if erow is not None:
            ts += [("LINEABOVE",(0,erow),(-1,erow),0.5,INK),
                   ("LINEBELOW",(0,erow),(-1,erow),0.5,INK),
                   ("BACKGROUND",(0,erow),(-1,erow),LINEN),
                   ("FONTNAME",(0,erow),(0,erow),SANS_B)]
        if totrow is not None:
            ts += [("LINEABOVE",(0,totrow),(-1,totrow),0.5,INK),
                   ("BACKGROUND",(0,totrow),(-1,totrow),LINEN),
                   ("FONTNAME",(0,totrow),(-1,totrow),SANS_B)]
        if highlight_rows_dict:
            for ri, bg in highlight_rows_dict.items():
                if bg is not None and 0<=ri<len(data):
                    ts.append(("BACKGROUND",(0,ri),(-1,ri),bg))
        t.setStyle(TableStyle(ts)); return t

    # ══════════════════════════════════════════════════════════════════════════
    # ASSEMBLE
    # ══════════════════════════════════════════════════════════════════════════
    buf = io.BytesIO()
    doc = BaseDocTemplate(
        buf, pagesize=A4,
        leftMargin=LM, rightMargin=RM,
        topMargin=TM+8*mm, bottomMargin=BM+8*mm,
    )
    frame = Frame(LM, BM+8*mm, BW, PH-TM-BM-16*mm, id="body")
    doc.addPageTemplates([PageTemplate(id="report", frames=[frame], onPage=_chrome)])
    story = []
    story += _p1(); story.append(PageBreak())
    story += _p2(); story.append(PageBreak())
    story += _p3(); story.append(PageBreak())
    story += _p4(); story.append(PageBreak())
    story += _p5(); story.append(PageBreak())
    story += _p6()
    doc.build(story)
    return buf.getvalue()



if "show_save_farm_form" not in st.session_state:
    st.session_state["show_save_farm_form"] = False

# Sidebar widget default values — overwritten when a farm is loaded
_SIDEBAR_DEFAULTS = {
    "roi_country":           "Germany",
    "roi_crop":              "Lettuce (Butterhead)",
    "roi_footprint":         1000,
    "roi_levels":            5,
    "roi_lights_tier":       "Basic",
    "roi_hvac":              "Standard",
    "roi_automation":        "Medium",
    "roi_price_scenario":    "base",
    "roi_harvest_mode":      "Single",
    "roi_price_override":    0.0,
    "roi_packaging_cost":    0.15,
    "roi_loss_rate":         5.0,
    "roi_net_grow_factor":   85.0,
    "roi_walkways_factor":   15.0,
    "roi_water_price":       2.0,
    "roi_kwh_override":      0.0,
    "roi_rent_monthly":      0.0,
    "roi_real_estate_capex": 0.0,
    "roi_depreciation_years": 10,
    "roi_tax_rate":          25.0,
    "roi_ltv":               60.0,
    "roi_interest_rate":     5.5,
    "roi_loan_term_years":   10,
}
for k, v in _SIDEBAR_DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── Consume pending farm load (written by load handlers, applied before any widget renders)
if "_pending_farm_load" in st.session_state:
    _pf = st.session_state.pop("_pending_farm_load")

    # ── Parse crop mix once ───────────────────────────────────────────────────
    _pf_mix_raw = _pf.get("crop_mix_json")
    _pf_parsed_mix = None
    if _pf_mix_raw:
        try:
            _parsed_attempt = json.loads(_pf_mix_raw) if isinstance(_pf_mix_raw, str) else _pf_mix_raw
            if isinstance(_parsed_attempt, list) and _parsed_attempt:
                # Migration: rename Sweet Pepper inside crop mix
                for _m_row in _parsed_attempt:
                    if _m_row.get("crop") == "Sweet Pepper":
                        _pf_src = (_pf.get("crop_source") or "greenhouse").lower()
                        if _pf_src == "polytunnel":
                            _m_row["crop"] = "Sweet Pepper (Polytunnel)"
                        else:
                            _m_row["crop"] = "Sweet Pepper (GH Substrate)"
                _pf_parsed_mix = _parsed_attempt
        except Exception:
            pass
    _pf_crop_fallback = _pf.get("crop", "Lettuce (Butterhead)")

    # ── Pop ALL widget-owned keys before writing so Streamlit treats them as
    #    freshly instantiated and respects the new values. ────────────────────
    _all_widget_keys = [
        # VF
        "roi_country", "roi_crop", "roi_footprint", "roi_levels",
        "roi_lights_tier", "roi_hvac", "roi_automation", "roi_price_scenario",
        "roi_harvest_mode", "roi_price_override", "roi_packaging_cost",
        "roi_loss_rate", "roi_net_grow_factor", "roi_walkways_factor",
        "roi_water_price", "roi_rent_monthly", "roi_real_estate_capex", "roi_kwh_override",
        "roi_depreciation_years", "roi_tax_rate", "roi_ltv",
        "roi_interest_rate", "roi_loan_term_years", "roi_multi_crop",
        # GH
        "gh_country", "gh_crop", "gh_footprint", "gh_crop_source",
        "gh_automation", "gh_price_scenario", "gh_harvest_mode",
        "gh_price_override", "gh_packaging_cost", "gh_loss_rate",
        "gh_net_grow_factor", "gh_walkways_factor", "gh_water_price",
        "gh_rent_monthly", "gh_real_estate_capex", "gh_depreciation_years",
        "gh_tax_rate", "gh_ltv", "gh_interest_rate", "gh_loan_term_years",
        "gh_discount_rate", "gh_multi_crop",
        # AQ
        "aq_country", "aq_plant_crop", "aq_plant_crop_source",
        "aq_plant_footprint", "aq_automation", "aq_price_scenario",
        "aq_harvest_mode", "aq_packaging_cost", "aq_loss_rate",
        "aq_net_grow_factor", "aq_walkways_factor", "aq_water_price",
        "aq_rent_monthly", "aq_real_estate_capex", "aq_depreciation_years",
        "aq_tax_rate", "aq_ltv", "aq_interest_rate", "aq_loan_term_years",
        "aq_discount_rate", "aq_multi_crop",
    ]
    for _k in _all_widget_keys:
        st.session_state.pop(_k, None)
    # Dynamic crop mix row keys (up to 6 rows, all three modality prefixes)
    for _i in range(6):
        for _pfx in ("roi", "gh", "aq"):
            st.session_state.pop(f"{_pfx}_mix_crop_{_i}", None)
            st.session_state.pop(f"{_pfx}_mix_pct_{_i}", None)

    # ── VF keys ───────────────────────────────────────────────────────────────
    st.session_state["roi_country"]            = _pf.get("country", "Germany")
    st.session_state["roi_crop"]               = _pf_crop_fallback
    st.session_state["roi_footprint"]          = int(_pf.get("footprint") or 1000)
    st.session_state["roi_levels"]             = int(_pf.get("levels") or 5)
    st.session_state["roi_lights_tier"]        = _pf.get("lights_tier", "Basic")
    st.session_state["roi_hvac"]               = _pf.get("hvac", "Standard")
    st.session_state["roi_automation"]         = _pf.get("automation", "Medium")
    st.session_state["roi_price_scenario"]     = _pf.get("price_scenario", "base")
    st.session_state["roi_harvest_mode"]       = _pf.get("harvest_mode", "Single")
    st.session_state["roi_price_override"]     = float(_pf.get("price_override") or 0.0)
    st.session_state["roi_packaging_cost"]     = float(_pf.get("packaging_cost") or 0.15)
    st.session_state["roi_loss_rate"]          = float(_pf.get("loss_rate") or 5.0)
    st.session_state["roi_net_grow_factor"]    = float(_pf.get("net_grow_factor") or 85.0)
    st.session_state["roi_walkways_factor"]    = float(_pf.get("walkways_factor") or 15.0)
    st.session_state["roi_water_price"]        = float(_pf.get("water_price") or 2.0)
    st.session_state["roi_rent_monthly"]       = float(_pf.get("rent_monthly") or 0.0)
    st.session_state["roi_real_estate_capex"]  = float(_pf.get("real_estate_capex") or 0.0)
    st.session_state["roi_depreciation_years"] = int(_pf.get("depreciation_years") or 10)
    st.session_state["roi_tax_rate"]           = float(_pf.get("tax_rate") or 25.0)
    st.session_state["roi_ltv"]                = float(_pf.get("ltv") or 60.0)
    st.session_state["roi_interest_rate"]      = float(_pf.get("interest_rate") or 5.5)
    st.session_state["roi_loan_term_years"]    = int(_pf.get("loan_term_years") or 10)
    # Filter mix to only crops valid in the VF CROPS dict
    _vf_valid_fallback = (_pf_crop_fallback if _pf_crop_fallback in CROPS
                          else list(CROPS.keys())[0])
    _vf_filtered_mix = ([row for row in _pf_parsed_mix if row.get("crop") in CROPS]
                        if _pf_parsed_mix else [])
    if _vf_filtered_mix:
        st.session_state["roi_multi_crop"] = True
        st.session_state["roi_crop_mix"]   = _vf_filtered_mix
        for _i, _row in enumerate(_vf_filtered_mix):
            st.session_state[f"roi_mix_crop_{_i}"] = _row.get("crop", _vf_valid_fallback)
            st.session_state[f"roi_mix_pct_{_i}"]  = int(_row.get("pct", 100))
    else:
        st.session_state["roi_multi_crop"] = False
        st.session_state["roi_crop_mix"]   = [{"crop": _vf_valid_fallback, "pct": 100}]

    # ── GH keys ───────────────────────────────────────────────────────────────
    _gh_crop_source_raw = (_pf.get("crop_source") or "greenhouse").lower()
    _gh_crop_source_val = "Polytunnel" if _gh_crop_source_raw == "polytunnel" else "Greenhouse"
    _gh_valid_dict      = POLYTUNNEL_CROPS if _gh_crop_source_val == "Polytunnel" else GREENHOUSE_CROPS
    _gh_crop_fallback   = _pf.get("crop", "Tomato (Beef)")
    # Migration: rename Sweet Pepper to modality-specific name after data_tables rename
    if _gh_crop_fallback == "Sweet Pepper":
        if _gh_crop_source_val == "Polytunnel":
            _gh_crop_fallback = "Sweet Pepper (Polytunnel)"
        else:
            _gh_crop_fallback = "Sweet Pepper (GH Substrate)"

    _gh_valid_fallback  = _gh_crop_fallback if _gh_crop_fallback in _gh_valid_dict else list(_gh_valid_dict.keys())[0]
    st.session_state["gh_country"]            = _pf.get("country", "Germany")
    st.session_state["gh_crop"]               = _gh_valid_fallback
    st.session_state["gh_footprint"]          = int(_pf.get("footprint") or 1000)
    st.session_state["gh_crop_source"]        = _gh_crop_source_val
    st.session_state["gh_automation"]         = _pf.get("automation", "Medium")
    st.session_state["gh_price_scenario"]     = _pf.get("price_scenario", "base")
    st.session_state["gh_harvest_mode"]       = _pf.get("harvest_mode", "Single")
    st.session_state["gh_price_override"]     = float(_pf.get("price_override") or 0.0)
    st.session_state["gh_packaging_cost"]     = float(_pf.get("packaging_cost") or 0.15)
    st.session_state["gh_loss_rate"]          = float(_pf.get("loss_rate") or 5.0)
    st.session_state["gh_net_grow_factor"]    = float(_pf.get("net_grow_factor") or 85.0)
    st.session_state["gh_walkways_factor"]    = float(_pf.get("walkways_factor") or 15.0)
    st.session_state["gh_water_price"]        = float(_pf.get("water_price") or 2.0)
    st.session_state["gh_rent_monthly"]       = float(_pf.get("rent_monthly") or 0.0)
    st.session_state["gh_real_estate_capex"]  = float(_pf.get("real_estate_capex") or 0.0)
    st.session_state["gh_depreciation_years"] = int(_pf.get("depreciation_years") or 10)
    st.session_state["gh_tax_rate"]           = float(_pf.get("tax_rate") or 25.0)
    st.session_state["gh_ltv"]                = float(_pf.get("ltv") or 60.0)
    st.session_state["gh_interest_rate"]      = float(_pf.get("interest_rate") or 5.5)
    st.session_state["gh_loan_term_years"]    = int(_pf.get("loan_term_years") or 10)
    st.session_state["gh_discount_rate"]      = float(_pf.get("discount_rate") or 8.0)
    # Filter mix to only crops valid in the GH crop source dict
    _gh_filtered_mix = ([row for row in _pf_parsed_mix if row.get("crop") in _gh_valid_dict]
                        if _pf_parsed_mix else [])
    if _gh_filtered_mix:
        st.session_state["gh_multi_crop"] = True
        st.session_state["gh_crop_mix"]   = _gh_filtered_mix
        for _i, _row in enumerate(_gh_filtered_mix):
            st.session_state[f"gh_mix_crop_{_i}"] = _row.get("crop", _gh_valid_fallback)
            st.session_state[f"gh_mix_pct_{_i}"]  = int(_row.get("pct", 100))
    else:
        st.session_state["gh_multi_crop"] = False
        st.session_state["gh_crop_mix"]   = [{"crop": _gh_valid_fallback, "pct": 100}]

    # ── AQ keys ───────────────────────────────────────────────────────────────
    _aq_crop_fallback = _pf.get("crop", "Lettuce (Romaine)")
    # Migration: same rename guard for aquaponics plant-side load
    if _aq_crop_fallback == "Sweet Pepper":
        if (_pf.get("crop_source") or "greenhouse").lower() == "polytunnel":
            _aq_crop_fallback = "Sweet Pepper (Polytunnel)"
        else:
            _aq_crop_fallback = "Sweet Pepper (GH Substrate)"

    st.session_state["aq_country"]            = _pf.get("country", "Germany")
    # Restore fish species if saved on the farm record
    if _pf.get("fish_species") and _pf["fish_species"] in list(FISH_SPECIES.keys()):
        st.session_state["aq_species"] = _pf["fish_species"]

    st.session_state["aq_plant_crop"]         = _aq_crop_fallback
    st.session_state["aq_plant_crop_source"]  = (
        "Polytunnel" if (_pf.get("crop_source") or "greenhouse").lower() == "polytunnel"
        else "Greenhouse"
    )
    st.session_state["aq_plant_footprint"]    = int(_pf.get("footprint") or 1000)
    st.session_state["aq_automation"]         = _pf.get("automation", "Medium")
    st.session_state["aq_price_scenario"]     = _pf.get("price_scenario", "base")
    st.session_state["aq_harvest_mode"]       = _pf.get("harvest_mode", "Single")
    st.session_state["aq_packaging_cost"]     = float(_pf.get("packaging_cost") or 0.15)
    st.session_state["aq_loss_rate"]          = float(_pf.get("loss_rate") or 5.0)
    st.session_state["aq_net_grow_factor"]    = float(_pf.get("net_grow_factor") or 90.0)
    st.session_state["aq_walkways_factor"]    = float(_pf.get("walkways_factor") or 10.0)
    st.session_state["aq_water_price"]        = float(_pf.get("water_price") or 2.0)
    st.session_state["aq_rent_monthly"]       = float(_pf.get("rent_monthly") or 0.0)
    st.session_state["aq_real_estate_capex"]  = float(_pf.get("real_estate_capex") or 0.0)
    st.session_state["aq_depreciation_years"] = int(_pf.get("depreciation_years") or 15)
    st.session_state["aq_tax_rate"]           = float(_pf.get("tax_rate") or 25.0)
    st.session_state["aq_ltv"]                = float(_pf.get("ltv") or 60.0)
    st.session_state["aq_interest_rate"]      = float(_pf.get("interest_rate") or 5.5)
    st.session_state["aq_loan_term_years"]    = int(_pf.get("loan_term_years") or 15)
    st.session_state["aq_discount_rate"]      = float(_pf.get("discount_rate") or 8.0)
    # AQ fish/tank params — stored in metadata jsonb if present
    _pf_meta = _pf.get("metadata") or {}
    if isinstance(_pf_meta, str):
        try:
            _pf_meta = json.loads(_pf_meta)
        except Exception:
            _pf_meta = {}
    if _pf_meta.get("species"):
        st.session_state["aq_species"]       = _pf_meta["species"]
    if _pf_meta.get("tank_volume_m3"):
        st.session_state["aq_tank_volume_m3"] = float(_pf_meta["tank_volume_m3"])
    if _pf_meta.get("target_temp_c"):
        st.session_state["aq_target_temp_c"]  = float(_pf_meta["target_temp_c"])
    # Filter mix to only crops valid in the AQ plant crop source dict
    _aq_crop_source_for_filter = (_pf.get("crop_source") or "greenhouse").lower()
    _aq_valid_dict = POLYTUNNEL_CROPS if _aq_crop_source_for_filter == "polytunnel" else GREENHOUSE_CROPS
    _aq_valid_fallback = (list(_aq_valid_dict.keys())[0]
                          if _aq_crop_fallback not in _aq_valid_dict else _aq_crop_fallback)
    _aq_filtered_mix = ([row for row in _pf_parsed_mix if row.get("crop") in _aq_valid_dict]
                        if _pf_parsed_mix else [])
    if _aq_filtered_mix:
        st.session_state["aq_multi_crop"] = True
        st.session_state["aq_crop_mix"]   = _aq_filtered_mix
        for _i, _row in enumerate(_aq_filtered_mix):
            st.session_state[f"aq_mix_crop_{_i}"] = _row.get("crop", _aq_valid_fallback)
            st.session_state[f"aq_mix_pct_{_i}"]  = int(_row.get("pct", 100))
    else:
        st.session_state["aq_multi_crop"] = False
        st.session_state["aq_crop_mix"]   = [{"crop": _aq_valid_fallback, "pct": 100}]

def _run_multicrop_generic(base_inputs: dict, crop_mix: list,
                            calc_fn, crop_data_dict: dict,
                            dli_key: str = "dli") -> dict:
    """
    Generic multi-crop aggregation engine used by all modalities.
    calc_fn:        the calculation function (calculate or calculate_greenhouse)
    crop_data_dict: the crop dictionary to read DLI from (CROPS, GREENHOUSE_CROPS, etc.)
    dli_key:        field name for DLI in the crop dict (default "dli")
    """
    import copy as _copy
    total_pct = sum(row["pct"] for row in crop_mix)
    if total_pct == 0:
        return calc_fn(base_inputs)

    # Per-crop sub-area runs
    crop_results = []
    for row in crop_mix:
        frac    = row["pct"] / total_pct
        sub_inp = _copy.deepcopy(base_inputs)
        sub_inp["crop"]      = row["crop"]
        sub_inp["footprint"] = base_inputs["footprint"] * frac
        crop_results.append((frac, row["crop"], calc_fn(sub_inp)))

    # Aggregate revenue and operating costs
    annual_revenue       = sum(cr["annual_revenue"]       for _, _, cr in crop_results)
    annual_variable_cost = sum(cr["annual_variable_cost"] for _, _, cr in crop_results)
    annual_water_cost    = sum(cr["annual_water_cost"]    for _, _, cr in crop_results)
    annual_labour_cost   = sum(cr["annual_labour_cost"]   for _, _, cr in crop_results)
    annual_rent          = crop_results[0][2]["annual_rent"]

    # Single energy calculation using area-weighted DLI
    _ref_crop    = max(crop_mix, key=lambda r: crop_data_dict[r["crop"]][dli_key])["crop"]
    weighted_dli = sum(crop_data_dict[cn][dli_key] * f for f, cn, _ in crop_results)
    _orig        = crop_data_dict[_ref_crop]
    _patched     = _copy.deepcopy(_orig)
    _patched[dli_key] = weighted_dli
    crop_data_dict[_ref_crop] = _patched
    _energy_inp  = _copy.deepcopy(base_inputs)
    _energy_inp["crop"] = _ref_crop
    try:
        _energy_r = calc_fn(_energy_inp)
    finally:
        crop_data_dict[_ref_crop] = _orig
    annual_energy_cost = _energy_r["annual_energy_cost"]
    total_annual_kwh   = _energy_r.get("total_annual_kwh") or _energy_r.get("annual_kwh", 0)

    # Single CAPEX calculation using highest-DLI crop
    _capex_inp = _copy.deepcopy(base_inputs)
    _capex_inp["crop"] = _ref_crop
    _capex_r = calc_fn(_capex_inp)
    total_capex        = _capex_r["total_capex"]
    annual_maintenance = _capex_r["annual_maintenance"]

    # EBITDA and financial layer
    total_annual_costs = (annual_energy_cost + annual_variable_cost +
                          annual_water_cost + annual_labour_cost +
                          annual_rent + annual_maintenance)
    ebitda        = annual_revenue - total_annual_costs
    ebitda_margin = ebitda / annual_revenue if annual_revenue > 0 else 0.0

    annual_depreciation = _capex_r["annual_depreciation"]
    annual_debt_service = _capex_r["annual_debt_service"]
    dscr            = ebitda / annual_debt_service if annual_debt_service > 0 else None
    equity_invested = total_capex * (1 - base_inputs["ltv"] / 100)
    annual_fcfe     = (ebitda - annual_debt_service
                       - annual_depreciation * (base_inputs["tax_rate"] / 100))
    payback_years   = (equity_invested / annual_fcfe
                       if annual_fcfe > 0 and equity_invested > 0 else None)

    _disc = base_inputs.get("discount_rate", 8.0) / 100
    dcf_cashflows  = []
    cumulative_npv = -equity_invested
    for yr in range(1, 11):
        pv = annual_fcfe / ((1 + _disc) ** yr)
        cumulative_npv += pv
        dcf_cashflows.append({"year": yr, "fcfe": annual_fcfe,
                               "pv": pv, "cumulative_npv": cumulative_npv})

    combined = dict(_capex_r)
    combined.update({
        "annual_revenue":       annual_revenue,
        "annual_energy_cost":   annual_energy_cost,
        "annual_variable_cost": annual_variable_cost,
        "annual_water_cost":    annual_water_cost,
        "annual_labour_cost":   annual_labour_cost,
        "annual_maintenance":   annual_maintenance,
        "annual_rent":          annual_rent,
        "total_annual_costs":   total_annual_costs,
        "ebitda":               ebitda,
        "ebitda_margin":        ebitda_margin,
        "total_capex":          total_capex,
        "total_annual_kwh":     total_annual_kwh,
        "annual_kwh":           total_annual_kwh,
        "annual_depreciation":  annual_depreciation,
        "annual_debt_service":  annual_debt_service,
        "dscr":                 dscr,
        "payback_years":        payback_years,
        "dcf_cashflows":        dcf_cashflows,
        "npv":                  cumulative_npv,
        "total_annual_kg":      sum(cr["total_annual_kg"] for _, _, cr in crop_results),
        "_is_multicrop":        True,
        "_crop_results": [
            {
                "crop":                 cn,
                "pct":                  f * 100,
                "annual_revenue":       cr["annual_revenue"],
                "annual_variable_cost": cr["annual_variable_cost"],
                "annual_labour_cost":   cr["annual_labour_cost"],
                "annual_water_cost":    cr["annual_water_cost"],
                "ebitda":               cr["ebitda"],
                "total_annual_kg":      cr["total_annual_kg"],
                "effective_price":      cr["effective_price"],
            }
            for f, cn, cr in crop_results
        ],
    })
    return combined

# Apply pending modality switch (set by load handlers, applied before radio renders)
if "_pending_modality" in st.session_state:
    st.session_state["cea_modality"] = st.session_state.pop("_pending_modality")

# ── Persistent farm context (top of sidebar, all modalities) ────────────
_active_farm_global = render_farm_context_sidebar(supabase=supabase)
 
# ── Setup gate — no farm loaded ──────────────────────────────────────────
if not _active_farm_global:
    st.title("CEA Feasibility Calculator") # Remove emoji from title
    st.info(
        "**No farm profile loaded.**\n\n"
        "Select an existing farm profile in the sidebar to run the analysis, "
        "or configure the parameters below and use **Save as Farm Profile** to create one.\n\n"
        "👈 Use the sidebar to load or create a farm."
    )
    st.markdown("---")
    st.caption( # Keep emoji in page link
        "First time? Pick a modality, fill in the parameters below, run the calculation, "
        "then click 💾 Save as Farm Profile at the bottom of the results."
    )

modality = st.radio(
    "Select farming modality", # Keep emojis in radio options
    options=[
        "🏭 Indoor Vertical Farm",
        "🌿 High-Tech Greenhouse",
        "🐟 Decoupled Aquaponics",
        "♻️ Coupled Aquaponics",
    ],
    horizontal=True,
    key="cea_modality",
)

if modality == "🏭 Indoor Vertical Farm":

    # ── Sidebar ───────────────────────────────────────────────────────────────────
    with st.sidebar:
        st.header("Farm Parameters")
    
        # ── Farm parameters ───────────────────────────────────────────────────────
        country_list = list(COUNTRIES.keys())
        _c_default   = st.session_state["roi_country"]
        country = st.selectbox("Country", country_list,
                               index=country_list.index(_c_default) if _c_default in country_list else 0,
                               key="roi_country")

        # ── Energy & Labour reference rates (energy_labour module) ────────────
        _el_rates  = get_rates_for_country_name(country)
        _el_e      = _el_rates["energy"]
        _el_l      = _el_rates["labour"]
        _model_kwh = COUNTRIES.get(country, {}).get("kwh", 0)
        if _el_rates["iso"]:
            _el_delta = _el_e["industrial"] - _model_kwh
            _el_arrow = "▲" if _el_delta > 0.005 else ("▼" if _el_delta < -0.005 else "≈")
            _el_col   = "#d4a845" if _el_delta > 0.005 else ("#52a066" if _el_delta < -0.005 else "#7a8070")
            _el_live  = " · ⚡ Live" if _el_e.get("live") else ""
            _el_html  = (
                f"<div style='font-size:11px;color:#9ba390;margin:-4px 0 8px 0;"
                f"padding:6px 8px;background:#252a25;border-radius:3px;"
                f"border:1px solid #363c36;border-left:3px solid {_el_col};'>"
                f"<b>Ref. rates ({_el_rates['iso']})</b> &nbsp;&middot;&nbsp; "
                f"Electricity industrial: <b style='color:{_el_col};'>${_el_e['industrial']:.3f}/kWh {_el_arrow}</b> "
                f"<span style='color:#7a8070;'>(model ${_model_kwh:.3f})</span>"
                f"&nbsp;&middot;&nbsp; Labour: <b>${_el_l['industrial_loaded']:.0f}/hr</b>"
                f"{_el_live}</div>"
            )
            st.markdown(_el_html, unsafe_allow_html=True)

        # ── Multi-crop toggle ─────────────────────────────────────────────────
        multi_crop_mode = st.toggle("Multi-crop mode",
                                     value=st.session_state.get("roi_multi_crop", False),
                                     key="roi_multi_crop")

        if not multi_crop_mode:
            crop_list   = list(CROPS.keys())
            _cr_default = st.session_state["roi_crop"]
            crop = st.selectbox("Crop", crop_list,
                                index=crop_list.index(_cr_default) if _cr_default in crop_list else 0,
                                key="roi_crop")
        else:
            crop = list(CROPS.keys())[0]  # placeholder — unused in multi-crop mode
            st.markdown("**Crop allocation** (must sum to 100%)")
            if "roi_crop_mix" not in st.session_state:
                st.session_state["roi_crop_mix"] = [{"crop": "Lettuce (Butterhead)", "pct": 100}]
            _mix = st.session_state["roi_crop_mix"]
            _crop_list_all = list(CROPS.keys())
            _to_remove = None
            for _ci, _row in enumerate(_mix):
                _rc1, _rc2, _rc3 = st.columns([4, 2, 1])
                with _rc1:
                    _mix[_ci]["crop"] = st.selectbox(
                        f"Crop {_ci+1}", _crop_list_all,
                        index=_crop_list_all.index(_row["crop"]) if _row["crop"] in _crop_list_all else 0,
                        key=f"roi_mix_crop_{_ci}")
                with _rc2:
                    _mix[_ci]["pct"] = st.number_input(
                        "%", min_value=1, max_value=100, value=int(_row["pct"]),
                        step=1, key=f"roi_mix_pct_{_ci}")
                with _rc3:
                    if len(_mix) > 1 and st.button("✕", key=f"roi_mix_del_{_ci}"):
                        _to_remove = _ci
            if _to_remove is not None:
                _mix.pop(_to_remove)
                st.session_state["roi_crop_mix"] = _mix
                st.rerun()
            _total_pct = sum(r["pct"] for r in _mix)
            st.caption(f"Total allocated: **{_total_pct}%**")
            if _total_pct != 100:
                st.warning(f"⚠️ Must sum to 100%. Currently {_total_pct}%.")
            if len(_mix) < 6:
                if st.button("➕ Add crop", key="roi_mix_add"):
                    _mix.append({"crop": _crop_list_all[0], "pct": 1})
                    st.session_state["roi_crop_mix"] = _mix
                    st.rerun()
            st.session_state["roi_crop_mix"] = _mix
    
        st.divider()
        footprint = st.number_input("Footprint (m²)", value=st.session_state["roi_footprint"],
                                    step=100, min_value=50, key="roi_footprint")
        levels    = st.number_input("Levels", value=st.session_state["roi_levels"],
                                    min_value=1, max_value=20, step=1, key="roi_levels")
    
        st.divider()
        lights_list    = list(LIGHTS.keys())
        _lt_default    = st.session_state["roi_lights_tier"]
        lights_tier    = st.selectbox("Lights", lights_list,
                                      index=lights_list.index(_lt_default) if _lt_default in lights_list else 1,
                                      key="roi_lights_tier")
    
        hvac_list   = ["Excellent conditions", "Standard", "High HVAC"]
        _hv_default = st.session_state["roi_hvac"]
        # Auto-suggest HVAC tier from saved ambient temperature if available
        _vf_amb_temp = st.session_state.get("active_farm", {}).get("ambient_temp_annual")
        if _vf_amb_temp is not None:
            _vf_suggested_hvac = (
                "Excellent conditions" if _vf_amb_temp >= 17 # Remove emoji from caption
                else "Standard" if _vf_amb_temp >= 12
                else "High HVAC"
            )
            if _hv_default != _vf_suggested_hvac:
                st.caption(
                    f"\U0001f4a1 Climate data suggests **{_vf_suggested_hvac}** "
                    f"(ambient {_vf_amb_temp:.1f}\u00b0C \u00b7 "
                    f"\u226517\u00b0C \u2192 Excellent \u00b7 12\u201317\u00b0C \u2192 Standard \u00b7 <12\u00b0C \u2192 High HVAC). " # Remove emoji from caption
                    f"Current selection: {_hv_default}."
                )
        hvac        = st.selectbox("HVAC conditions",
                                   hvac_list,
                                   index=hvac_list.index(_hv_default) if _hv_default in hvac_list else 1,
                                   key="roi_hvac",
                                   help=(
                                       "Controls total energy multiplier (lighting + HVAC + pumps + controls). "
                                       "Excellent = 1.70\u00d7 \u00b7 Standard = 1.83\u00d7 \u00b7 High HVAC = 2.025\u00d7. "
                                       "Reflects facility insulation quality and climate severity. "
                                       "Set farm coordinates in the Farm Intelligence Map to get an automatic suggestion."
                                   ))
    
        auto_list   = ["None", "Low", "Medium", "High"]
        _au_default = st.session_state["roi_automation"]
        automation  = st.selectbox("Automation", auto_list,
                                   index=auto_list.index(_au_default) if _au_default in auto_list else 2,
                                   key="roi_automation")
    
        ps_list     = ["base", "low", "high"]
        _ps_default = st.session_state["roi_price_scenario"]
        price_scenario = st.selectbox("Price Scenario", ps_list,
                                      index=ps_list.index(_ps_default) if _ps_default in ps_list else 0,
                                      key="roi_price_scenario")
    
        # Harvest mode — only enable multi-harvest if crop supports it
        crop_data = CROPS[crop]
        if crop_data["days_between"] > 0:
            hm_list     = ["Single", "2 Harvests", "3 Harvests"]
            _hm_default = st.session_state["roi_harvest_mode"]
            harvest_mode = st.selectbox("Harvest Mode", hm_list,
                                        index=hm_list.index(_hm_default) if _hm_default in hm_list else 0,
                                        key="roi_harvest_mode")
        else:
            st.selectbox("Harvest Mode", ["Single"], disabled=True,
                         help="This crop only supports single harvest.")
            harvest_mode = "Single"
            st.session_state["roi_harvest_mode"] = "Single"
    
        st.divider()
        st.subheader("Advanced")
        price_override    = st.number_input("Price Override ($/kg, 0=auto)",
                                            value=st.session_state["roi_price_override"],
                                            step=0.1, min_value=0.0, key="roi_price_override")
        packaging_cost    = st.number_input("Packaging ($/kg)",
                                            value=st.session_state["roi_packaging_cost"],
                                            step=0.01, min_value=0.0, key="roi_packaging_cost")
        loss_rate         = st.number_input("Loss Rate (%)",
                                            value=st.session_state["roi_loss_rate"],
                                            step=0.5, min_value=0.0, max_value=100.0, key="roi_loss_rate")
        net_grow_factor   = st.number_input("Net Grow Factor (%)",
                                            value=st.session_state["roi_net_grow_factor"],
                                            step=1.0, min_value=1.0, max_value=100.0, key="roi_net_grow_factor")
        walkways_factor   = st.number_input("Walkways Factor (%)",
                                            value=st.session_state["roi_walkways_factor"],
                                            step=1.0, min_value=0.0, max_value=50.0, key="roi_walkways_factor")
        water_price       = st.number_input("Water Price ($/m³)",
                                            value=st.session_state["roi_water_price"],
                                            step=0.1, min_value=0.0, key="roi_water_price")
        _vf_kwh_default   = COUNTRIES.get(country, {}).get("kwh", 0.0)
        kwh_override      = st.number_input(
            "Electricity Price ($/kWh)",
            value=float(st.session_state.get("roi_kwh_override") or _vf_kwh_default),
            step=0.005, min_value=0.001, format="%.4f", key="roi_kwh_override",
            help=(
                f"Country default (from IEA/Eurostat table): ${_vf_kwh_default:.4f}/kWh. "
                "Override with your actual site tariff if different. "
                "Industrial/commercial rate is typically 30–60% lower than residential."
            )
        )
        rent_monthly      = st.number_input("Monthly Rent ($)",
                                            value=st.session_state["roi_rent_monthly"],
                                            step=100.0, min_value=0.0, key="roi_rent_monthly")
        real_estate_capex = st.number_input("Real Estate CAPEX ($)",
                                            value=st.session_state["roi_real_estate_capex"],
                                            step=10000.0, min_value=0.0, key="roi_real_estate_capex")
    
        st.divider()
        st.subheader("Financial Structure")
        depreciation_years  = st.number_input("Depreciation (years)",
                                              value=st.session_state["roi_depreciation_years"],
                                              step=1, min_value=1, key="roi_depreciation_years")
        tax_rate_input      = st.number_input("Tax Rate (%)",
                                              value=st.session_state["roi_tax_rate"],
                                              step=1.0, min_value=0.0, max_value=100.0, key="roi_tax_rate")
        ltv_input           = st.number_input("LTV (%)",
                                              value=st.session_state["roi_ltv"],
                                              step=5.0, min_value=0.0, max_value=100.0, key="roi_ltv")
        interest_rate_input = st.number_input("Interest Rate (%)",
                                              value=st.session_state["roi_interest_rate"],
                                              step=0.1, min_value=0.0, key="roi_interest_rate")
        loan_term_years     = st.number_input("Loan Term (years)",
                                              value=st.session_state["roi_loan_term_years"],
                                              step=1, min_value=1, key="roi_loan_term_years")

    _multi_crop_mode = st.session_state.get("roi_multi_crop", False)
    _crop_mix        = st.session_state.get("roi_crop_mix", [])

    # Sanitise locally — only keep crops valid in VF CROPS dict
    _crop_mix  = [row for row in _crop_mix if row.get("crop") in CROPS]
    if not _crop_mix:
        _multi_crop_mode = False

    _mix_total       = sum(row["pct"] for row in _crop_mix)
    _mix_valid       = _multi_crop_mode and len(_crop_mix) > 0 and _mix_total == 100
    
    # ── Run calculation ───────────────────────────────────────────────────────────
    inputs = {
        "country":           country,
        "crop":              crop,
        "footprint":         footprint,
        "levels":            levels,
        "lights_tier":       lights_tier,
        "hvac":              hvac,
        "automation":        automation,
        "price_scenario":    price_scenario,
        "price_override":    price_override,
        "packaging_cost":    packaging_cost,
        "loss_rate":         loss_rate,
        "net_grow_factor":   net_grow_factor,
        "walkways_factor":   walkways_factor,
        "water_price":       water_price,
        "rent_monthly":      rent_monthly,
        "real_estate_capex": real_estate_capex,
        "harvest_mode":      harvest_mode,
        "depreciation_years": depreciation_years,
        "tax_rate":          tax_rate_input,
        "ltv":               ltv_input,
        "interest_rate":     interest_rate_input,
        "loan_term_years":   loan_term_years,
        "discount_rate":     8.0,
        "crop_mix_json":     json.dumps(_crop_mix) if _mix_valid else None,
    }
    
    def run_multicrop(base_inputs: dict, crop_mix: list) -> dict:
        import core.data_tables as _dt
        return _run_multicrop_generic(base_inputs, crop_mix, calculate, _dt.CROPS)

    # Guard: ensure all dict lookups are valid before calculation.
    # Fields can be None if a non-VF farm was loaded and session state bled over.
    from core.data_tables import LIGHTS as _LIGHTS_CHECK, HVAC_FACTORS as _HVAC_CHECK
    if not inputs.get("lights_tier") or inputs["lights_tier"] not in _LIGHTS_CHECK:
        inputs["lights_tier"] = "Basic"
    if not inputs.get("hvac") or inputs["hvac"] not in _HVAC_CHECK:
        inputs["hvac"] = "Standard"
    if not inputs.get("crop") or inputs["crop"] not in CROPS:
        inputs["crop"] = list(CROPS.keys())[0]
    if not inputs.get("country") or inputs["country"] not in COUNTRIES:
        inputs["country"] = "Germany"

    if _multi_crop_mode and not _mix_valid:
        st.warning("⚠️ Fix crop allocation (must sum to 100%) before results are shown.")
        st.stop()

    # Apply electricity price override (patches COUNTRIES entry temporarily)
    _vf_kwh_original = COUNTRIES[country]["kwh"]
    if abs(kwh_override - _vf_kwh_original) > 0.0001:
        COUNTRIES[country]["kwh"] = kwh_override
    if _mix_valid:
        r = run_multicrop(inputs, _crop_mix)
    else:
        r = calculate(inputs)
    COUNTRIES[country]["kwh"] = _vf_kwh_original  # always restore

    # ── Climate profile display ───────────────────────────────────────────────
    _active_farm_data = st.session_state.get("active_farm")
    if _active_farm_data and _active_farm_data.get("mean_annual_dli"):
        _loc_dli  = _active_farm_data["mean_annual_dli"]
        _loc_temp = _active_farm_data["ambient_temp_annual"]
        _crop_dli = CROPS[crop]["dli"]
        _nat_frac = compute_natural_dli_fraction(_loc_dli, _crop_dli)
        st.caption(
            f"**Climate profile active** — " # Remove emoji from caption
            f"Mean annual DLI: {_loc_dli:.1f} mol/m²/day · "
            f"Ambient temperature: {_loc_temp:.1f}°C · "
            f"Natural DLI coverage for {crop}: {_nat_frac*100:.0f}% "
            f"({'supplemental lighting required' if _nat_frac < 1.0 else 'no supplemental lighting required'})"
        )
    
    # ── Data Sources panel ───────────────────────────────────────────────────
    _vf_farm_has_climate = bool(st.session_state.get("active_farm", {}).get("mean_annual_dli"))
    _vf_active_data      = st.session_state.get("active_farm") or {} # Remove emoji from expander title
    with st.expander("ℹ️ Data sources & calculation transparency", expanded=False):
        _di1, _di2 = st.columns(2)
        with _di1:
            st.markdown("**📡 Automatic — from Open-Meteo Archive API**")
            if _vf_farm_has_climate:
                _vf_dli  = _vf_active_data.get("mean_annual_dli", 0)
                _vf_temp = _vf_active_data.get("ambient_temp_annual", 0)
                _vf_sugg = "Excellent conditions" if _vf_temp >= 17 else ("Standard" if _vf_temp >= 12 else "High HVAC") # Remove emoji from markdown
                st.markdown(
                    f"- **Mean annual DLI: {_vf_dli:.1f} mol/m²/day** "
                    f"— shown for information only. VF uses fully artificial lighting; "
                    f"location solar irradiance does not affect energy calculations.\n"
                    f"- **Ambient temperature: {_vf_temp:.1f}°C** "
                    f"— used to suggest HVAC tier (suggests **{_vf_sugg}** for this location). "
                    f"Does not directly enter the energy formula; the HVAC selectbox is the active input.\n"
                    f"- Source: Open-Meteo 10-year historical archive (`archive-api.open-meteo.com/v1/archive`). " # Remove emoji from markdown
                    f"Fetched once at farm save time, stored in Supabase."
                )
            else:
                st.markdown(
                    "- ⚠️ **No climate data available** for this farm.\n"
                    "- Set farm coordinates in the **Farm Intelligence Map**, then re-save the farm profile.\n"
                    "- Until then, the HVAC tier selection has no location-based suggestion."
                ) # Keep warning emoji
        with _di2:
            st.markdown("**🎛️ Manual inputs — set in this calculator**")
            _hvac_cur = st.session_state.get("roi_hvac", "Standard")
            _hvac_factors_map = {"Excellent conditions": "1.70×", "Standard": "1.83×", "High HVAC": "2.025×"}
            st.markdown(
                f"- **HVAC tier: {_hvac_cur}** (multiplier: {_hvac_factors_map.get(_hvac_cur, '1.83×')}) "
                f"— applied to lighting energy to yield total facility electricity (lighting + HVAC + pumps + controls). "
                f"Reflects insulation quality and climate severity. See Assumptions §3.\n"
                f"- Lights tier, automation, crop, footprint, levels\n"
                f"- Country → electricity price (IEA/Eurostat, see §5)\n"
                f"- All financial structure inputs (see §17.7)"
            )
        st.caption( # Remove emoji from caption
            "ℹ️ Formula: Energy = DLI × 0.2778 / LED efficacy × cycle days "
            "× HVAC factor × cycles/yr × EGA × €/kWh. Full derivation in Assumptions §3."
        )


    # ── Energy & Labour calibration callout ──────────────────────────────────
    _el_r2    = get_rates_for_country_name(inputs["country"])
    _el_e2    = _el_r2["energy"]
    _el_l2    = _el_r2["labour"]
    _mkwh2    = COUNTRIES.get(inputs["country"], {}).get("kwh", 0)
    _mlabour2 = COUNTRIES.get(inputs["country"], {}).get("labour", 0)
    if _el_r2["iso"]:
        _e_flag  = abs(_el_e2["industrial"] - _mkwh2) > 0.01 # Keep warning emoji
        _l_flag  = abs(_el_l2["industrial_loaded"] - _mlabour2) > 3.0
        _exp_lbl = "⚠️ Verify your input assumptions" if (_e_flag or _l_flag) else "✅ Input assumptions cross-check"
        with st.expander(_exp_lbl, expanded=(_e_flag or _l_flag)):
            _rc1, _rc2 = st.columns(2)
            with _rc1:
                st.markdown("**⚡ Electricity**")
                _e_dir = "higher" if _el_e2["industrial"] > _mkwh2 else "lower"
                _e_pct = abs(_el_e2["industrial"] - _mkwh2) / _mkwh2 * 100 if _mkwh2 else 0
                if _e_flag:
                    st.warning(
                        f"Reference industrial rate: **${_el_e2['industrial']:.3f}/kWh** "
                        f"({_e_pct:.0f}% {_e_dir} than model’s ${_mkwh2:.3f}/kWh). "
                        f"Verify country default or use a site-specific override if your tariff differs."
                    )
                else:
                    st.success(f"Model electricity (${_mkwh2:.3f}/kWh) aligns with reference industrial rate (${_el_e2['industrial']:.3f}/kWh).")
                st.caption(f"Source: {_el_e2['source']}" + (f" · ⚡ {_el_e2['live_note']}" if _el_e2.get("live") else ""))
            with _rc2:
                st.markdown("**👷 Labour**")
                _l_dir = "higher" if _el_l2["industrial_loaded"] > _mlabour2 else "lower"
                _l_pct = abs(_el_l2["industrial_loaded"] - _mlabour2) / _mlabour2 * 100 if _mlabour2 else 0
                if _l_flag:
                    st.warning(
                        f"Reference fully-loaded industrial: **${_el_l2['industrial_loaded']:.0f}/hr** "
                        f"({_l_pct:.0f}% {_l_dir} than model’s ${_mlabour2:.0f}/hr). "
                        f"Overhead {_el_l2['overhead_pct']} applied (base ${_el_l2['industrial_base']:.0f}/hr)."
                    )
                else:
                    st.success(f"Model labour (${_mlabour2:.0f}/hr) aligns with reference (${_el_l2['industrial_loaded']:.0f}/hr, overhead {_el_l2['overhead_pct']}).")
                st.caption(f"Source: {_el_l2['source']}")

    # ── PDF Report Generator ─────────────────────────────────────────────────
    def generate_pdf_report(inputs: dict, r: dict) -> bytes:
        _fn = st.session_state.get("active_farm", {}).get("name", "")
        def _vf_sens(kwh_m=1.0, lab_m=1.0, yld_m=1.0, prc_m=1.0):
            return run_with_multipliers(inputs, kwh_mult=kwh_m, labour_mult=lab_m,
                                        yield_mult=yld_m, price_mult=prc_m)
        return _build_feasibility_pdf(r, inputs, "vf", farm_name=_fn,
                                      run_sens_fn=_vf_sens)

    # ── Excel Report Generator ────────────────────────────────────────────────
    def generate_excel_report(inputs: dict, r: dict) -> bytes:
        """
        Build an investment-grade Excel workbook from the VF model results.
        Sheets: Cover | Dashboard | Inputs | P&L | CAPEX & Debt | DCF
        Returns raw bytes for st.download_button.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        # ── Colour palette ────────────────────────────────────────────────────
        C_SAGE       = "2F5D3A"
        C_SAGE_LIGHT = "E6EDE4"
        C_LINEN      = "F4F1EA"
        C_LINEN2     = "FBF9F4"
        C_INK        = "161A16"
        C_INK2       = "4A524A"
        C_INK3       = "7A807A"
        C_RULE       = "D6D2C4"
        C_AMBER      = "C08A2E"
        C_CLAY       = "B85C38"
        C_WHITE      = "FFFFFF"
        C_RED_LIGHT  = "FDF0EA"
        C_GREEN_LIGHT= "E6EDE4"
        C_AMBER_LIGHT= "FDF6E3"

        _today      = date.today()
        _date_str   = _today.strftime("%d %B %Y")
        _farm_name  = st.session_state.get("active_farm", {}).get("name", "—")
        _modality   = "Indoor Vertical Farm"
        # Determine the primary crop name for the report
        _multi_crop_mode_excel = st.session_state.get("roi_multi_crop", False)
        _crop_mix_excel = st.session_state.get("roi_crop_mix", [])
        if _multi_crop_mode_excel and _crop_mix_excel:
            _crop = _crop_mix_excel[0]["crop"]
        else:
            _crop = inputs.get("crop", "—")

        _country    = inputs.get("country", "—")
        _doc_id     = f"XLS-VF-{_today.strftime('%Y%m%d')}"

        # ── Helper styles ─────────────────────────────────────────────────────
        def _font(bold=False, size=10, color=C_INK, italic=False, name="Calibri"):
            return Font(name=name, bold=bold, size=size, color=color, italic=italic)

        def _fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def _align(h="left", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        def _border(style="thin", color=C_RULE):
            s = Side(style=style, color=color)
            return Border(left=s, right=s, top=s, bottom=s)

        def _bottom_border(style="thin", color=C_INK2):
            s = Side(style=style, color=color)
            return Border(bottom=s)

        def _set_col_width(ws, col_letter, width):
            ws.column_dimensions[col_letter].width = width

        def _hdr(ws, row, col, value, bg=C_SAGE, fg=C_WHITE, size=10, bold=True,
                 align="left", merge_to=None, italic=False):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font      = _font(bold=bold, size=size, color=fg, italic=italic)
            cell.fill      = _fill(bg)
            cell.alignment = _align(h=align, v="center")
            if merge_to:
                ws.merge_cells(
                    start_row=row, start_column=col,
                    end_row=row,   end_column=merge_to
                )
            return cell

        def _cell(ws, row, col, value, bold=False, size=10, color=C_INK,
                  align="left", bg=None, fmt=None, italic=False):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font      = _font(bold=bold, size=size, color=color, italic=italic)
            cell.alignment = _align(h=align, v="center")
            if bg:
                cell.fill = _fill(bg)
            if fmt:
                cell.number_format = fmt
            return cell

        def _num(ws, row, col, value, bold=False, color=C_INK, bg=None,
                 fmt='#,##0', negative_red=False):
            cell = ws.cell(row=row, column=col, value=value)
            if negative_red and isinstance(value, (int, float)) and value < 0:
                color = C_CLAY
            cell.font          = _font(bold=bold, size=10, color=color, name="Consolas")
            cell.alignment     = _align(h="right", v="center")
            cell.number_format = fmt
            if bg:
                cell.fill = _fill(bg)
            return cell

        def _section_hdr(ws, row, col_from, col_to, label):
            """Sage section separator row."""
            ws.row_dimensions[row].height = 18
            for c in range(col_from, col_to + 1):
                ws.cell(row=row, column=c).fill = _fill(C_SAGE_LIGHT)
            cell = ws.cell(row=row, column=col_from, value=label.upper())
            cell.font      = _font(bold=True, size=9, color=C_SAGE)
            cell.alignment = _align(h="left", v="center")
            ws.merge_cells(start_row=row, start_column=col_from,
                           end_row=row, end_column=col_to)

        def _divider(ws, row, col_from, col_to):
            for c in range(col_from, col_to + 1):
                ws.cell(row=row, column=c).border = _bottom_border(color=C_RULE)

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 1 — COVER
        # ══════════════════════════════════════════════════════════════════════
        ws_cov = wb.create_sheet("Cover")
        ws_cov.sheet_view.showGridLines = False
        _set_col_width(ws_cov, "A", 2)
        _set_col_width(ws_cov, "B", 28)
        _set_col_width(ws_cov, "C", 40)
        _set_col_width(ws_cov, "D", 20)
        _set_col_width(ws_cov, "E", 2)

        # Sage accent column
        for r_idx in range(1, 50):
            ws_cov.cell(row=r_idx, column=1).fill = _fill(C_SAGE)
            ws_cov.row_dimensions[r_idx].height = 16

        # Title block (rows 3-8)
        ws_cov.row_dimensions[3].height = 10
        ws_cov.row_dimensions[4].height = 36
        ws_cov.row_dimensions[5].height = 20
        ws_cov.row_dimensions[6].height = 14
        ws_cov.row_dimensions[7].height = 14
        ws_cov.row_dimensions[8].height = 20

        _cell(ws_cov, 4, 2, "CEA FEASIBILITY MODEL", bold=True, size=22,
              color=C_INK, align="left")
        ws_cov.merge_cells("B4:D4")

        _cell(ws_cov, 5, 2, _modality, bold=False, size=13, color=C_SAGE, align="left")
        ws_cov.merge_cells("B5:D5")

        _cell(ws_cov, 6, 2, f"Farm: {_farm_name}", size=10, color=C_INK2, italic=True)
        ws_cov.merge_cells("B6:D6")
        _cell(ws_cov, 7, 2, f"Crop: {_crop}  ·  Country: {_country}", size=10, color=C_INK2, italic=True)
        ws_cov.merge_cells("B7:D7")

        # Separator
        for c in range(2, 5):
            ws_cov.cell(row=9, column=c).fill = _fill(C_SAGE)
        ws_cov.row_dimensions[9].height = 3

        # Metadata table (rows 11-18)
        meta = [
            ("Document ID",    _doc_id),
            ("Report Date",    _date_str),
            ("Modality",       _modality),
            ("Farm",           _farm_name),
            ("Primary Crop",   _crop),
            ("Country",        _country),
            ("Currency",       "USD ($)"),
            ("Model Version",  "AgriPortal V2"),
        ]
        for i, (k, v) in enumerate(meta):
            row = 11 + i
            ws_cov.row_dimensions[row].height = 16
            _cell(ws_cov, row, 2, k, bold=True, size=9, color=C_INK2)
            _cell(ws_cov, row, 3, v, size=10, color=C_INK)

        # Disclaimer block (rows 22-28)
        ws_cov.row_dimensions[22].height = 14
        _cell(ws_cov, 22, 2, "DISCLAIMER", bold=True, size=9, color=C_CLAY)
        disclaimer = (
            "This workbook is generated by the Agricultural Intelligence Portal and is "
            "intended for indicative feasibility analysis only. All figures are model outputs "
            "based on the assumptions and parameters entered by the user. This document does "
            "not constitute investment advice, a prospectus, or a financial projection for "
            "fundraising purposes. Independent technical, financial, and legal due diligence "
            "is required before making any investment decision."
        )
        dc = ws_cov.cell(row=23, column=2, value=disclaimer)
        dc.font      = _font(size=9, color=C_INK3, italic=True)
        dc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws_cov.merge_cells("B23:D28")
        ws_cov.row_dimensions[23].height = 72

        # Sheet index (rows 32 onward)
        _cell(ws_cov, 32, 2, "WORKBOOK CONTENTS", bold=True, size=9, color=C_INK2)
        sheets_index = [
            ("Dashboard",    "Executive KPI summary — key metrics at a glance"),
            ("Inputs",       "All model parameters as entered (audit trail)"),
            ("P&L",          "Annual profit & loss statement, cost structure"),
            ("CAPEX & Debt", "Capital expenditure breakdown and debt schedule"),
            ("DCF",          "10-year discounted cash flow and IRR"),
        ]
        for i, (sname, sdesc) in enumerate(sheets_index):
            row = 34 + i
            ws_cov.row_dimensions[row].height = 15
            _cell(ws_cov, row, 2, sname, bold=True, size=9, color=C_SAGE)
            _cell(ws_cov, row, 3, sdesc, size=9, color=C_INK2)

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 2 — DASHBOARD
        # ══════════════════════════════════════════════════════════════════════
        ws_db = wb.create_sheet("Dashboard")
        ws_db.sheet_view.showGridLines = False
        for col, w in zip("ABCDEFGH", [2, 22, 18, 18, 18, 18, 18, 2]):
            _set_col_width(ws_db, col, w)

        # Accent column
        for r_idx in range(1, 60):
            ws_db.cell(row=r_idx, column=1).fill = _fill(C_SAGE)
            ws_db.row_dimensions[r_idx].height = 16

        # Title
        ws_db.row_dimensions[3].height = 28
        _cell(ws_db, 3, 2, "EXECUTIVE DASHBOARD", bold=True, size=16, color=C_INK)
        ws_db.merge_cells("B3:G3")
        ws_db.row_dimensions[4].height = 14
        _cell(ws_db, 4, 2, f"{_farm_name}  ·  {_crop}  ·  {_country}  ·  {_date_str}",
              size=9, color=C_INK3, italic=True)
        ws_db.merge_cells("B4:G4")

        # Viability signal
        _energy_pct = r.get("annual_energy_cost", 0) / r.get("annual_revenue", 1) * 100 if r.get("annual_revenue") else 0
        if _energy_pct > 60:
            _viab_label = "STRUCTURALLY CHALLENGED"
            _viab_bg    = C_CLAY
        elif _energy_pct > 35:
            _viab_label = "MARGINAL — REVIEW ASSUMPTIONS"
            _viab_bg    = C_AMBER
        else:
            _viab_label = "VIABLE"
            _viab_bg    = C_SAGE

        ws_db.row_dimensions[6].height = 20
        vc = ws_db.cell(row=6, column=2,
                        value=f"Viability Signal: {_viab_label}  |  Energy = {_energy_pct:.1f}% of Revenue")
        vc.font      = _font(bold=True, size=10, color=C_WHITE)
        vc.fill      = _fill(_viab_bg)
        vc.alignment = _align(h="center", v="center")
        ws_db.merge_cells("B6:G6")

        # KPI tiles — row 8-10 (Revenue, EBITDA, EBITDA Margin, CAPEX)
        #            row 12-14 (Payback, DSCR, Energy%, NPV)
        def _kpi_block(ws, start_row, start_col, label, value_str, sub=""):
            ws.row_dimensions[start_row].height   = 13
            ws.row_dimensions[start_row+1].height = 28
            ws.row_dimensions[start_row+2].height = 13
            lc = ws.cell(row=start_row, column=start_col, value=label.upper())
            lc.font      = _font(size=7.5, color=C_INK3, bold=True, name="Consolas")
            lc.alignment = _align(h="center")
            lc.fill      = _fill(C_LINEN2)
            vc2 = ws.cell(row=start_row+1, column=start_col, value=value_str)
            vc2.font      = _font(size=18, bold=True, color=C_INK, name="Consolas")
            vc2.alignment = _align(h="center")
            vc2.fill      = _fill(C_LINEN2)
            sc = ws.cell(row=start_row+2, column=start_col, value=sub)
            sc.font      = _font(size=7.5, color=C_INK3)
            sc.alignment = _align(h="center")
            sc.fill      = _fill(C_LINEN2)
            for r2 in range(start_row, start_row+3):
                ws.cell(row=r2, column=start_col).border = _border(color=C_RULE)

        _payback_str = f"{r['payback_years']:.1f} yrs" if r.get("payback_years") else "N/A"
        _dscr_str    = f"{r['dscr']:.2f}×" if r.get("dscr") else "N/A"
        _npv_val     = r.get("npv", 0)

        kpis_r1 = [
            ("Annual Revenue",  f"${r.get('annual_revenue',0)/1e3:.0f}K",    "total farm revenue"),
            ("EBITDA",          f"${r.get('ebitda',0)/1e3:.0f}K",           "earnings before int/tax/d&a"),
            ("EBITDA Margin",   f"{r.get('ebitda_margin',0)*100:.1f}%",      "ebitda / revenue"),
            ("Total CAPEX",     f"${r.get('total_capex',0)/1e3:.0f}K",      "total capital expenditure"),
            ("Payback Period",  _payback_str,                                "years to recover equity"),
            ("DSCR",            _dscr_str,                                   "debt service coverage ratio"),
        ]
        for i, (lbl, val, sub) in enumerate(kpis_r1):
            _kpi_block(ws_db, 8, 2 + i, lbl, val, sub)

        # Row 2 of KPIs
        kpis_r2 = [
            ("Energy % Rev",    f"{_energy_pct:.1f}%",                       "key viability indicator"),
            ("NPV (10yr)",      f"${_npv_val/1e3:.0f}K",                    "net present value"),
            ("Annual kg",       f"{r.get('total_annual_kg',0):,.0f}",        "total production"),
            ("Price ($/kg)",    f"${r.get('effective_price',0):.2f}",        "effective selling price"),
            ("Grow Area (m²)",  f"{r.get('effective_grow_area',0):,.0f}",    "net canopy area"),
            ("Cycles/yr",       f"{r.get('cycles_per_year',0):.1f}",         "production cycles per year"),
        ]
        for i, (lbl, val, sub) in enumerate(kpis_r2):
            _kpi_block(ws_db, 13, 2 + i, lbl, val, sub)

        # Cost structure summary (rows 18 onward)
        ws_db.row_dimensions[18].height = 6
        _section_hdr(ws_db, 19, 2, 7, "Annual Cost Structure")
        ws_db.row_dimensions[20].height = 14

        cost_items = [
            ("Energy",      r.get("annual_energy_cost",   0)),
            ("Labour",      r.get("annual_labour_cost",   0)),
            ("Nutrients/Var",r.get("annual_variable_cost", 0)),
            ("Water",       r.get("annual_water_cost",    0)),
            ("Maintenance", r.get("annual_maintenance",   0)),
            ("Rent",        r.get("annual_rent",          0)),
        ]
        _hdr(ws_db, 20, 2, "Cost Item",   bg=C_LINEN, fg=C_INK2, size=9, bold=True)
        _hdr(ws_db, 20, 3, "Annual ($)",  bg=C_LINEN, fg=C_INK2, size=9, bold=True, align="right")
        _hdr(ws_db, 20, 4, "% of Revenue",bg=C_LINEN, fg=C_INK2, size=9, bold=True, align="right")
        _hdr(ws_db, 20, 5, "% of Total Costs",bg=C_LINEN,fg=C_INK2,size=9,bold=True,align="right")
        ws_db.merge_cells("E20:G20")

        _rev  = r.get("annual_revenue", 1) or 1
        _tcost= r.get("total_annual_costs", 1) or 1
        for i, (name, val) in enumerate(cost_items):
            row = 21 + i
            ws_db.row_dimensions[row].height = 15
            bg = C_LINEN2 if i % 2 == 0 else C_WHITE
            _cell(ws_db, row, 2, name, size=9, bg=bg)
            _num( ws_db, row, 3, val,  bg=bg, fmt='$#,##0')
            _num( ws_db, row, 4, val/_rev*100,  bg=bg, fmt='0.0"%"')
            _num( ws_db, row, 5, val/_tcost*100,bg=bg, fmt='0.0"%"')
            ws_db.merge_cells(f"E{row}:G{row}")

        # Total row
        row = 21 + len(cost_items)
        ws_db.row_dimensions[row].height = 16
        _cell(ws_db, row, 2, "TOTAL COSTS", bold=True, size=9, bg=C_LINEN)
        _num( ws_db, row, 3, r.get("total_annual_costs",0), bold=True, bg=C_LINEN, fmt='$#,##0')
        _num( ws_db, row, 4, r.get("total_annual_costs",0)/_rev*100, bold=True, bg=C_LINEN, fmt='0.0"%"')
        _cell(ws_db, row, 5, "", bg=C_LINEN)
        ws_db.merge_cells(f"E{row}:G{row}")

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 3 — INPUTS (audit trail)
        # ══════════════════════════════════════════════════════════════════════
        ws_inp = wb.create_sheet("Inputs")
        ws_inp.sheet_view.showGridLines = False
        for col, w in zip("ABCD", [2, 36, 22, 20]):
            _set_col_width(ws_inp, col, w)
        for r_idx in range(1, 80):
            ws_inp.cell(row=r_idx, column=1).fill = _fill(C_SAGE)
            ws_inp.row_dimensions[r_idx].height = 15

        ws_inp.row_dimensions[3].height = 24
        _cell(ws_inp, 3, 2, "MODEL INPUTS — AUDIT TRAIL", bold=True, size=14, color=C_INK)
        ws_inp.merge_cells("B3:D3")
        _cell(ws_inp, 4, 2, f"Snapshot: {_date_str}  ·  {_doc_id}", size=9, color=C_INK3, italic=True)
        ws_inp.merge_cells("B4:D4")

        _hdr(ws_inp, 6, 2, "Parameter", bg=C_SAGE, fg=C_WHITE, size=9)
        _hdr(ws_inp, 6, 3, "Value",     bg=C_SAGE, fg=C_WHITE, size=9, align="right")
        _hdr(ws_inp, 6, 4, "Unit / Note", bg=C_SAGE, fg=C_WHITE, size=9)

        _inp_rows = [
            ("FARM CONFIGURATION", None, None),
            ("Country",              inputs.get("country"),          "—"),
            ("Primary Crop",         inputs.get("crop"),             "—"),
            ("Total Footprint",      inputs.get("footprint"),        "m²"),
            ("Stacking Levels",      inputs.get("levels"),           "—"),
            ("LED Lights Tier",      inputs.get("lights_tier"),      "—"),
            ("HVAC Tier",            inputs.get("hvac"),             "—"),
            ("Automation Level",     inputs.get("automation"),       "Low / Medium / High"),
            ("Harvest Mode",         inputs.get("harvest_mode"),     "Single / Continuous"),
            ("PRODUCTION PARAMETERS", None, None),
            ("Net Grow Factor",      inputs.get("net_grow_factor"),  "% of gross area"),
            ("Walkways Factor",      inputs.get("walkways_factor"),  "% deducted"),
            ("Loss Rate",            inputs.get("loss_rate"),        "%"),
            ("Price Scenario",       inputs.get("price_scenario"),   "base / high / custom"),
            ("Price Override",       inputs.get("price_override"),   "$/kg (0 = auto)"),
            ("Packaging Cost",       inputs.get("packaging_cost"),   "$/kg"),
            ("Water Price",          inputs.get("water_price"),      "$/m³"),
            ("FINANCIAL STRUCTURE", None, None),
            ("Monthly Rent",         inputs.get("rent_monthly"),     "$/month"),
            ("Real Estate CAPEX",    inputs.get("real_estate_capex"),"$"),
            ("Depreciation Period",  inputs.get("depreciation_years"),"years"),
            ("Tax Rate",             inputs.get("tax_rate"),         "%"),
            ("LTV (Loan-to-Value)",  inputs.get("ltv"),              "%"),
            ("Interest Rate",        inputs.get("interest_rate"),    "%"),
            ("Loan Term",            inputs.get("loan_term_years"),  "years"),
            ("Discount Rate (DCF)",  inputs.get("discount_rate", 8.0), "%"),
        ]

        data_row = 7
        for param, val, unit in _inp_rows:
            if unit is None:
                # Section header
                _section_hdr(ws_inp, data_row, 2, 4, param)
            else:
                bg = C_LINEN2 if data_row % 2 == 0 else C_WHITE
                _cell(ws_inp, data_row, 2, param, size=9, bg=bg)
                if isinstance(val, float) and val == int(val):
                    val = int(val)
                _cell(ws_inp, data_row, 3, str(val) if val is not None else "—",
                      size=9, bg=bg, align="right")
                _cell(ws_inp, data_row, 4, unit, size=9, color=C_INK3, bg=bg, italic=True)
            data_row += 1

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 4 — P&L
        # ══════════════════════════════════════════════════════════════════════
        ws_pl = wb.create_sheet("P&L")
        ws_pl.sheet_view.showGridLines = False
        for col, w in zip("ABCDE", [2, 38, 20, 20, 2]):
            _set_col_width(ws_pl, col, w)
        for r_idx in range(1, 65):
            ws_pl.cell(row=r_idx, column=1).fill = _fill(C_SAGE)
            ws_pl.row_dimensions[r_idx].height = 15

        ws_pl.row_dimensions[3].height = 24
        _cell(ws_pl, 3, 2, "ANNUAL PROFIT & LOSS STATEMENT", bold=True, size=14, color=C_INK)
        ws_pl.merge_cells("B3:D3")
        _cell(ws_pl, 4, 2, f"{_farm_name}  ·  Year 1 Steady-State  ·  {_date_str}", size=9, color=C_INK3, italic=True)
        ws_pl.merge_cells("B4:D4")

        _hdr(ws_pl, 6, 2, "Line Item",      bg=C_SAGE, fg=C_WHITE, size=9)
        _hdr(ws_pl, 6, 3, "Annual ($)",     bg=C_SAGE, fg=C_WHITE, size=9, align="right")
        _hdr(ws_pl, 6, 4, "% of Revenue",   bg=C_SAGE, fg=C_WHITE, size=9, align="right")

        _rev2  = r.get("annual_revenue", 0) or 1
        _pl_rows = [
            # (label, value, bold, is_section, bg_override, note)
            ("REVENUE",                           None,  True,  True,  C_SAGE_LIGHT, None),
            ("Gross Revenue",                     _rev2, False, False, None, None),
            ("Less: Post-harvest Loss",           -(_rev2 * inputs.get("loss_rate",5)/100), False, False, None, f"{inputs.get('loss_rate',5):.1f}% of gross revenue"),
            ("Net Revenue",                       r.get("annual_revenue",0), True, False, C_LINEN, None),
            ("",                                  None,  False, False, None, None),
            ("OPERATING COSTS",                   None,  True,  True,  C_SAGE_LIGHT, None),
            ("Energy",                            r.get("annual_energy_cost",0),   False, False, None, None),
            ("Labour",                            r.get("annual_labour_cost",0),   False, False, None, None),
            ("Nutrients & Variable Costs",        r.get("annual_variable_cost",0), False, False, None, None),
            ("Water",                             r.get("annual_water_cost",0),    False, False, None, None),
            ("Maintenance",                       r.get("annual_maintenance",0),   False, False, None, None),
            ("Rent",                              r.get("annual_rent",0),          False, False, None, None),
            ("Total Operating Costs",             r.get("total_annual_costs",0),   True,  False, C_LINEN, None),
            ("",                                  None,  False, False, None, None),
            ("EBITDA",                            r.get("ebitda",0),               True,  True,  C_SAGE_LIGHT, None),
            ("",                                  None,  False, False, None, None),
            ("NON-CASH & FINANCING",              None,  True,  True,  C_SAGE_LIGHT, None),
            ("Depreciation & Amortisation",       -r.get("annual_depreciation",0), False, False, None, None),
            ("EBIT (Operating Income)",           r.get("ebit",0),                 True,  False, C_LINEN, None),
            ("Interest Expense",                  -(r.get("annual_debt_service",0) - r.get("annual_depreciation",0) * 0), False, False, None, "approx. — see CAPEX & Debt for schedule"),
            ("EBT (Pre-Tax Income)",              r.get("ebit",0),                 True,  False, C_LINEN, None),
            ("Income Tax",                        -r.get("tax_charge",0),          False, False, None, f"{inputs.get('tax_rate',25):.0f}% of EBT"),
            ("NET INCOME",                        r.get("net_income",0),           True,  True,  C_SAGE_LIGHT if r.get("net_income",0)>=0 else C_RED_LIGHT, None),
        ]

        pl_row = 7
        for lbl, val, bold, is_sect, bg_ov, note in _pl_rows:
            ws_pl.row_dimensions[pl_row].height = 15 if not is_sect else 17
            if lbl == "":
                pl_row += 1
                continue
            if is_sect and val is None:
                # Section header row
                bg = bg_ov or C_SAGE_LIGHT
                lc2 = ws_pl.cell(row=pl_row, column=2, value=lbl)
                lc2.font      = _font(bold=True, size=9, color=C_SAGE)
                lc2.fill      = _fill(bg)
                lc2.alignment = _align(h="left")
                ws_pl.cell(row=pl_row, column=3).fill = _fill(bg)
                ws_pl.cell(row=pl_row, column=4).fill = _fill(bg)
                pl_row += 1
                continue

            bg = bg_ov or (C_LINEN2 if pl_row % 2 == 0 else C_WHITE)
            color = C_CLAY if (isinstance(val, (int,float)) and val < 0) else C_INK

            _cell(ws_pl, pl_row, 2, lbl, bold=bold, size=9, bg=bg, color=C_INK)
            if val is not None:
                nc = ws_pl.cell(row=pl_row, column=3, value=round(val, 0))
                nc.font          = _font(bold=bold, size=10, color=color, name="Consolas")
                nc.alignment     = _align(h="right", v="center")
                nc.number_format = '$#,##0;[Red]-$#,##0'
                nc.fill          = _fill(bg)

                pct_val = val / _rev2 * 100 if _rev2 else 0
                pc = ws_pl.cell(row=pl_row, column=4, value=round(pct_val, 1))
                pc.font          = _font(size=9, color=C_INK3, name="Consolas")
                pc.alignment     = _align(h="right")
                pc.number_format = '0.0"%"'
                pc.fill          = _fill(bg)
            else:
                ws_pl.cell(row=pl_row, column=3).fill = _fill(bg)
                ws_pl.cell(row=pl_row, column=4).fill = _fill(bg)

            if note:
                _cell(ws_pl, pl_row, 4, f"  {note}", size=8, color=C_INK3, italic=True, bg=bg)

            pl_row += 1

        # Unit economics block
        pl_row += 1
        _section_hdr(ws_pl, pl_row, 2, 4, "Unit Economics")
        pl_row += 1
        _total_kg = r.get("total_annual_kg", 0) or 1
        _ega      = r.get("effective_grow_area", 0) or 1
        unit_rows = [
            ("Revenue per m²",          r.get("annual_revenue",0)     / _ega,    "$/m²/yr"),
            ("EBITDA per m²",           r.get("ebitda",0)             / _ega,    "$/m²/yr"),
            ("Total Cost per kg",       r.get("total_annual_costs",0) / _total_kg, "$/kg"),
            ("Energy Cost per kg",      r.get("annual_energy_cost",0) / _total_kg, "$/kg"),
            ("Labour Cost per kg",      r.get("annual_labour_cost",0) / _total_kg, "$/kg"),
            ("Annual Production",       r.get("total_annual_kg",0),               "kg/yr"),
            ("kWh per kg",              r.get("total_annual_kwh",r.get("annual_kwh",0)) / _total_kg, "kWh/kg"),
        ]
        for lbl, val, unit in unit_rows:
            bg = C_LINEN2 if pl_row % 2 == 0 else C_WHITE
            _cell(ws_pl, pl_row, 2, lbl, size=9, bg=bg)
            nc2 = ws_pl.cell(row=pl_row, column=3, value=round(val, 2))
            nc2.font          = _font(size=10, name="Consolas")
            nc2.alignment     = _align(h="right")
            nc2.number_format = '#,##0.00'
            nc2.fill          = _fill(bg)
            _cell(ws_pl, pl_row, 4, unit, size=8, color=C_INK3, italic=True, bg=bg)
            pl_row += 1

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 5 — CAPEX & DEBT
        # ══════════════════════════════════════════════════════════════════════
        ws_cx = wb.create_sheet("CAPEX & Debt")
        ws_cx.sheet_view.showGridLines = False
        for col, w in zip("ABCDE", [2, 32, 20, 20, 2]):
            _set_col_width(ws_cx, col, w)
        for r_idx in range(1, 80):
            ws_cx.cell(row=r_idx, column=1).fill = _fill(C_SAGE)
            ws_cx.row_dimensions[r_idx].height = 15

        ws_cx.row_dimensions[3].height = 24
        _cell(ws_cx, 3, 2, "CAPEX BREAKDOWN & DEBT SCHEDULE", bold=True, size=14, color=C_INK)
        ws_cx.merge_cells("B3:D3")
        _cell(ws_cx, 4, 2, f"{_farm_name}  ·  {_date_str}", size=9, color=C_INK3, italic=True)
        ws_cx.merge_cells("B4:D4")

        _hdr(ws_cx, 6, 2, "CAPEX Component",  bg=C_SAGE, fg=C_WHITE, size=9)
        _hdr(ws_cx, 6, 3, "Amount ($)",        bg=C_SAGE, fg=C_WHITE, size=9, align="right")
        _hdr(ws_cx, 6, 4, "% of Total CAPEX",  bg=C_SAGE, fg=C_WHITE, size=9, align="right")

        _total_cx = r.get("total_capex", 0) or 1
        capex_items = [
            ("LED Lighting",          r.get("led_capex",0)),
            ("HVAC",                  r.get("hvac_capex",0)),
            ("Racking",               r.get("racks_capex",0)),
            ("Building & Enclosure",  r.get("building_capex",0)),
            ("Automation & Controls", r.get("automation_capex",0)),
            ("Robotics",              r.get("robotics_capex",0)),
            ("Electrical",            r.get("electrical_capex",0)),
            ("Water & Irrigation",    r.get("water_capex",0)),
            ("Installation",          r.get("installation_capex",0)),
            ("Real Estate",           inputs.get("real_estate_capex",0)),
        ]
        cx_row = 7
        for lbl, val in capex_items:
            bg = C_LINEN2 if cx_row % 2 == 0 else C_WHITE
            _cell(ws_cx, cx_row, 2, lbl, size=9, bg=bg)
            nc3 = ws_cx.cell(row=cx_row, column=3, value=round(val,0))
            nc3.font = _font(size=10, name="Consolas"); nc3.alignment = _align(h="right")
            nc3.number_format = '$#,##0'; nc3.fill = _fill(bg)
            pc3 = ws_cx.cell(row=cx_row, column=4, value=round(val/_total_cx*100,1))
            pc3.font = _font(size=9, color=C_INK3, name="Consolas"); pc3.alignment = _align(h="right")
            pc3.number_format = '0.0"%"'; pc3.fill = _fill(bg)
            cx_row += 1

        # Total CAPEX
        _cell(ws_cx, cx_row, 2, "TOTAL CAPEX", bold=True, size=9, bg=C_LINEN)
        nc4 = ws_cx.cell(row=cx_row, column=3, value=round(r.get("total_capex",0),0))
        nc4.font=_font(bold=True,size=10,name="Consolas"); nc4.alignment=_align(h="right")
        nc4.number_format='$#,##0'; nc4.fill=_fill(C_LINEN)
        nc4b=ws_cx.cell(row=cx_row,column=4,value=100.0)
        nc4b.font=_font(bold=True,size=9,name="Consolas"); nc4b.alignment=_align(h="right")
        nc4b.number_format='0.0"%"'; nc4b.fill=_fill(C_LINEN)
        cx_row += 2

        # Financing summary
        _section_hdr(ws_cx, cx_row, 2, 4, "Financing Structure")
        cx_row += 1
        _debt     = r.get("debt_amount", 0)
        _equity   = r.get("equity_invested", r.get("total_capex",0) * (1 - inputs.get("ltv",0)/100))
        _ann_ds   = r.get("annual_debt_service", 0)
        _ann_depr = r.get("annual_depreciation", 0)
        fin_rows = [
            ("Total CAPEX",        r.get("total_capex",0),  "$"),
            ("Debt (LTV)",         _debt,                   f"$  —  {inputs.get('ltv',0):.0f}% LTV"),
            ("Equity Invested",    _equity,                 "$"),
            ("Interest Rate",      inputs.get("interest_rate",0), "%"),
            ("Loan Term",          inputs.get("loan_term_years",0), "years"),
            ("Annual Debt Service",_ann_ds,                 "$/yr"),
            ("Annual Depreciation",_ann_depr,               "$/yr  — straight-line"),
            ("DSCR",               r.get("dscr") or 0,      "× (min 1.25 recommended)"),
        ]
        for lbl, val, unit in fin_rows:
            bg = C_LINEN2 if cx_row % 2 == 0 else C_WHITE
            _cell(ws_cx, cx_row, 2, lbl, size=9, bg=bg)
            nf = ws_cx.cell(row=cx_row, column=3, value=round(val,2))
            nf.font=_font(size=10,name="Consolas"); nf.alignment=_align(h="right")
            nf.number_format='#,##0.00'; nf.fill=_fill(bg)
            _cell(ws_cx, cx_row, 4, unit, size=8, color=C_INK3, italic=True, bg=bg)
            cx_row += 1

        # Depreciation schedule (straight-line, 10 years)
        cx_row += 1
        _section_hdr(ws_cx, cx_row, 2, 4, "Depreciation Schedule (Straight-Line)")
        cx_row += 1
        _hdr(ws_cx, cx_row, 2, "Year", bg=C_LINEN, fg=C_INK2, size=9)
        _hdr(ws_cx, cx_row, 3, "Annual Depreciation ($)", bg=C_LINEN, fg=C_INK2, size=9, align="right")
        _hdr(ws_cx, cx_row, 4, "Cumulative Depreciated ($)", bg=C_LINEN, fg=C_INK2, size=9, align="right")
        cx_row += 1
        _dep_yrs = inputs.get("depreciation_years", 10)
        for yr in range(1, min(_dep_yrs, 10) + 1):
            bg = C_LINEN2 if cx_row % 2 == 0 else C_WHITE
            _cell(ws_cx, cx_row, 2, f"Year {yr}", size=9, bg=bg)
            nd = ws_cx.cell(row=cx_row, column=3, value=round(_ann_depr,0))
            nd.font=_font(size=10,name="Consolas"); nd.alignment=_align(h="right")
            nd.number_format='$#,##0'; nd.fill=_fill(bg)
            nd2 = ws_cx.cell(row=cx_row, column=4, value=round(_ann_depr*yr,0))
            nd2.font=_font(size=10,name="Consolas"); nd2.alignment=_align(h="right")
            nd2.number_format='$#,##0'; nd2.fill=_fill(bg)
            cx_row += 1

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 6 — DCF
        # ══════════════════════════════════════════════════════════════════════
        ws_dcf = wb.create_sheet("DCF")
        ws_dcf.sheet_view.showGridLines = False
        dcf_cols = ["A","B","C","D","E","F","G","H","I"]
        for col, w in zip(dcf_cols, [2, 10, 16, 16, 16, 16, 16, 16, 2]):
            _set_col_width(ws_dcf, col, w)
        for r_idx in range(1, 30):
            ws_dcf.cell(row=r_idx, column=1).fill = _fill(C_SAGE)
            ws_dcf.row_dimensions[r_idx].height = 15

        ws_dcf.row_dimensions[3].height = 24
        _cell(ws_dcf, 3, 2, "10-YEAR DISCOUNTED CASH FLOW", bold=True, size=14, color=C_INK)
        ws_dcf.merge_cells("B3:H3")
        _cell(ws_dcf, 4, 2,
              f"Discount rate: {inputs.get('discount_rate',8.0):.1f}%  ·  Equity: ${_equity:,.0f}  ·  {_date_str}",
              size=9, color=C_INK3, italic=True)
        ws_dcf.merge_cells("B4:H4")

        # Headers
        dcf_hdrs = ["Year", "FCFE ($)", "Discount Factor", "PV of FCFE ($)", "Cumulative NPV ($)", "Revenue ($)", "EBITDA ($)"]
        for i, lbl in enumerate(dcf_hdrs):
            _hdr(ws_dcf, 6, 2+i, lbl, bg=C_SAGE, fg=C_WHITE, size=9,
                 align="right" if i > 0 else "left")

        # Year 0 (equity outflow)
        ws_dcf.row_dimensions[7].height = 15
        _cell(ws_dcf, 7, 2, "0 (Investment)", size=9, color=C_INK, bg=C_LINEN2)
        _num(ws_dcf, 7, 3, -round(_equity,0), bg=C_LINEN2, fmt='$#,##0;[Red]-$#,##0', negative_red=True)
        _num(ws_dcf, 7, 4, 1.000, bg=C_LINEN2, fmt='0.000')
        _num(ws_dcf, 7, 5, -round(_equity,0), bg=C_LINEN2, fmt='$#,##0;[Red]-$#,##0', negative_red=True)
        _num(ws_dcf, 7, 6, -round(_equity,0), bg=C_LINEN2, fmt='$#,##0;[Red]-$#,##0', negative_red=True)
        _cell(ws_dcf, 7, 7, "—", size=9, color=C_INK3, bg=C_LINEN2, align="right")
        _cell(ws_dcf, 7, 8, "—", size=9, color=C_INK3, bg=C_LINEN2, align="right")

        _dcf_cashflows = r.get("dcf_cashflows", [])
        for i, yr_data in enumerate(_dcf_cashflows):
            row = 8 + i
            ws_dcf.row_dimensions[row].height = 15
            yr    = yr_data.get("year", i+1)
            fcfe  = yr_data.get("fcfe", 0)
            pv    = yr_data.get("pv", 0)
            cum   = yr_data.get("cumulative_npv", 0)
            _disc_rt = inputs.get("discount_rate", 8.0) / 100
            df    = 1 / ((1 + _disc_rt) ** yr)

            bg = C_LINEN2 if row % 2 == 0 else C_WHITE
            _cum_color = C_SAGE if cum >= 0 else C_CLAY
            _cell(ws_dcf, row, 2, f"Year {yr}", size=9, bg=bg)
            _num(ws_dcf, row, 3, round(fcfe,0),  bg=bg, fmt='$#,##0;[Red]-$#,##0')
            _num(ws_dcf, row, 4, round(df,4),    bg=bg, fmt='0.0000')
            _num(ws_dcf, row, 5, round(pv,0),    bg=bg, fmt='$#,##0;[Red]-$#,##0')
            nc_cum = ws_dcf.cell(row=row, column=6, value=round(cum,0))
            nc_cum.font=_font(bold=(cum>=0), size=10, color=_cum_color, name="Consolas")
            nc_cum.alignment=_align(h="right"); nc_cum.number_format='$#,##0;[Red]-$#,##0'
            nc_cum.fill=_fill(bg)
            _num(ws_dcf, row, 7, round(r.get("annual_revenue",0),0), bg=bg, fmt='$#,##0')
            _num(ws_dcf, row, 8, round(r.get("ebitda",0),0),         bg=bg, fmt='$#,##0')

        # Summary metrics at bottom
        _summary_row = 8 + len(_dcf_cashflows) + 2
        _section_hdr(ws_dcf, _summary_row, 2, 8, "Investment Return Summary")
        _summary_row += 1
        _npv_final = _dcf_cashflows[-1]["cumulative_npv"] if _dcf_cashflows else 0

        # IRR approximation (Newton-Raphson)
        def _calc_irr(equity, annual_fcfe, n=10):
            try:
                cashflows = [-equity] + [annual_fcfe] * n
                rate = 0.1
                for _ in range(200):
                    npv = sum(cf / (1+rate)**t for t, cf in enumerate(cashflows))
                    d_npv = sum(-t*cf / (1+rate)**(t+1) for t, cf in enumerate(cashflows))
                    if abs(d_npv) < 1e-10: break
                    rate -= npv / d_npv
                    if rate <= -1: return None
                return rate if 0 < rate < 10 else None
            except Exception:
                return None

        _fcfe_val = _dcf_cashflows[0]["fcfe"] if _dcf_cashflows else r.get("ebitda",0)
        _irr      = _calc_irr(_equity, _fcfe_val)

        summary_kpis = [
            ("NPV (10yr, equity basis)",       f"${_npv_final:,.0f}"),
            ("IRR (approx.)",                  f"{_irr*100:.1f}%" if _irr else "N/A"),
            ("Payback Period",                 f"{r['payback_years']:.1f} yrs" if r.get("payback_years") else "N/A"),
            ("DSCR",                           f"{r['dscr']:.2f}×" if r.get("dscr") else "N/A"),
            ("Equity Invested",                f"${_equity:,.0f}"),
            ("Discount Rate",                  f"{inputs.get('discount_rate',8.0):.1f}%"),
        ]
        for lbl, val in summary_kpis:
            bg = C_LINEN2 if _summary_row % 2 == 0 else C_WHITE
            _cell(ws_dcf, _summary_row, 2, lbl, size=9, bg=bg)
            vc3 = ws_dcf.cell(row=_summary_row, column=3, value=val)
            vc3.font=_font(bold=True, size=10, name="Consolas"); vc3.alignment=_align(h="right")
            vc3.fill=_fill(bg)
            ws_dcf.merge_cells(
                start_row=_summary_row, start_column=3,
                end_row=_summary_row, end_column=8
            )
            _summary_row += 1

        # ── Freeze panes & set active sheet ───────────────────────────────────
        for ws in [ws_db, ws_inp, ws_pl, ws_cx, ws_dcf]:
            ws.freeze_panes = ws.cell(row=7, column=2)

        wb.active = ws_cov  # open on Cover

        # ── Write to buffer ───────────────────────────────────────────────────
        _buf = io.BytesIO()
        wb.save(_buf)
        _buf.seek(0)
        return _buf.read()

    # ── Key metrics ───────────────────────────────────────────────────────────────
    # ── PDF Download Button ───────────────────────────────────────────────────
    pdf_col1, pdf_col2, pdf_col3 = st.columns([4, 1, 1])
    with pdf_col2:
        if st.button("📄 Download PDF Report", use_container_width=True):
            with st.spinner("Generating PDF..."):
                pdf_bytes = generate_pdf_report(inputs, r)
                # Correct naming: use primary crop from mix if multi-crop is valid
                _rep_crop = _crop_mix[0]["crop"] if _mix_valid and _crop_mix else inputs["crop"]
                _rep_crop_safe = _rep_crop.replace(' ', '_').replace('/', '')
                filename = f"CEA_Report_{_rep_crop_safe}_{inputs['country']}_{date.today().strftime('%Y%m%d')}.pdf"
                st.download_button(
                    label="⬇️ Save PDF",
                    data=pdf_bytes,
                    file_name=filename,
                    mime="application/pdf",
                    use_container_width=True,
                )
    with pdf_col3:
        if st.button("📊 Download Excel Model", use_container_width=True):
            with st.spinner("Building Excel workbook..."):
                xl_bytes = generate_excel_report(inputs, r)
                # Correct naming: use primary crop from mix if multi-crop is valid
                _xl_crop = _crop_mix[0]["crop"] if _mix_valid and _crop_mix else inputs["crop"]
                _xl_crop_safe = _xl_crop.replace(' ', '_').replace('/', '')
                xl_filename = f"CEA_Model_{_xl_crop_safe}_{inputs['country']}_{date.today().strftime('%Y%m%d')}.xlsx"
                st.download_button(
                    label="⬇️ Save Excel",
                    data=xl_bytes,
                    file_name=xl_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
    st.divider()
    
    st.subheader("Key Metrics")
    
    # Break-even yield: minimum kg/m²/cycle needed to cover all costs
    # Back-solved from: revenue = total_costs
    # breakeven_yield = total_costs / (price * (1-loss) * cycles * ega)
    _loss = inputs.get("loss_rate", 5) / 100
    _denominator = r["effective_price"] * (1 - _loss) * r["cycles_per_year"] * r["effective_grow_area"]
    breakeven_yield = r["total_annual_costs"] / _denominator if _denominator > 0 else None
    projected_yield = CROPS[crop]["yield"]
    
    if breakeven_yield is not None:
        yield_gap_pct = (projected_yield - breakeven_yield) / breakeven_yield * 100
        yield_gap_str = f"{yield_gap_pct:+.1f}%"
    else:
        yield_gap_str = "N/A"
    
    m1, m2, m3, m4, m5, m6, m7 = st.columns(7)
    m1.metric("Annual Revenue",   f"${r['annual_revenue']:,.0f}")
    m2.metric("EBITDA",           f"${r['ebitda']:,.0f}")
    m3.metric("EBITDA Margin",    f"{r['ebitda_margin']*100:.1f}%")
    m4.metric("Total CAPEX",      f"${r['total_capex']:,.0f}")
    m5.metric("Payback",          f"{r['payback_years']:.1f} yrs" if r["payback_years"] else "N/A")
    m6.metric("DSCR",             f"{r['dscr']:.2f}" if r["dscr"] else "N/A")
    m7.metric(
        "Break-even Yield",
        f"{breakeven_yield:.2f} kg/m²/cycle" if breakeven_yield else "N/A",
        delta=yield_gap_str if breakeven_yield else None,
        delta_color="normal",
        help="Minimum yield needed to cover all costs. Delta = projected crop yield vs this threshold."
    )
    
    # ── Multi-crop breakdown ──────────────────────────────────────────────────
    if r.get("_is_multicrop"):
        st.divider()
        st.subheader("Per-Crop Breakdown")
        _mc_rows = []
        for _mc in r["_crop_results"]:
            _mc_rows.append({
                "Crop":           _mc["crop"],
                "Area %":         f"{_mc['pct']:.0f}%",
                "Annual kg":      f"{_mc['total_annual_kg']:,.0f}",
                "Price ($/kg)":   f"${_mc['effective_price']:.2f}",
                "Revenue":        f"${_mc['annual_revenue']:,.0f}",
                "Variable Cost":  f"${_mc['annual_variable_cost']:,.0f}",
                "Labour":         f"${_mc['annual_labour_cost']:,.0f}",
                "EBITDA contrib": f"${_mc['ebitda']:,.0f}",
            })
        st.dataframe(pd.DataFrame(_mc_rows), use_container_width=True, hide_index=True)
        st.caption(
            "Energy cost and CAPEX are shared farm infrastructure — "
            "computed once on the full farm area and shown in the combined metrics above."
        )
        st.divider()

    # ── DSCR warning ─────────────────────────────────────────────────────────────
    if r.get("dscr") is not None and r["dscr"] < 1.0:
        st.warning(
            f"⚠️ **Debt service coverage is low (DSCR = {r['dscr']:.2f}x).** "
            f"Annual debt repayment (${r['annual_debt_service']:,.0f}) exceeds EBITDA (${r['ebitda']:,.0f}). "
            f"Consider reducing LTV, extending the loan term, increasing farm scale, or selecting a higher-margin crop."
        )

    # ── Save as Farm Profile ──────────────────────────────────────────────────
    st.divider()
    save_col1, save_col2 = st.columns([5, 1]) # Keep emoji in button
    with save_col2:
        if st.button("💾 Save as Farm Profile", use_container_width=True):
            st.session_state["show_save_farm_form"] = True
    
    if st.session_state["show_save_farm_form"]:
        with st.container(border=True):
            _vf_active  = st.session_state.get("active_farm")
            _save_lat = st.session_state.get("shared_lat")
            _save_lon = st.session_state.get("shared_lng")
            _climate_data = {}
            if _save_lat and _save_lon:
                with st.spinner("🌤️ Fetching climate profile for this location…"):
                    try:
                        from core.climate import fetch_climate_profile
                        _climate_data = fetch_climate_profile(_save_lat, _save_lon)
                    except Exception:
                        _climate_data = {}
            _vf_payload = {
                "country":           inputs["country"],
                "crop":              (_crop_mix[0]["crop"] if _mix_valid and _crop_mix else inputs["crop"]),
                "footprint":         inputs["footprint"],
                "levels":            inputs["levels"],
                "lights_tier":       inputs["lights_tier"],
                "hvac":              inputs["hvac"],
                "automation":        inputs["automation"],
                "price_scenario":    inputs["price_scenario"],
                "price_override":    inputs["price_override"],
                "packaging_cost":    inputs["packaging_cost"],
                "loss_rate":         inputs["loss_rate"],
                "net_grow_factor":   inputs["net_grow_factor"],
                "walkways_factor":   inputs["walkways_factor"],
                "water_price":       inputs["water_price"],
                "rent_monthly":      inputs["rent_monthly"],
                "real_estate_capex": inputs["real_estate_capex"],
                "harvest_mode":      inputs["harvest_mode"],
                "depreciation_years": inputs["depreciation_years"],
                "tax_rate":          inputs["tax_rate"],
                "ltv":               inputs["ltv"],
                "interest_rate":     inputs["interest_rate"],
                "loan_term_years":   inputs["loan_term_years"],
                "lat":               st.session_state.get("shared_lat"),
                "lon":               st.session_state.get("shared_lng"),
                "ambient_temp_annual": _climate_data.get("ambient_temp_annual"),
                "mean_annual_dli":     _climate_data.get("mean_annual_dli"),
                "agriculture_type":  "vertical_farm",
                "modality":          "vertical_farm",
                "metadata":          {},
                "model_snapshot":    json.dumps(r),
                "model_updated_at":  date.today().isoformat(),
                "crop_mix_json":     json.dumps(_crop_mix) if _mix_valid else None,
                "notes":             None,
            }
            if _vf_active:
                st.markdown(f"**Update** existing farm **{_vf_active['name']}**, or save as a new profile.")
                _vu1, _vu2, _vu3 = st.columns([2, 2, 1])
                with _vu1:
                    if st.button("✅ Update existing farm", use_container_width=True, key="vf_update_btn"):
                        try: # Keep emoji in success message
                            supabase.table("farms").update(_vf_payload).eq("id", _vf_active["id"]).execute()
                            st.session_state["active_farm"] = {**_vf_active, **_vf_payload}
                            st.success(f"✅ Farm **{_vf_active['name']}** updated.")
                            if _climate_data.get("mean_annual_dli"):
                                st.caption(f"🌤️ Climate profile saved — Mean DLI: {_climate_data['mean_annual_dli']:.1f} mol/m²/day · Ambient temp: {_climate_data['ambient_temp_annual']:.1f}°C")
                            st.session_state["show_save_farm_form"] = False
                            st.rerun()
                        except Exception as e:
                            st.error(f"Update failed: {e}")
                with _vu2:
                    farm_profile_name = st.text_input("New farm name", key="farm_profile_name_input", placeholder="Enter name for new profile")
                    if st.button("➕ Save as new farm", use_container_width=True, key="vf_saveas_btn"):
                        if not farm_profile_name.strip():
                            st.error("Please enter a name for the new farm profile.") # Keep emoji in success message
                        else:
                            try:
                                supabase.table("farms").insert({**_vf_payload, "name": farm_profile_name.strip(), "owner_id": current_user()}).execute()
                                st.success(f"✅ New farm profile '{farm_profile_name.strip()}' saved.")
                                if _climate_data.get("mean_annual_dli"):
                                    st.caption(f"🌤️ Climate profile saved — Mean DLI: {_climate_data['mean_annual_dli']:.1f} mol/m²/day · Ambient temp: {_climate_data['ambient_temp_annual']:.1f}°C")
                                st.session_state["show_save_farm_form"] = False
                                st.rerun()
                            except Exception as e:
                                st.error(f"Could not save: {e}")
                with _vu3:
                    if st.button("✖ Cancel", use_container_width=True, key="vf_cancel_save"):
                        st.session_state["show_save_farm_form"] = False
                        st.rerun()
            else:
                st.markdown("**Save current configuration as a Farm Profile**")
                st.caption("Saves all parameters so you can track harvests in the Harvest Tracker.")
                farm_profile_name = st.text_input("Farm name", key="farm_profile_name_input") # Keep emoji in button
                _vn1, _vn2 = st.columns([3, 1])
                with _vn1:
                    if st.button("✅ Confirm Save", use_container_width=True, key="vf_confirm_save"):
                        if not farm_profile_name.strip():
                            st.error("Please enter a farm name.")
                        else:
                            try:
                                supabase.table("farms").insert({**_vf_payload, "name": farm_profile_name.strip(), "owner_id": current_user()}).execute()
                                st.success(f"Farm profile '{farm_profile_name.strip()}' saved.")
                                if _climate_data.get("mean_annual_dli"):
                                    st.caption(f"🌤️ Climate profile saved — Mean DLI: {_climate_data['mean_annual_dli']:.1f} mol/m²/day · Ambient temp: {_climate_data['ambient_temp_annual']:.1f}°C")
                                st.session_state["show_save_farm_form"] = False
                                st.rerun()
                            except Exception as e:
                                st.error(f"Could not save farm profile: {e}")
                with _vn2:
                    if st.button("✖ Cancel", use_container_width=True, key="vf_cancel_new"):
                        st.session_state["show_save_farm_form"] = False
                        st.rerun()
    
    st.divider()
    
    # ── EBITDA Bridge ─────────────────────────────────────────────────────────
    st.subheader("EBITDA Bridge")
    bridge_labels = ["Revenue", "Variable", "Water", "Energy", "Labour", "Rent", "Maintenance", "EBITDA"]
    bridge_values = [
        r["annual_revenue"],
        -r["annual_variable_cost"],
        -r["annual_water_cost"],
        -r["annual_energy_cost"],
        -r["annual_labour_cost"],
        -r["annual_rent"],
        -r["annual_maintenance"],
        r["ebitda"],
    ]
    bar_colors = []
    for i, v in enumerate(bridge_values):
        if i == 0:
            bar_colors.append("rgba(0,229,160,0.85)")
        elif i == len(bridge_values) - 1:
            bar_colors.append("rgba(0,229,160,0.85)" if v >= 0 else "rgba(255,77,77,0.85)")
        else:
            bar_colors.append("rgba(255,77,77,0.6)")
    
    fig_bridge = go.Figure(go.Bar(
        x=bridge_labels,
        y=bridge_values,
        marker_color=bar_colors,
        text=[f"${abs(v):,.0f}" for v in bridge_values],
        textposition="outside",
    ))
    fig_bridge.update_layout(
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        font_color="#161a16", showlegend=False,
        yaxis=dict(showgrid=False),
        xaxis=dict(showgrid=False),
        height=380, margin=dict(t=30, b=20),
    )
    style_fig(fig_bridge)
    st.plotly_chart(fig_bridge, use_container_width=True)
    
    st.divider()
    
    # ── Cost breakdown + CAPEX breakdown ─────────────────────────────────────────
    col_cost, col_capex = st.columns(2)
    
    with col_cost:
        st.subheader("Annual Cost Breakdown")
        cost_labels = ["Energy", "Labour", "Variable", "Water", "Maintenance", "Rent"]
        cost_values = [
            r["annual_energy_cost"], r["annual_labour_cost"],
            r["annual_variable_cost"], r["annual_water_cost"],
            r["annual_maintenance"], r["annual_rent"],
        ]
        fig_cost = go.Figure(go.Pie(
            labels=cost_labels,
            values=cost_values,
            hole=0.45,
            marker_colors=["#ff4d4d", "#ffc13d", "#00e5a0", "#4fc3f7", "#ba68c8", "#ef9a9a"],
        ))
        fig_cost.update_layout(
            plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            font_color="#161a16", height=320, margin=dict(t=10, b=10),
        )
        style_fig(fig_cost)
        st.plotly_chart(fig_cost, use_container_width=True)
    
    with col_capex:
        st.subheader("CAPEX Breakdown")
        capex_labels = ["LED", "HVAC", "Racks", "Building", "Automation",
                        "Robotics", "Electrical", "Water", "Installation"]
        capex_values = [
            r["led_capex"], r["hvac_capex"], r["racks_capex"], r["building_capex"],
            r["automation_capex"], r["robotics_capex"], r["electrical_capex"],
            r["water_capex"], r["installation_capex"],
        ]
        fig_capex = go.Figure(go.Pie(
            labels=capex_labels,
            values=capex_values,
            hole=0.45,
            marker_colors=["#00e5a0", "#26c6da", "#66bb6a", "#ffa726", "#ab47bc",
                           "#ef5350", "#42a5f5", "#26a69a", "#8d6e63"],
        ))
        fig_capex.update_layout(
            plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            font_color="#161a16", height=320, margin=dict(t=10, b=10),
        )
        style_fig(fig_capex)
        st.plotly_chart(fig_capex, use_container_width=True)
    
    st.divider()
    
    # ── DCF chart ─────────────────────────────────────────────────────────────────
    st.subheader("Cumulative NPV — 10-year DCF")
    dcf_years      = [d["year"] for d in r["dcf_cashflows"]]
    dcf_cumulative = [d["cumulative_npv"] for d in r["dcf_cashflows"]]
    
    fig_dcf = go.Figure()
    fig_dcf.add_trace(go.Scatter(
        x=dcf_years,
        y=dcf_cumulative,
        mode="lines+markers",
        line=dict(color="#00e5a0", width=2),
        fill="tozeroy",
        fillcolor="rgba(0,229,160,0.1)",
    ))
    fig_dcf.add_hline(y=0, line_dash="dash", line_color="rgba(255,255,255,0.3)")
    fig_dcf.update_layout(
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        font_color="#161a16", height=300,
        xaxis=dict(title="Year", showgrid=False),
        yaxis=dict(title="Cumulative NPV ($)", showgrid=False),
        margin=dict(t=10, b=10),
    )
    style_fig(fig_dcf)
    st.plotly_chart(fig_dcf, use_container_width=True)
    
    st.divider()
    
    # ── Full results table ────────────────────────────────────────────────────────
    st.subheader("Full Results")
    
    summary_data = {
        "Metric": [
            "Effective Grow Area (m²)", "Gross Area (m²)",
            "Cycles / Year", "Effective Cycle Days", "Harvest Mode",
            "Total Annual kg", "Price ($/kg)",
            "Annual Revenue", "Energy Cost", "Variable Cost",
            "Water Cost", "Labour Cost", "Maintenance", "Rent",
            "Total Annual Costs", "EBITDA", "EBITDA Margin",
            "Total CAPEX", "Payback (years)",
            "Annual Labour Hours", "Net Income", "NPV (10yr)",
        ],
        "Value": [
            f"{r['effective_grow_area']:,.0f}",
            f"{r['gross_area']:,.0f}",
            str(r["cycles_per_year"]),
            str(r["effective_cycle_days"]),
            r["harvest_mode"],
            f"{r['total_annual_kg']:,.0f}",
            f"${r['effective_price']:.2f}",
            f"${r['annual_revenue']:,.0f}",
            f"${r['annual_energy_cost']:,.0f}",
            f"${r['annual_variable_cost']:,.0f}",
            f"${r['annual_water_cost']:,.0f}",
            f"${r['annual_labour_cost']:,.0f}",
            f"${r['annual_maintenance']:,.0f}",
            f"${r['annual_rent']:,.0f}",
            f"${r['total_annual_costs']:,.0f}",
            f"${r['ebitda']:,.0f}",
            f"{r['ebitda_margin']*100:.1f}%",
            f"${r['total_capex']:,.0f}",
            f"{r['payback_years']:.1f}" if r["payback_years"] else "N/A",
            f"{r['annual_labour_hours']:,.1f}",
            f"${r['net_income']:,.0f}",
            f"${r['npv']:,.0f}",
        ],
    }
    
    def highlight_alternating_rows(row):
        # Apply a subtle background to alternating rows for readability
        return [MATCH if row.name % 2 == 0 else ""] * len(row)

    _summary_df = pd.DataFrame(summary_data)
    st.dataframe(
        _summary_df.style.apply(highlight_alternating_rows, axis=1),
        use_container_width=True, hide_index=True,
    )
    
    # ═════════════════════════════════════════════════════════════════════════════
    # COUNTRY & CROP COMPARISON
    # ═════════════════════════════════════════════════════════════════════════════
    st.divider()
    st.subheader("🌍 Viability Comparison")
    st.caption("All calculations use the current sidebar inputs. Only the dimension being compared changes.")
    
    comp_tab1, comp_tab2 = st.tabs(["Compare Countries", "Compare Crops"])
    
    # ── TAB 1: Country Comparison ─────────────────────────────────────────────────
    with comp_tab1:
        country_metric = st.selectbox(
            "Rank by",
            ["EBITDA", "Energy % of Revenue", "Payback (years)", "EBITDA Margin (%)"],
            key="country_metric_select",
        )
    
        country_results = []
        for c_name in COUNTRIES.keys():
            c_inputs = copy.deepcopy(inputs)
            c_inputs["country"] = c_name
            try:
                c_r = calculate(c_inputs)
                energy_pct = c_r["annual_energy_cost"] / c_r["annual_revenue"] * 100 if c_r["annual_revenue"] > 0 else 999
                country_results.append({
                    "Country":              c_name,
                    "EBITDA":               c_r["ebitda"],
                    "Energy % of Revenue":  energy_pct,
                    "Payback (years)":      c_r["payback_years"] if c_r["payback_years"] else 999,
                    "EBITDA Margin (%)":    c_r["ebitda_margin"] * 100,
                    "revenue":              c_r["annual_revenue"],
                    "energy_cost":          c_r["annual_energy_cost"],
                })
            except Exception:
                continue
    
        if country_results:
            df_countries = pd.DataFrame(country_results)
    
            # Sort: for EBITDA and margin, descending (higher is better).
            # For energy % and payback, ascending (lower is better).
            ascending = country_metric in ("Energy % of Revenue", "Payback (years)")
            df_countries = df_countries.sort_values(country_metric, ascending=ascending).reset_index(drop=True)
    
            def country_bar_color(row, metric):
                if metric == "EBITDA":
                    return "rgba(0,229,160,0.75)" if row["EBITDA"] >= 0 else "rgba(255,77,77,0.75)"
                elif metric == "Energy % of Revenue":
                    return "rgba(0,229,160,0.75)" if row["Energy % of Revenue"] < 40 else "rgba(255,77,77,0.75)"
                elif metric == "Payback (years)":
                    # Payback under 10 years = green, over = red, 999 = no payback = red
                    return "rgba(0,229,160,0.75)" if 0 < row["Payback (years)"] < 10 else "rgba(255,77,77,0.75)"
                elif metric == "EBITDA Margin (%)":
                    return "rgba(0,229,160,0.75)" if row["EBITDA Margin (%)"] >= 0 else "rgba(255,77,77,0.75)"
                return "rgba(105,105,105,0.75)"
    
            bar_colors_country = [
                country_bar_color(row, country_metric)
                for _, row in df_countries.iterrows()
            ]
    
            fig_countries = go.Figure(go.Bar(
                x=df_countries[country_metric],
                y=df_countries["Country"],
                orientation="h",
                marker_color=bar_colors_country,
                text=df_countries[country_metric].apply(
                    lambda v: f"${v:,.0f}" if country_metric == "EBITDA"
                    else (f"{v:.1f}%" if "%" in country_metric
                    else (f"{v:.1f} yrs" if v < 900 else "N/A"))
                ),
                textposition="outside",
            ))
    
            # Add viability threshold line for energy % chart
            if country_metric == "Energy % of Revenue":
                fig_countries.add_vline(
                    x=40,
                    line_dash="dash",
                    line_color="rgba(255,193,61,0.6)",
                    annotation_text="40% threshold",
                    annotation_position="top",
                    annotation_font_color="#ffc13d",
                )
    
            fig_countries.update_layout(
                plot_bgcolor="#ffffff",
                paper_bgcolor="#ffffff",
                font_color="#161a16",
                height=max(500, len(df_countries) * 22),
                margin=dict(l=10, r=100, t=20, b=20),
                xaxis=dict(showgrid=False, zeroline=False),
                yaxis=dict(showgrid=False, autorange="reversed"),
            )
            style_fig(fig_countries)
            st.plotly_chart(fig_countries, use_container_width=True)
    
            # Summary note
            viable = (df_countries["Energy % of Revenue"] < 40).sum()
            st.caption(f"**{viable} of {len(df_countries)} countries** show energy below 40% of revenue for this crop and setup. Green = below threshold, Red = above threshold.")
    
    # ── TAB 2: Crop Comparison ────────────────────────────────────────────────────
    with comp_tab2:
        crop_col1, crop_col2 = st.columns(2)
        with crop_col1:
            crop_metric = st.selectbox(
                "Rank by",
                ["EBITDA", "EBITDA Margin (%)", "Energy % of Revenue", "Payback (years)"],
                key="crop_metric_select",
            )
        with crop_col2:
            # Category filter
            CROP_CATEGORIES = {
                "All crops": None,
                "Leafy greens & lettuce": ["Lettuce", "Leaf ", "Lollo", "Batavia", "Frisee",
                                            "Escarole", "Baby Lettuce", "Spinach", "Baby Spinach",
                                            "Kale", "Baby Kale", "Mizuna", "Arugula", "Tatsoi",
                                            "Pak Choi", "Shanghai", "Mustard", "Mibuna", "Choi Sum",
                                            "Komatsuna", "Swiss Chard", "Purslane", "Sorrel",
                                            "Watercress", "Upland Cress", "Claytonia", "Turnip Greens",
                                            "Broccoli", "Brussels", "Radicchio", "Chicory",
                                            "Endive", "Fennel"],
                "Herbs": ["Basil", "Mint", "Parsley", "Cilantro", "Chives", "Dill", "Oregano",
                          "Thyme", "Sage", "Rosemary", "Lemongrass", "Stevia", "Shiso",
                          "Wasabi", "Celery", "Chinese Celery", "Garlic Chives", "Spring Onions"],
                "Microgreens": ["Microgreens", "Pea Shoots"],
                "Fruiting & other": ["Strawberr", "Tomato", "Pepper", "Chili", "Eggplant",
                                      "Cucumber", "Dwarf Bean", "Okra", "Edible Flower",
                                      "Carrot", "Radish", "Beet", "Micro-Turnip"],
                "Mushrooms": ["Mushroom"],
            }
            selected_category = st.selectbox("Filter by category", list(CROP_CATEGORIES.keys()), key="crop_category_select")
    
        # Apply category filter
        filter_keywords = CROP_CATEGORIES[selected_category]
        if filter_keywords:
            filtered_crops = [c for c in CROPS.keys() if any(kw in c for kw in filter_keywords)]
        else:
            filtered_crops = list(CROPS.keys())
    
        crop_results = []
        for cr_name in filtered_crops:
            cr_inputs = copy.deepcopy(inputs)
            cr_inputs["crop"] = cr_name
            cr_inputs["harvest_mode"] = "Single"   # normalise to single for fair comparison
            cr_inputs["price_override"] = 0        # use each crop's own price
            try:
                cr_r = calculate(cr_inputs)
                energy_pct = cr_r["annual_energy_cost"] / cr_r["annual_revenue"] * 100 if cr_r["annual_revenue"] > 0 else 999
                crop_results.append({
                    "Crop":                 cr_name,
                    "EBITDA":               cr_r["ebitda"],
                    "EBITDA Margin (%)":    cr_r["ebitda_margin"] * 100,
                    "Energy % of Revenue":  energy_pct,
                    "Payback (years)":      cr_r["payback_years"] if cr_r["payback_years"] else 999,
                })
            except Exception:
                continue
    
        if crop_results:
            df_crops = pd.DataFrame(crop_results)
    
            ascending_crop = crop_metric in ("Energy % of Revenue", "Payback (years)")
            df_crops = df_crops.sort_values(crop_metric, ascending=ascending_crop).reset_index(drop=True)
    
            def crop_bar_color(row, metric):
                if metric == "EBITDA":
                    return "rgba(0,229,160,0.75)" if row["EBITDA"] >= 0 else "rgba(255,77,77,0.75)"
                elif metric == "Energy % of Revenue":
                    return "rgba(0,229,160,0.75)" if row["Energy % of Revenue"] < 40 else "rgba(255,77,77,0.75)"
                elif metric == "Payback (years)":
                    return "rgba(0,229,160,0.75)" if 0 < row["Payback (years)"] < 10 else "rgba(255,77,77,0.75)"
                elif metric == "EBITDA Margin (%)":
                    return "rgba(0,229,160,0.75)" if row["EBITDA Margin (%)"] >= 0 else "rgba(255,77,77,0.75)"
                return "rgba(105,105,105,0.75)"
    
            bar_colors_crop = [
                crop_bar_color(row, crop_metric)
                for _, row in df_crops.iterrows()
            ]
    
            fig_crops = go.Figure(go.Bar(
                x=df_crops[crop_metric],
                y=df_crops["Crop"],
                orientation="h",
                marker_color=bar_colors_crop,
                text=df_crops[crop_metric].apply(
                    lambda v: f"${v:,.0f}" if crop_metric == "EBITDA"
                    else (f"{v:.1f}%" if "%" in crop_metric
                    else (f"{v:.1f} yrs" if v < 900 else "N/A"))
                ),
                textposition="outside",
            ))
    
            if crop_metric == "Energy % of Revenue":
                fig_crops.add_vline(
                    x=40,
                    line_dash="dash",
                    line_color="rgba(255,193,61,0.6)",
                    annotation_text="40% threshold",
                    annotation_position="top",
                    annotation_font_color="#ffc13d",
                )
    
            fig_crops.update_layout(
                plot_bgcolor="#ffffff",
                paper_bgcolor="#ffffff",
                font_color="#161a16",
                height=max(500, len(df_crops) * 22),
                margin=dict(l=10, r=100, t=20, b=20),
                xaxis=dict(showgrid=False, zeroline=False),
                yaxis=dict(showgrid=False, autorange="reversed"),
            )
            style_fig(fig_crops)
            st.plotly_chart(fig_crops, use_container_width=True)
    
            viable_crops = (df_crops["Energy % of Revenue"] < 40).sum()
            st.caption(f"**{viable_crops} of {len(df_crops)} crops** show energy below 40% of revenue for this country and setup. Normalised to Single harvest for fair comparison.")
    
    # ═════════════════════════════════════════════════════════════════════════════
    # SENSITIVITY ANALYSIS
    # ═════════════════════════════════════════════════════════════════════════════
    st.divider()
    st.subheader("🔬 Sensitivity Analysis")
    
    # ── Helper: run calculate() with a temporarily modified country entry ─────────
    def run_with_country_override(base_inputs: dict, kwh_mult: float = 1.0, labour_mult: float = 1.0) -> dict:
        """
        Runs calculate() with energy and/or labour prices modified by a multiplier.
        Uses a deep copy of the country entry so the global COUNTRIES dict is never mutated.
        """
        from core.data_tables import COUNTRIES
        import core.data_tables as dt
    
        country_name = base_inputs["country"]
        original = COUNTRIES[country_name]
    
        modified = copy.deepcopy(original)
        modified["kwh"]    = original["kwh"]    * kwh_mult
        modified["labour"] = original["labour"] * labour_mult
    
        # Temporarily patch the global dict
        dt.COUNTRIES[country_name] = modified
        try:
            result = calculate(base_inputs)
        finally:
            # Always restore — even if calculate() raises an exception
            dt.COUNTRIES[country_name] = original
    
        return result
    
    # ── Helper: run calculate() with yield or price multiplier ───────────────────
    def run_with_multipliers(
        base_inputs: dict,
        kwh_mult: float = 1.0,
        labour_mult: float = 1.0,
        yield_mult: float = 1.0,
        price_mult: float = 1.0,
    ) -> dict:
        """
        Runs calculate() with energy/labour via country override,
        and yield/price via modified inputs dict.
        """
        modified_inputs = copy.deepcopy(base_inputs)
    
        # Yield override: scale the crop's yield by setting price_override and
        # a yield multiplier. Since calculate() reads yield from CROPS directly,
        # we use price_override for price and a crop patch for yield.
        import core.data_tables as dt
        from core.data_tables import CROPS as _CROPS_LOCAL
        from core.data_tables import COUNTRIES as _COUNTRIES_LOCAL
    
        crop_name = base_inputs["crop"]
        original_crop = _CROPS_LOCAL[crop_name]
    
        modified_crop = copy.deepcopy(original_crop)
        modified_crop["yield"]    = original_crop["yield"]    * yield_mult
        modified_crop["yield_h2"] = original_crop["yield_h2"]  # fractions stay the same
        modified_crop["yield_h3"] = original_crop["yield_h3"]
    
        # Price override: if user already set a price_override, scale that;
        # otherwise compute base effective price and scale it
        if base_inputs.get("price_override", 0) > 0:
            modified_inputs["price_override"] = base_inputs["price_override"] * price_mult
        else:
            country = _COUNTRIES_LOCAL[base_inputs["country"]]
            base_price = original_crop[f"price_{base_inputs['price_scenario']}"] * country["food_index"]
            modified_inputs["price_override"] = base_price * price_mult
    
        country_name = base_inputs["country"]
        original_country = _COUNTRIES_LOCAL[country_name]
        modified_country = copy.deepcopy(original_country)
        modified_country["kwh"]    = original_country["kwh"]    * kwh_mult
        modified_country["labour"] = original_country["labour"] * labour_mult
    
        dt.CROPS[crop_name]              = modified_crop
        dt.COUNTRIES[country_name]       = modified_country
        try:
            if _mix_valid:
                result = run_multicrop(modified_inputs, _crop_mix)
            else:
                result = calculate(modified_inputs)
        finally:
            dt.CROPS[crop_name]        = original_crop
            dt.COUNTRIES[country_name] = original_country
    
        return result
    
    # ─────────────────────────────────────────────────────────────────────────────
    # COMPONENT 1 — TORNADO CHART
    # ─────────────────────────────────────────────────────────────────────────────
    st.markdown("#### Tornado Chart — EBITDA Sensitivity")
    st.caption("Each bar shows how EBITDA changes when one variable is stressed while all others stay at base values.")
    
    base_ebitda = r["ebitda"]
    
    # Define variables: (label, kwh_mult_pess, kwh_mult_opt, labour_mult_pess, labour_mult_opt, yield_mult_pess, yield_mult_opt, price_mult_pess, price_mult_opt)
    tornado_vars = [
        {
            "label":        "Energy Price",
            "pess":         run_with_multipliers(inputs, kwh_mult=1.50)["ebitda"],
            "opt":          run_with_multipliers(inputs, kwh_mult=0.70)["ebitda"],
            "pess_label":   "Energy price +50% (stress: energy crisis scenario)",
            "opt_label":    "Energy price −30% (upside: efficiency gains / cheaper tariff)",
            "rationale":    "Asymmetric: energy spikes are the dominant CEA risk in Europe. +50% reflects 2021–2022 style shock; −30% reflects realistic downside from long-term contracts or solar integration.",
        },
        {
            "label":        "Selling Price",
            "pess":         run_with_multipliers(inputs, price_mult=0.80)["ebitda"],
            "opt":          run_with_multipliers(inputs, price_mult=1.20)["ebitda"],
            "pess_label":   "Selling price −20% (stress: market price pressure)",
            "opt_label":    "Selling price +20% (upside: premium positioning or scarcity)",
            "rationale":    "Symmetric ±20%. Reflects realistic wholesale price volatility for premium fresh produce in European markets.",
        },
        {
            "label":        "Yield",
            "pess":         run_with_multipliers(inputs, yield_mult=0.80)["ebitda"],
            "opt":          run_with_multipliers(inputs, yield_mult=1.20)["ebitda"],
            "pess_label":   "Yield −20% (stress: underperformance vs crop profile)",
            "opt_label":    "Yield +20% (upside: optimised cultivar or growing conditions)",
            "rationale":    "Symmetric ±20%. Crop profile yields are benchmarks; actual yields vary with cultivar selection, growing protocol, and operator experience.",
        },
        {
            "label":        "Labour Cost",
            "pess":         run_with_multipliers(inputs, labour_mult=1.30)["ebitda"],
            "opt":          run_with_multipliers(inputs, labour_mult=0.80)["ebitda"],
            "pess_label":   "Labour cost +30% (stress: wage inflation or high turnover)",
            "opt_label":    "Labour cost −20% (upside: automation or lower-cost market)",
            "rationale":    "Asymmetric: labour costs in CEA tend to rise over time. +30% reflects structural wage inflation or higher-than-modelled staff requirements; −20% reflects partial automation gains.",
        },
    ]
    
    # Compute deltas from base and sort by total swing (widest bar on top)
    for tv in tornado_vars:
        tv["delta_pess"] = tv["pess"] - base_ebitda
        tv["delta_opt"]  = tv["opt"]  - base_ebitda
        tv["swing"]      = abs(tv["delta_opt"] - tv["delta_pess"])
    
    tornado_vars.sort(key=lambda x: x["swing"], reverse=True)
    
    fig_tornado = go.Figure()
    
    for tv in tornado_vars:
        # Pessimistic bar (always negative impact or less positive)
        fig_tornado.add_trace(go.Bar(
            name="Pessimistic",
            y=[tv["label"]],
            x=[tv["delta_pess"]],
            orientation="h",
            marker_color="rgba(255,77,77,0.75)",
            showlegend=(tv == tornado_vars[0]),
            text=f"${tv['delta_pess']:,.0f}",
            textposition="outside",
            customdata=[[tv["pess_label"], tv["rationale"]]],
            hovertemplate=(
                "<b>%{y}</b><br>"
                "EBITDA delta: %{x:$,.0f}<br>"
                "<i>%{customdata[0]}</i><br>"
                "<br><b>Rationale:</b> %{customdata[1]}<extra></extra>"
            ),
        ))
        # Optimistic bar
        fig_tornado.add_trace(go.Bar(
            name="Optimistic",
            y=[tv["label"]],
            x=[tv["delta_opt"]],
            orientation="h",
            marker_color="rgba(0,229,160,0.75)",
            showlegend=(tv == tornado_vars[0]),
            text=f"${tv['delta_opt']:,.0f}",
            textposition="outside",
            customdata=[[tv["opt_label"], tv["rationale"]]],
            hovertemplate=(
                "<b>%{y}</b><br>"
                "EBITDA delta: %{x:$,.0f}<br>"
                "<i>%{customdata[0]}</i><br>"
                "<br><b>Rationale:</b> %{customdata[1]}<extra></extra>"
            ),
        ))
    
    fig_tornado.add_vline(x=0, line_dash="dash", line_color="rgba(255,255,255,0.4)")
    fig_tornado.update_layout(
        barmode="overlay",
        plot_bgcolor="#ffffff",
        paper_bgcolor="#ffffff",
        font_color="#161a16",
        height=320,
        margin=dict(l=10, r=80, t=20, b=20),
        xaxis=dict(title="EBITDA delta from base ($)", showgrid=False, zeroline=False),
        yaxis=dict(showgrid=False),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    )
    style_fig(fig_tornado)
    st.plotly_chart(fig_tornado, use_container_width=True)
    
    # ─────────────────────────────────────────────────────────────────────────────
    # COMPONENT 2 — SCENARIO COMPARISON
    # ─────────────────────────────────────────────────────────────────────────────
    st.divider()
    st.markdown("#### Scenario Comparison")
    st.caption("Define up to 4 named scenarios by adjusting multipliers. The Base Case always shows the current sidebar inputs.")
    
    if "scenarios" not in st.session_state:
        try:
            resp = supabase.table("scenarios").select("*").is_("farm_id", "null").order("created_at", desc=False).execute()
            loaded = []
            for row in resp.data:
                try:
                    result = json.loads(row["outputs_json"]) if row.get("outputs_json") else {}
                    loaded.append({
                        "name":         row["name"],
                        "energy_mult":  row["energy_factor"],
                        "yield_mult":   row["yield_factor"],
                        "price_mult":   row["price_factor"],
                        "labour_mult":  row["labour_factor"],
                        "is_actual":    row["is_actual"],
                        "result":       result,
                        "supabase_id":  row["id"],
                    })
                except Exception:
                    continue
            st.session_state["scenarios"] = loaded
        except Exception:
            st.session_state["scenarios"] = []
    
    # ── Scenario builder ──────────────────────────────────────────────────────────
    with st.expander("➕ Define a new scenario", expanded=len(st.session_state["scenarios"]) == 0):
        sc_name = st.text_input("Scenario name", value="Scenario 1", key="sc_name_input")
    
        sc_col1, sc_col2 = st.columns(2)
        with sc_col1:
            sc_energy   = st.slider("Energy price multiplier",  min_value=0.3, max_value=3.0, value=1.0, step=0.05, key="sc_energy")
            sc_yield    = st.slider("Yield multiplier",         min_value=0.3, max_value=2.0, value=1.0, step=0.05, key="sc_yield")
        with sc_col2:
            sc_price    = st.slider("Selling price multiplier", min_value=0.3, max_value=2.0, value=1.0, step=0.05, key="sc_price")
            sc_labour   = st.slider("Labour cost multiplier",   min_value=0.3, max_value=3.0, value=1.0, step=0.05, key="sc_labour")
    
        if st.button("💾 Save Scenario", use_container_width=True):
            if len(st.session_state["scenarios"]) >= 4:
                st.warning("Maximum 4 scenarios reached. Delete one before adding another.")
            else:
                sc_result = run_with_multipliers(
                    inputs,
                    kwh_mult=sc_energy,
                    labour_mult=sc_labour,
                    yield_mult=sc_yield,
                    price_mult=sc_price,
                )
                new_scenario = {
                    "name":        sc_name,
                    "energy_mult": sc_energy,
                    "yield_mult":  sc_yield,
                    "price_mult":  sc_price,
                    "labour_mult": sc_labour,
                    "result":      sc_result,
                    "is_actual":   False,
                    "supabase_id": None,
                }
                try:
                    sb_resp = supabase.table("scenarios").insert({
                        "name":          sc_name,
                        "energy_factor": sc_energy,
                        "yield_factor":  sc_yield,
                        "price_factor":  sc_price,
                        "labour_factor": sc_labour,
                        "capex_factor":  1.0,
                        "is_actual":     False,
                        "farm_id":       None,
                        "outputs_json":  json.dumps(sc_result),
                    }).execute()
                    if sb_resp.data:
                        new_scenario["supabase_id"] = sb_resp.data[0]["id"]
                    st.success(f"Scenario '{sc_name}' saved.")
                except Exception as e:
                    st.error(f"Saved locally but could not write to database: {e}")
                    st.success(f"Scenario '{sc_name}' saved locally.")
                st.session_state["scenarios"].append(new_scenario)
    
        if st.session_state["scenarios"]:
            if st.button("🗑️ Clear all scenarios", use_container_width=True):
                try:
                    supabase.table("scenarios").delete().is_("farm_id", "null").execute()
                except Exception as e:
                    st.error(f"Could not delete from database: {e}")
                st.session_state["scenarios"] = []
                st.rerun()
    
    # ── Comparison display ────────────────────────────────────────────────────────
    if st.session_state["scenarios"]:
    
        # Per-scenario delete buttons
        for idx, sc in enumerate(st.session_state["scenarios"]):
            del_col1, del_col2 = st.columns([8, 1])
            with del_col1:
                st.caption(f"**{sc['name']}** — energy ×{sc['energy_mult']} / yield ×{sc['yield_mult']} / price ×{sc['price_mult']} / labour ×{sc['labour_mult']}")
            with del_col2:
                if st.button("🗑️", key=f"del_sc_{idx}"):
                    if sc.get("supabase_id"):
                        try:
                            supabase.table("scenarios").delete().eq("id", sc["supabase_id"]).execute()
                        except Exception as e:
                            st.error(f"Could not delete from database: {e}")
                    st.session_state["scenarios"].pop(idx)
                    st.rerun()
    
        scenario_names   = ["Base Case"] + [s["name"] for s in st.session_state["scenarios"]]
        scenario_results = [r] + [s["result"] for s in st.session_state["scenarios"]]
    
        # Grouped bar chart
        metrics_to_plot = {
            "EBITDA":       [res["ebitda"]             for res in scenario_results],
            "Revenue":      [res["annual_revenue"]      for res in scenario_results],
            "Energy Cost":  [res["annual_energy_cost"]  for res in scenario_results],
            "Labour Cost":  [res["annual_labour_cost"]  for res in scenario_results],
        }
    
        bar_colors_chart = ["#00e5a0", "#ffc13d", "#4fc3f7", "#ba68c8", "#ef9a9a"]
    
        fig_compare = go.Figure()
        for i, (metric_name, values) in enumerate(metrics_to_plot.items()):
            fig_compare.add_trace(go.Bar(
                name=metric_name,
                x=scenario_names,
                y=values,
                marker_color=bar_colors_chart[i],
                text=[f"${v:,.0f}" for v in values],
                textposition="outside",
            ))
    
        fig_compare.update_layout(
            barmode="group",
            plot_bgcolor="#ffffff",
            paper_bgcolor="#ffffff",
            font_color="#161a16",
            height=420,
            margin=dict(t=30, b=20),
            xaxis=dict(showgrid=False),
            yaxis=dict(showgrid=False, title="$"),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        )
        style_fig(fig_compare)
        st.plotly_chart(fig_compare, use_container_width=True)
    
        # Summary table
        def energy_pct(res):
            if res["annual_revenue"] > 0:
                return f"{res['annual_energy_cost'] / res['annual_revenue'] * 100:.1f}%"
            return "N/A"
    
        def fmt_payback(res):
            return f"{res['payback_years']:.1f}" if res["payback_years"] else "N/A"
    
        table_rows = []
        for name, res in zip(scenario_names, scenario_results):
            table_rows.append({
                "Scenario":           name,
                "Revenue":            f"${res['annual_revenue']:,.0f}",
                "Energy Cost":        f"${res['annual_energy_cost']:,.0f}",
                "Labour Cost":        f"${res['annual_labour_cost']:,.0f}",
                "EBITDA":             f"${res['ebitda']:,.0f}",
                "EBITDA Margin":      f"{res['ebitda_margin']*100:.1f}%",
                "Payback (yrs)":      fmt_payback(res),
                "Energy % of Revenue": energy_pct(res),
            })
    
        scenario_df = pd.DataFrame(table_rows)
    
        def highlight_energy_pct(row):
            style = severity_cell(row["Energy % of Revenue"], hi=60, mid=40)
            return [style] * len(row)

        st.dataframe(
            scenario_df.style.apply(highlight_energy_pct, axis=1),
            use_container_width=True,
            hide_index=True,
        )
    
    else:
        st.info("No scenarios saved yet. Use the form above to define and save your first scenario.")

elif modality == "🌿 High-Tech Greenhouse":

    st.markdown("### 🌿 High-Tech Greenhouse & Polytunnel Calculator")

    # ── Session state defaults ────────────────────────────────────────────────
    _GH_DEFAULTS = {
        "gh_country":           "Germany",
        "gh_crop_source":       "Greenhouse",
        "gh_crop":              "Tomato (Beef)",
        "gh_footprint":         1000,
        "gh_automation":        "Medium",
        "gh_price_scenario":    "base",
        "gh_harvest_mode":      "Single",
        "gh_price_override":    0.0,
        "gh_packaging_cost":    0.15,
        "gh_loss_rate":         5.0,
        "gh_net_grow_factor":   90.0,
        "gh_walkways_factor":   10.0,
        "gh_water_price":       2.0,
        "gh_kwh_override":      0.0,
        "gh_rent_monthly":      0.0,
        "gh_real_estate_capex": 0.0,
        "gh_depreciation_years": 15,
        "gh_tax_rate":          25.0,
        "gh_ltv":               60.0,
        "gh_interest_rate":     5.5,
        "gh_loan_term_years":   15,
        "gh_discount_rate":     8.0,
        "gh_multi_crop":        False,
        "gh_crop_mix":          [{"crop": "Tomato (Beef)", "pct": 100}],
    }
    for _k, _v in _GH_DEFAULTS.items():
        if _k not in st.session_state:
            st.session_state[_k] = _v

    if "gh_show_save_form" not in st.session_state:
        st.session_state["gh_show_save_form"] = False

    # ── Sidebar ───────────────────────────────────────────────────────────────
    with st.sidebar:
        st.header("Greenhouse Parameters")

        st.divider()

        # ── Farm parameters ───────────────────────────────────────────────────
        gh_country_list = list(COUNTRIES.keys())
        gh_country = st.selectbox("Country", gh_country_list,
                                  index=gh_country_list.index(st.session_state["gh_country"]) if st.session_state["gh_country"] in gh_country_list else 0,
                                  key="gh_country")

        # ── Energy & Labour reference rates (energy_labour module) ────────────
        _el_rates  = get_rates_for_country_name(gh_country)
        _el_e      = _el_rates["energy"]
        _el_l      = _el_rates["labour"]
        _model_kwh = COUNTRIES.get(gh_country, {}).get("kwh", 0)
        if _el_rates["iso"]:
            _el_delta = _el_e["industrial"] - _model_kwh
            _el_arrow = "▲" if _el_delta > 0.005 else ("▼" if _el_delta < -0.005 else "≈")
            _el_col   = "#d4a845" if _el_delta > 0.005 else ("#52a066" if _el_delta < -0.005 else "#7a8070")
            _el_live  = " · ⚡ Live" if _el_e.get("live") else ""
            _el_html  = (
                f"<div style='font-size:11px;color:#9ba390;margin:-4px 0 8px 0;"
                f"padding:6px 8px;background:#252a25;border-radius:3px;"
                f"border:1px solid #363c36;border-left:3px solid {_el_col};'>"
                f"<b>Ref. rates ({_el_rates['iso']})</b> &nbsp;&middot;&nbsp; "
                f"Electricity industrial: <b style='color:{_el_col};'>${_el_e['industrial']:.3f}/kWh {_el_arrow}</b> "
                f"<span style='color:#7a8070;'>(model ${_model_kwh:.3f})</span>"
                f"&nbsp;&middot;&nbsp; Labour: <b>${_el_l['industrial_loaded']:.0f}/hr</b>"
                f"{_el_live}</div>"
            )
            st.markdown(_el_html, unsafe_allow_html=True)

        gh_crop_source = st.radio("Crop source", ["Greenhouse", "Polytunnel"],
                                  index=0 if st.session_state["gh_crop_source"] == "Greenhouse" else 1,
                                  horizontal=True, key="gh_crop_source")

        if gh_crop_source == "Greenhouse":
            gh_crop_list = list(GREENHOUSE_CROPS.keys())
        else:
            gh_crop_list = list(POLYTUNNEL_CROPS.keys())

        # ── Multi-crop toggle ─────────────────────────────────────────────────
        gh_multi_crop_mode = st.toggle("Multi-crop mode", value=False, key="gh_multi_crop")

        if not gh_multi_crop_mode:
            _gh_crop_default = st.session_state["gh_crop"]
            gh_crop = st.selectbox("Crop", gh_crop_list,
                                   index=gh_crop_list.index(_gh_crop_default) if _gh_crop_default in gh_crop_list else 0,
                                   key="gh_crop")
        else:
            gh_crop = gh_crop_list[0]  # placeholder — unused in multi-crop mode
            st.markdown("**Crop allocation** (must sum to 100%)")
            if "gh_crop_mix" not in st.session_state:
                st.session_state["gh_crop_mix"] = [{"crop": gh_crop_list[0], "pct": 100}]
            _gh_mix = st.session_state["gh_crop_mix"]
            # Validate all mix crops exist in current crop list
            _gh_mix = [row for row in _gh_mix if row["crop"] in gh_crop_list]
            if not _gh_mix:
                _gh_mix = [{"crop": gh_crop_list[0], "pct": 100}]
            _gh_to_remove = None
            for _gci, _grow in enumerate(_gh_mix):
                _gc1, _gc2, _gc3 = st.columns([4, 2, 1])
                with _gc1:
                    _gh_mix[_gci]["crop"] = st.selectbox(
                        f"Crop {_gci+1}", gh_crop_list,
                        index=gh_crop_list.index(_grow["crop"]) if _grow["crop"] in gh_crop_list else 0,
                        key=f"gh_mix_crop_{_gci}")
                with _gc2:
                    _gh_mix[_gci]["pct"] = st.number_input(
                        "%", min_value=1, max_value=100, value=int(_grow["pct"]),
                        step=1, key=f"gh_mix_pct_{_gci}")
                with _gc3:
                    if len(_gh_mix) > 1 and st.button("✕", key=f"gh_mix_del_{_gci}"):
                        _gh_to_remove = _gci
            if _gh_to_remove is not None:
                _gh_mix.pop(_gh_to_remove)
                st.session_state["gh_crop_mix"] = _gh_mix
                st.rerun()
            _gh_total_pct = sum(r["pct"] for r in _gh_mix)
            st.caption(f"Total allocated: **{_gh_total_pct}%**")
            if _gh_total_pct != 100:
                st.warning(f"⚠️ Must sum to 100%. Currently {_gh_total_pct}%.")
            if len(_gh_mix) < 6:
                if st.button("➕ Add crop", key="gh_mix_add"):
                    _gh_mix.append({"crop": gh_crop_list[0], "pct": 1})
                    st.session_state["gh_crop_mix"] = _gh_mix
                    st.rerun()
            st.session_state["gh_crop_mix"] = _gh_mix

        st.divider()
        gh_footprint = st.number_input("Footprint (m²)", value=st.session_state["gh_footprint"],
                                       step=100, min_value=50, key="gh_footprint")

        st.divider()
        gh_automation = st.selectbox("Automation", ["None", "Low", "Medium", "High"],
                                     index=["None", "Low", "Medium", "High"].index(st.session_state["gh_automation"]),
                                     key="gh_automation")

        ps_list_gh = ["base", "low", "high"]
        gh_price_scenario = st.selectbox("Price Scenario", ps_list_gh,
                                         index=ps_list_gh.index(st.session_state["gh_price_scenario"]),
                                         key="gh_price_scenario")

        if gh_crop_source == "Greenhouse":
            gh_crop_data = GREENHOUSE_CROPS[gh_crop]
        else:
            gh_crop_data = POLYTUNNEL_CROPS[gh_crop]

        if gh_crop_data["days_between"] > 0:
            hm_list_gh = ["Single", "2 Harvests", "3 Harvests"]
            _gh_hm_default = st.session_state["gh_harvest_mode"]
            gh_harvest_mode = st.selectbox("Harvest Mode", hm_list_gh,
                                           index=hm_list_gh.index(_gh_hm_default) if _gh_hm_default in hm_list_gh else 0,
                                           key="gh_harvest_mode")
        else:
            st.selectbox("Harvest Mode", ["Single"], disabled=True,
                         help="This crop only supports single harvest.")
            gh_harvest_mode = "Single"
            st.session_state["gh_harvest_mode"] = "Single"

        st.divider()
        st.subheader("Advanced")
        gh_price_override    = st.number_input("Price Override ($/kg, 0=auto)",
                                               value=st.session_state["gh_price_override"],
                                               step=0.1, min_value=0.0, key="gh_price_override")
        gh_packaging_cost    = st.number_input("Packaging ($/kg)",
                                               value=st.session_state["gh_packaging_cost"],
                                               step=0.01, min_value=0.0, key="gh_packaging_cost")
        gh_loss_rate         = st.number_input("Loss Rate (%)",
                                               value=st.session_state["gh_loss_rate"],
                                               step=0.5, min_value=0.0, max_value=100.0, key="gh_loss_rate")
        gh_net_grow_factor   = st.number_input("Net Grow Factor (%)",
                                               value=st.session_state["gh_net_grow_factor"],
                                               step=1.0, min_value=1.0, max_value=100.0, key="gh_net_grow_factor")
        gh_walkways_factor   = st.number_input("Walkways Factor (%)",
                                               value=st.session_state["gh_walkways_factor"],
                                               step=1.0, min_value=0.0, max_value=50.0, key="gh_walkways_factor")
        gh_water_price       = st.number_input("Water Price ($/m³)",
                                               value=st.session_state["gh_water_price"],
                                               step=0.1, min_value=0.0, key="gh_water_price")
        _gh_kwh_default      = COUNTRIES.get(gh_country, {}).get("kwh", 0.0)
        gh_kwh_override      = st.number_input(
            "Electricity Price ($/kWh)",
            value=float(st.session_state.get("gh_kwh_override") or _gh_kwh_default),
            step=0.005, min_value=0.001, format="%.4f", key="gh_kwh_override",
            help=(
                f"Country default: ${_gh_kwh_default:.4f}/kWh. "
                "Override with your actual site tariff. "
                "Industrial rate is typically 30–60% lower than residential."
            )
        )
        gh_rent_monthly      = st.number_input("Monthly Rent ($)",
                                               value=st.session_state["gh_rent_monthly"],
                                               step=100.0, min_value=0.0, key="gh_rent_monthly")
        gh_real_estate_capex = st.number_input("Real Estate CAPEX ($)",
                                               value=st.session_state["gh_real_estate_capex"],
                                               step=10000.0, min_value=0.0, key="gh_real_estate_capex")

        st.divider()
        st.subheader("Financial Structure")
        gh_depreciation_years = st.number_input("Depreciation (years)",
                                                value=st.session_state["gh_depreciation_years"],
                                                step=1, min_value=1, key="gh_depreciation_years")
        gh_tax_rate           = st.number_input("Tax Rate (%)",
                                                value=st.session_state["gh_tax_rate"],
                                                step=1.0, min_value=0.0, max_value=100.0, key="gh_tax_rate")
        gh_ltv                = st.number_input("LTV (%)",
                                                value=st.session_state["gh_ltv"],
                                                step=5.0, min_value=0.0, max_value=100.0, key="gh_ltv")
        gh_interest_rate      = st.number_input("Interest Rate (%)",
                                                value=st.session_state["gh_interest_rate"],
                                                step=0.1, min_value=0.0, key="gh_interest_rate")
        gh_loan_term_years    = st.number_input("Loan Term (years)",
                                                value=st.session_state["gh_loan_term_years"],
                                                step=1, min_value=1, key="gh_loan_term_years")
        gh_discount_rate      = st.number_input("Discount Rate (%)",
                                                value=st.session_state["gh_discount_rate"],
                                                step=0.5, min_value=0.0, key="gh_discount_rate")

    _gh_multi_crop_mode = st.session_state.get("gh_multi_crop", False)
    _gh_mix             = st.session_state.get("gh_crop_mix", [])
    # Validate all mix crops exist in current crop pool
    _gh_crop_pool = list(GREENHOUSE_CROPS.keys()) if gh_crop_source == "Greenhouse" else list(POLYTUNNEL_CROPS.keys())
    _gh_mix = [row for row in _gh_mix if row["crop"] in _gh_crop_pool]
    _gh_mix_total       = sum(row["pct"] for row in _gh_mix)
    _gh_mix_valid       = _gh_multi_crop_mode and len(_gh_mix) > 0 and _gh_mix_total == 100

    # ── Run calculation ───────────────────────────────────────────────────────
    gh_inputs = {
        "country":            gh_country,
        "crop":               gh_crop,
        "crop_source":        gh_crop_source.lower(),
        "footprint":          gh_footprint,
        "automation":         gh_automation,
        "price_scenario":     gh_price_scenario,
        "price_override":     gh_price_override,
        "packaging_cost":     gh_packaging_cost,
        "loss_rate":          gh_loss_rate,
        "net_grow_factor":    gh_net_grow_factor,
        "walkways_factor":    gh_walkways_factor,
        "water_price":        gh_water_price,
        "rent_monthly":       gh_rent_monthly,
        "real_estate_capex":  gh_real_estate_capex,
        "harvest_mode":       gh_harvest_mode,
        "depreciation_years": gh_depreciation_years,
        "tax_rate":           gh_tax_rate,
        "ltv":                gh_ltv,
        "interest_rate":      gh_interest_rate,
        "loan_term_years":    gh_loan_term_years,
        "discount_rate":      gh_discount_rate,
        "ambient_temp_annual":    st.session_state.get("active_farm", {}).get("ambient_temp_annual"),
        "mean_annual_dli":        st.session_state.get("active_farm", {}).get("mean_annual_dli"),
        "crop_mix_json":      json.dumps(_gh_mix) if _gh_mix_valid else None,
    }

    if _gh_multi_crop_mode and not _gh_mix_valid:
        st.warning("⚠️ Fix greenhouse crop allocation (must sum to 100%) before results are shown.")
        st.stop()

    _gh_crop_data_dict = GREENHOUSE_CROPS if gh_crop_source == "Greenhouse" else POLYTUNNEL_CROPS

    # Apply electricity price override
    _gh_kwh_original = COUNTRIES[gh_country]["kwh"]
    if abs(gh_kwh_override - _gh_kwh_original) > 0.0001:
        COUNTRIES[gh_country]["kwh"] = gh_kwh_override
    if _gh_mix_valid:
        gh_r = _run_multicrop_generic(gh_inputs, _gh_mix,
                                       calculate_greenhouse, _gh_crop_data_dict)
    else:
        gh_r = calculate_greenhouse(gh_inputs)
    COUNTRIES[gh_country]["kwh"] = _gh_kwh_original  # always restore

    # ── Climate profile display ─────────────────────────────────────────────────
    _gh_active2 = st.session_state.get("active_farm")
    if _gh_active2 and _gh_active2.get("mean_annual_dli"):
        _gh_loc_dli2  = _gh_active2["mean_annual_dli"]
        _gh_loc_temp2 = _gh_active2["ambient_temp_annual"] # Remove emoji from caption
        _gh_crop_dli2 = gh_crop_data["dli"]
        _gh_nat_frac2 = compute_natural_dli_fraction(_gh_loc_dli2, _gh_crop_dli2)
        st.caption(
            f"🌤️ **Climate profile active** — "
            f"Mean annual DLI: {_gh_loc_dli2:.1f} mol/m²/day · "
            f"Ambient temperature: {_gh_loc_temp2:.1f}°C · "
            f"Natural DLI coverage for {gh_crop}: {_gh_nat_frac2*100:.0f}% "
            f"({'supplemental lighting required' if _gh_nat_frac2 < 1.0 else 'no supplemental lighting required'})"
        )

    # ── Data Sources panel ───────────────────────────────────────────────────
    _gh_has_climate  = bool(st.session_state.get("active_farm", {}).get("mean_annual_dli")) # Remove emoji from expander title
    _gh_active_data  = st.session_state.get("active_farm") or {}
    with st.expander("ℹ️ Data sources & calculation transparency", expanded=False):
        _gdi1, _gdi2 = st.columns(2)
        with _gdi1:
            st.markdown("**📡 Automatic — from Open-Meteo Archive API**")
            if _gh_has_climate:
                _gh_dli2  = _gh_active_data.get("mean_annual_dli", 0)
                _gh_temp2 = _gh_active_data.get("ambient_temp_annual", 0)
                _gh_nat2  = compute_natural_dli_fraction(_gh_dli2, gh_crop_data["dli"])
                st.markdown(
                    f"- **Mean annual DLI: {_gh_dli2:.1f} mol/m²/day** "
                    f"— determines natural DLI fraction ({_gh_nat2*100:.0f}% for {gh_crop}). "
                    f"Supplemental lighting kWh scales with the shortfall from crop DLI requirement.\n"
                    f"- **Ambient temperature: {_gh_temp2:.1f}°C** "
                    f"— directly drives heating energy. ΔT = target − {_gh_temp2:.1f}°C. "
                    f"A colder location increases heating OPEX automatically.\n"
                    f"- Source: Open-Meteo 10-year historical archive. Stored in Supabase at farm save time."
                )
            else:
                st.markdown(
                    "- ⚠️ **No climate data available** for this farm.\n"
                    "- Set coordinates in the **Farm Intelligence Map**, then re-save the farm profile.\n"
                    "- Until then, heating energy uses a generic temperate climate assumption "
                    "and supplemental lighting is based on crop's built-in natural DLI fraction." # Keep warning emoji
                )
        with _gdi2:
            st.markdown("**🎛️ Manual inputs — set in this calculator**")
            st.markdown(
                "- Crop → auto-selects structure type (Polytunnel / Multi-span / Venlo) "
                "and base natural DLI fraction. See Assumptions §11.\n"
                "- Country → electricity price (IEA/Eurostat, see §5)\n"
                "- Footprint, automation, price scenario\n"
                "- All financial structure inputs (see §17.7)"
            )
        st.caption(
            "ℹ️ GH heating formula: ΔT × 10 W/m²/°C × footprint × 8,760 hrs ÷ 1,000. "
            "Supplemental lighting: (crop DLI − natural DLI) × 0.0216 kWh/mol × operating days. "
            "Full derivation in Assumptions §11.2."
        )


    # ── Energy & Labour calibration callout ──────────────────────────────────
    _el_r2    = get_rates_for_country_name(gh_inputs["country"])
    _el_e2    = _el_r2["energy"]
    _el_l2    = _el_r2["labour"]
    _mkwh2    = COUNTRIES.get(gh_inputs["country"], {}).get("kwh", 0)
    _mlabour2 = COUNTRIES.get(gh_inputs["country"], {}).get("labour", 0)
    if _el_r2["iso"]:
        _e_flag  = abs(_el_e2["industrial"] - _mkwh2) > 0.01 # Keep warning emoji
        _l_flag  = abs(_el_l2["industrial_loaded"] - _mlabour2) > 3.0
        _exp_lbl = "⚠️ Verify your input assumptions" if (_e_flag or _l_flag) else "✅ Input assumptions cross-check"
        with st.expander(_exp_lbl, expanded=(_e_flag or _l_flag)):
            _rc1, _rc2 = st.columns(2)
            with _rc1:
                st.markdown("**⚡ Electricity**")
                _e_dir = "higher" if _el_e2["industrial"] > _mkwh2 else "lower"
                _e_pct = abs(_el_e2["industrial"] - _mkwh2) / _mkwh2 * 100 if _mkwh2 else 0 # Keep warning emoji
                if _e_flag:
                    st.warning(
                        f"Reference industrial rate: **${_el_e2['industrial']:.3f}/kWh** "
                        f"({_e_pct:.0f}% {_e_dir} than model’s ${_mkwh2:.3f}/kWh). "
                        f"Verify country default or use a site-specific override if your tariff differs."
                    )
                else:
                    st.success(f"Model electricity (${_mkwh2:.3f}/kWh) aligns with reference industrial rate (${_el_e2['industrial']:.3f}/kWh).")
                st.caption(f"Source: {_el_e2['source']}" + (f" · {_el_e2['live_note']}" if _el_e2.get("live") else "")) # Remove emoji from caption
            with _rc2:
                st.markdown("**👷 Labour**")
                _l_dir = "higher" if _el_l2["industrial_loaded"] > _mlabour2 else "lower"
                _l_pct = abs(_el_l2["industrial_loaded"] - _mlabour2) / _mlabour2 * 100 if _mlabour2 else 0
                if _l_flag:
                    st.warning(
                        f"Reference fully-loaded industrial: **${_el_l2['industrial_loaded']:.0f}/hr** "
                        f"({_l_pct:.0f}% {_l_dir} than model’s ${_mlabour2:.0f}/hr). "
                        f"Overhead {_el_l2['overhead_pct']} applied (base ${_el_l2['industrial_base']:.0f}/hr)."
                    ) # Keep warning emoji
                else:
                    st.success(f"Model labour (${_mlabour2:.0f}/hr) aligns with reference (${_el_l2['industrial_loaded']:.0f}/hr, overhead {_el_l2['overhead_pct']}).")
                st.caption(f"Source: {_el_l2['source']}")

    # ── Key metrics ───────────────────────────────────────────────────────────
    # ── PDF Report ────────────────────────────────────────────────────────────
    def generate_gh_pdf_report(gh_inputs: dict, gh_r: dict) -> bytes:
        _fn = st.session_state.get("active_farm", {}).get("name", "")
        _mc = "pt" if gh_inputs.get("crop_source","").lower() == "polytunnel" else "gh"
        def _gh_sens(kwh_m=1.0, lab_m=1.0, yld_m=1.0, prc_m=1.0):
            return _gh_run_mult(gh_inputs, kwh_m=kwh_m, lab_m=lab_m,
                                yld_m=yld_m, prc_m=prc_m)
        return _build_feasibility_pdf(gh_r, gh_inputs, _mc, farm_name=_fn,
                                      run_sens_fn=_gh_sens)

    gh_pdf_col1, gh_pdf_col2 = st.columns([5, 1])
    with gh_pdf_col2:
        if st.button("📄 Download PDF Report", key="gh_pdf_btn", use_container_width=True): # Keep emoji in button
            with st.spinner("Generating PDF..."):
                gh_pdf_bytes = generate_gh_pdf_report(gh_inputs, gh_r)
                # Correct naming: use primary crop from mix if multi-crop is valid
                _gh_rep_crop = _gh_mix[0]["crop"] if _gh_mix_valid and _gh_mix else gh_inputs["crop"]
                _gh_rep_crop_safe = _gh_rep_crop.replace(' ','_').replace('/','').replace('(','').replace(')','_')
                gh_filename = f"GH_Report_{_gh_rep_crop_safe}_{gh_inputs['country']}_{date.today().strftime('%Y%m%d')}.pdf"
                st.download_button(label="⬇️ Save PDF", data=gh_pdf_bytes,
                                   file_name=gh_filename, mime="application/pdf",
                                   use_container_width=True, key="gh_pdf_dl")
    st.divider()

    st.subheader("Key Metrics")

    gh_energy_pct = (gh_r["annual_energy_cost"] / gh_r["annual_revenue"] * 100
                     if gh_r["annual_revenue"] > 0 else 0)
    _gh_loss  = gh_inputs["loss_rate"] / 100
    _gh_denom = (gh_r["effective_price"] * (1 - _gh_loss)
                 * gh_r["cycles_per_year"] * gh_r["effective_grow_area"])
    gh_be_yield = gh_r["total_annual_costs"] / _gh_denom if _gh_denom > 0 else None
    _gh_projected_yield = gh_crop_data["yield"]
    if gh_be_yield is not None:
        _gh_yield_gap = (_gh_projected_yield - gh_be_yield) / gh_be_yield * 100
        _gh_yield_gap_str = f"{_gh_yield_gap:+.1f}%"
    else:
        _gh_yield_gap_str = "N/A"

    gm1, gm2, gm3, gm4, gm5, gm6, gm7 = st.columns(7)
    gm1.metric("Annual Revenue",   f"${gh_r['annual_revenue']:,.0f}")
    gm2.metric("EBITDA",           f"${gh_r['ebitda']:,.0f}")
    gm3.metric("EBITDA Margin",    f"{gh_r['ebitda_margin']*100:.1f}%")
    gm4.metric("Total CAPEX",      f"${gh_r['total_capex']:,.0f}")
    gm5.metric("Payback",          f"{gh_r['payback_years']:.1f} yrs" if gh_r["payback_years"] else "N/A")
    gm6.metric("DSCR",             f"{gh_r['dscr']:.2f}" if gh_r["dscr"] else "N/A")
    gm7.metric(
        "Break-even Yield",
        f"{gh_be_yield:.2f} kg/m²/cycle" if gh_be_yield else "N/A",
        delta=_gh_yield_gap_str if gh_be_yield else None,
        delta_color="normal",
        help="Minimum yield needed to cover all costs. Delta = projected crop yield vs this threshold."
    )

    if gh_r["supplemental_dli"] > 0:
        st.caption(
            f"💡 Supplemental DLI: {gh_r['supplemental_dli']:.1f} mol/m²/day "
            f"({(1 - gh_crop_data['natural_dli_fraction'])*100:.0f}% artificial) — "
            f"{gh_r['lighting_kwh_m2_year']:.1f} kWh/m²/year for lighting"
        )
    else:
        st.caption("☀️ Natural light only — no supplemental lighting required.")

    if gh_r["dscr"] is not None and gh_r["dscr"] < 1.0:
        st.warning( # Keep warning emoji
            f"⚠️ **Debt service coverage is low (DSCR = {gh_r['dscr']:.2f}x).** "
            f"Annual debt repayment (${gh_r['annual_debt_service']:,.0f}) exceeds EBITDA (${gh_r['ebitda']:,.0f}). "
            f"Consider reducing LTV, extending the loan term, increasing farm scale, or selecting a higher-margin crop."
        )

    # ── Multi-crop breakdown ──────────────────────────────────────────────────
    if gh_r.get("_is_multicrop"):
        st.divider()
        st.subheader("Per-Crop Breakdown")
        _gh_mc_rows = []
        for _mc in gh_r["_crop_results"]:
            _gh_mc_rows.append({
                "Crop":           _mc["crop"],
                "Area %":         f"{_mc['pct']:.0f}%",
                "Annual kg":      f"{_mc['total_annual_kg']:,.0f}",
                "Price ($/kg)":   f"${_mc['effective_price']:.2f}",
                "Revenue":        f"${_mc['annual_revenue']:,.0f}",
                "Variable Cost":  f"${_mc['annual_variable_cost']:,.0f}",
                "Labour":         f"${_mc['annual_labour_cost']:,.0f}",
                "EBITDA contrib": f"${_mc['ebitda']:,.0f}",
            })
        st.dataframe(pd.DataFrame(_gh_mc_rows), use_container_width=True, hide_index=True)
        st.caption(
            "Energy cost and CAPEX are shared farm infrastructure — "
            "computed once on the full farm area and shown in the combined metrics above."
        )
        st.divider()
    # ── Save as Farm Profile ──────────────────────────────────────────────────
    st.divider()
    gh_save_col1, gh_save_col2 = st.columns([5, 1])
    with gh_save_col2: # Keep emoji in button
        if st.button("💾 Save as Farm Profile", key="gh_save_btn", use_container_width=True):
            st.session_state["gh_show_save_form"] = True

    if st.session_state["gh_show_save_form"]:
        with st.container(border=True):
            _gh_active   = st.session_state.get("active_farm")
            _save_lat = st.session_state.get("shared_lat")
            _save_lon = st.session_state.get("shared_lng")
            _climate_data = {}
            if _save_lat and _save_lon:
                with st.spinner("🌤️ Fetching climate profile for this location…"):
                    try: # Keep emoji in spinner
                        from core.climate import fetch_climate_profile
                        _climate_data = fetch_climate_profile(_save_lat, _save_lon)
                    except Exception:
                        _climate_data = {}
            _gh_modality = "polytunnel" if gh_crop_source == "Polytunnel" else "greenhouse"
            _gh_payload  = {
                "country":           gh_inputs["country"],
                "crop":              (_gh_crop_mix[0]["crop"] if _gh_mix_valid and _gh_crop_mix else gh_inputs["crop"]),
                "footprint":         gh_inputs["footprint"],
                "automation":        gh_inputs["automation"],
                "price_scenario":    gh_inputs["price_scenario"],
                "price_override":    gh_inputs["price_override"],
                "packaging_cost":    gh_inputs["packaging_cost"],
                "loss_rate":         gh_inputs["loss_rate"],
                "net_grow_factor":   gh_inputs["net_grow_factor"],
                "walkways_factor":   gh_inputs["walkways_factor"],
                "water_price":       gh_inputs["water_price"],
                "rent_monthly":      gh_inputs["rent_monthly"],
                "real_estate_capex": gh_inputs["real_estate_capex"],
                "harvest_mode":      gh_inputs["harvest_mode"],
                "depreciation_years": gh_inputs["depreciation_years"],
                "tax_rate":          gh_inputs["tax_rate"],
                "ltv":               gh_inputs["ltv"],
                "interest_rate":     gh_inputs["interest_rate"],
                "loan_term_years":   gh_inputs["loan_term_years"],
                "lat":               st.session_state.get("shared_lat"),
                "lon":               st.session_state.get("shared_lng"),
                "ambient_temp_annual": _climate_data.get("ambient_temp_annual"),
                "mean_annual_dli":     _climate_data.get("mean_annual_dli"),
                "agriculture_type":  _gh_modality,
                "modality":          _gh_modality,
                "crop_source":       gh_inputs["crop_source"],
                "discount_rate":     gh_inputs["discount_rate"],
                "metadata":          {},
                "model_snapshot":    json.dumps(gh_r),
                "model_updated_at":  date.today().isoformat(),
                "crop_mix_json":     json.dumps(_gh_crop_mix) if _gh_mix_valid else None,
                "notes":             None,
            }
            if _gh_active:
                st.markdown(f"**Update** existing farm **{_gh_active['name']}**, or save as a new profile.")
                _gu1, _gu2, _gu3 = st.columns([2, 2, 1])
                with _gu1:
                    if st.button("✅ Update existing farm", use_container_width=True, key="gh_update_btn"):
                        try: # Keep emoji in success message
                            supabase.table("farms").update(_gh_payload).eq("id", _gh_active["id"]).execute()
                            st.session_state["active_farm"] = {**_gh_active, **_gh_payload}
                            st.success(f"✅ Farm **{_gh_active['name']}** updated.")
                            if _climate_data.get("mean_annual_dli"):
                                st.caption(f"🌤️ Climate profile saved — Mean DLI: {_climate_data['mean_annual_dli']:.1f} mol/m²/day · Ambient temp: {_climate_data['ambient_temp_annual']:.1f}°C")
                            st.session_state["gh_show_save_form"] = False
                            st.rerun()
                        except Exception as e:
                            st.error(f"Update failed: {e}")
                with _gu2:
                    gh_farm_name = st.text_input("New farm name", key="gh_farm_name_input", placeholder="Enter name for new profile")
                    if st.button("➕ Save as new farm", use_container_width=True, key="gh_saveas_btn"):
                        if not gh_farm_name.strip():
                            st.error("Please enter a name for the new farm profile.") # Keep emoji in success message
                        else:
                            try:
                                supabase.table("farms").insert({**_gh_payload, "name": gh_farm_name.strip(), "owner_id": current_user()}).execute()
                                st.success(f"✅ New farm profile '{gh_farm_name.strip()}' saved.")
                                if _climate_data.get("mean_annual_dli"):
                                    st.caption(f"🌤️ Climate profile saved — Mean DLI: {_climate_data['mean_annual_dli']:.1f} mol/m²/day · Ambient temp: {_climate_data['ambient_temp_annual']:.1f}°C")
                                st.session_state["gh_show_save_form"] = False
                                st.rerun()
                            except Exception as e:
                                st.error(f"Could not save: {e}")
                with _gu3:
                    if st.button("✖ Cancel", use_container_width=True, key="gh_cancel_save"):
                        st.session_state["gh_show_save_form"] = False
                        st.rerun()
            else:
                st.markdown("**Save current configuration as a Farm Profile**")
                st.caption("Saves all parameters so you can track harvests in the Harvest Tracker.")
                gh_farm_name = st.text_input("Farm name", key="gh_farm_name_input") # Keep emoji in button
                _gn1, _gn2 = st.columns([3, 1])
                with _gn1:
                    if st.button("✅ Confirm Save", key="gh_confirm_save", use_container_width=True):
                        if not gh_farm_name.strip():
                            st.error("Please enter a farm name.")
                        else:
                            try:
                                supabase.table("farms").insert({**_gh_payload, "name": gh_farm_name.strip(), "owner_id": current_user()}).execute()
                                st.success(f"Farm profile '{gh_farm_name.strip()}' saved.")
                                if _climate_data.get("mean_annual_dli"):
                                    st.caption(f"🌤️ Climate profile saved — Mean DLI: {_climate_data['mean_annual_dli']:.1f} mol/m²/day · Ambient temp: {_climate_data['ambient_temp_annual']:.1f}°C")
                                st.session_state["gh_show_save_form"] = False
                                st.rerun()
                            except Exception as e:
                                st.error(f"Could not save farm profile: {e}")
                with _gn2:
                    if st.button("✖ Cancel", key="gh_cancel_new", use_container_width=True):
                        st.session_state["gh_show_save_form"] = False
                        st.rerun()

    st.divider()

    # ── EBITDA Bridge ─────────────────────────────────────────────────────────
    st.subheader("EBITDA Bridge")
    gh_bridge_labels = ["Revenue", "Variable", "Water", "Energy", "Labour", "Rent", "Maintenance", "EBITDA"]
    gh_bridge_values = [
        gh_r["annual_revenue"], -gh_r["annual_variable_cost"], -gh_r["annual_water_cost"],
        -gh_r["annual_energy_cost"], -gh_r["annual_labour_cost"],
        -gh_r["annual_rent"], -gh_r["annual_maintenance"], gh_r["ebitda"],
    ]
    gh_bar_colors = []
    for _i, _v in enumerate(gh_bridge_values):
        if _i == 0:
            gh_bar_colors.append("rgba(0,229,160,0.85)")
        elif _i == len(gh_bridge_values) - 1:
            gh_bar_colors.append("rgba(0,229,160,0.85)" if _v >= 0 else "rgba(255,77,77,0.85)")
        else:
            gh_bar_colors.append("rgba(255,77,77,0.6)")
    gh_fig_bridge = go.Figure(go.Bar(
        x=gh_bridge_labels, y=gh_bridge_values,
        marker_color=gh_bar_colors,
        text=[f"${abs(_v):,.0f}" for _v in gh_bridge_values],
        textposition="outside",
    ))
    gh_fig_bridge.update_layout(
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        font_color="#161a16", showlegend=False,
        yaxis=dict(showgrid=False), xaxis=dict(showgrid=False),
        height=380, margin=dict(t=30, b=20),
    )
    style_fig(gh_fig_bridge)
    st.plotly_chart(gh_fig_bridge, use_container_width=True)

    st.divider()

    # ── Cost + CAPEX donuts ───────────────────────────────────────────────────
    gh_col_cost, gh_col_capex = st.columns(2)
    with gh_col_cost:
        st.subheader("Annual Cost Breakdown")
        gh_fig_cost = go.Figure(go.Pie(
            labels=["Energy", "Labour", "Variable", "Water", "Maintenance", "Rent"],
            values=[gh_r["annual_energy_cost"], gh_r["annual_labour_cost"],
                    gh_r["annual_variable_cost"], gh_r["annual_water_cost"],
                    gh_r["annual_maintenance"], gh_r["annual_rent"]],
            hole=0.45,
            marker_colors=["#ff4d4d", "#ffc13d", "#00e5a0", "#4fc3f7", "#ba68c8", "#ef9a9a"],
        ))
        gh_fig_cost.update_layout(
            plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            font_color="#161a16", height=320, margin=dict(t=10, b=10),
        )
        style_fig(gh_fig_cost)
        st.plotly_chart(gh_fig_cost, use_container_width=True)
    with gh_col_capex:
        st.subheader("CAPEX Breakdown")
        gh_fig_capex = go.Figure(go.Pie(
            labels=["Structure", "Climate", "Irrigation", "Lighting", "Automation", "Real Estate"],
            values=[gh_r["structure_capex"], gh_r["climate_capex"], gh_r["irrigation_capex"],
                    gh_r["lighting_capex"], gh_r["automation_capex"], gh_r["real_estate_capex"]],
            hole=0.45,
            marker_colors=["#00e5a0", "#26c6da", "#66bb6a", "#ffa726", "#ab47bc", "#8d6e63"],
        ))
        gh_fig_capex.update_layout(
            plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            font_color="#161a16", height=320, margin=dict(t=10, b=10),
        )
        style_fig(gh_fig_capex)
        st.plotly_chart(gh_fig_capex, use_container_width=True)

    st.divider()

    # ── DCF chart ─────────────────────────────────────────────────────────────
    st.subheader("Cumulative NPV — 10-year DCF")
    gh_fig_dcf = go.Figure()
    gh_fig_dcf.add_trace(go.Scatter(
        x=[d["year"] for d in gh_r["dcf_cashflows"]],
        y=[d["cumulative_npv"] for d in gh_r["dcf_cashflows"]],
        mode="lines+markers", line=dict(color="#00e5a0", width=2),
        fill="tozeroy", fillcolor="rgba(0,229,160,0.1)",
    ))
    gh_fig_dcf.add_hline(y=0, line_dash="dash", line_color="rgba(255,255,255,0.3)")
    gh_fig_dcf.update_layout(
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        font_color="#161a16", height=300,
        xaxis=dict(title="Year", showgrid=False),
        yaxis=dict(title="Cumulative NPV ($)", showgrid=False),
        margin=dict(t=10, b=10),
    )
    style_fig(gh_fig_dcf)
    st.plotly_chart(gh_fig_dcf, use_container_width=True)

    st.divider()

    # ── Full results table ────────────────────────────────────────────────────
    st.subheader("Full Results")
    _gh_res_df = pd.DataFrame({
        "Metric": [
            "Effective Grow Area (m²)", "Gross Area (m²)", "Structure Type",
            "Cycles / Year", "Harvest Mode", "Total Annual kg", "Price ($/kg)",
            "Annual Revenue", "Energy Cost", "Variable Cost", "Water Cost",
            "Labour Cost", "Maintenance", "Rent", "Total Annual Costs",
            "EBITDA", "EBITDA Margin", "Total CAPEX", "Payback (years)",
            "Annual Labour Hours", "Annual kWh", "Net Income", "NPV (10yr)",
        ],
        "Value": [
            f"{gh_r['effective_grow_area']:,.0f}",
            f"{gh_r['gross_area']:,.0f}",
            gh_r["structure_type"],
            str(gh_r["cycles_per_year"]),
            gh_r["harvest_mode"],
            f"{gh_r['total_annual_kg']:,.0f}",
            f"${gh_r['effective_price']:.2f}",
            f"${gh_r['annual_revenue']:,.0f}",
            f"${gh_r['annual_energy_cost']:,.0f}",
            f"${gh_r['annual_variable_cost']:,.0f}",
            f"${gh_r['annual_water_cost']:,.0f}",
            f"${gh_r['annual_labour_cost']:,.0f}",
            f"${gh_r['annual_maintenance']:,.0f}",
            f"${gh_r['annual_rent']:,.0f}",
            f"${gh_r['total_annual_costs']:,.0f}",
            f"${gh_r['ebitda']:,.0f}",
            f"{gh_r['ebitda_margin']*100:.1f}%",
            f"${gh_r['total_capex']:,.0f}",
            f"{gh_r['payback_years']:.1f}" if gh_r["payback_years"] else "N/A",
            f"{gh_r['annual_labour_hours']:,.0f}",
            f"{gh_r['annual_kwh']:,.0f}",
            f"${gh_r['net_income']:,.0f}",
            f"${gh_r['npv']:,.0f}",
        ],
    })
    st.dataframe(
        _gh_res_df.style.apply(lambda r: [MATCH if r.name % 2 == 0 else ""] * len(r), axis=1),
        use_container_width=True, hide_index=True,
    )

    # ═══════════════════════════════════════════════════════════════════════════
    # COUNTRY & CROP COMPARISON
    # ═══════════════════════════════════════════════════════════════════════════
    st.divider()
    st.subheader("🌍 Viability Comparison")
    st.caption("All calculations use the current sidebar inputs. Only the dimension being compared changes.")

    gh_comp_tab1, gh_comp_tab2 = st.tabs(["Compare Countries", "Compare Crops"])

    with gh_comp_tab1:
        gh_country_metric = st.selectbox(
            "Rank by",
            ["EBITDA", "Energy % of Revenue", "Payback (years)", "EBITDA Margin (%)"],
            key="gh_country_metric_select",
        )
        gh_country_results = []
        for _cn in COUNTRIES.keys():
            _ci = dict(gh_inputs); _ci["country"] = _cn
            try:
                _cr = calculate_greenhouse(_ci)
                _ep = _cr["annual_energy_cost"]/_cr["annual_revenue"]*100 if _cr["annual_revenue"]>0 else 999
                gh_country_results.append({
                    "Country": _cn, "EBITDA": _cr["ebitda"],
                    "Energy % of Revenue": _ep,
                    "Payback (years)": _cr["payback_years"] if _cr["payback_years"] else 999,
                    "EBITDA Margin (%)": _cr["ebitda_margin"]*100,
                })
            except Exception:
                continue
        if gh_country_results:
            _gh_df_c = pd.DataFrame(gh_country_results)
            _gh_asc  = gh_country_metric in ("Energy % of Revenue","Payback (years)")
            _gh_df_c = _gh_df_c.sort_values(gh_country_metric, ascending=_gh_asc).reset_index(drop=True)
            def _gh_cbar(row, m):
                if m=="EBITDA": return "rgba(0,229,160,0.75)" if row["EBITDA"]>=0 else "rgba(255,77,77,0.75)"
                if m=="Energy % of Revenue": return "rgba(0,229,160,0.75)" if row["Energy % of Revenue"]<40 else "rgba(255,77,77,0.75)"
                if m=="Payback (years)": return "rgba(0,229,160,0.75)" if 0<row["Payback (years)"]<10 else "rgba(255,77,77,0.75)"
                return "rgba(0,229,160,0.75)" if row["EBITDA Margin (%)"]>=0 else "rgba(255,77,77,0.75)"
            _gh_bc = [_gh_cbar(r, gh_country_metric) for _,r in _gh_df_c.iterrows()]
            _gh_fig_c = go.Figure(go.Bar(
                x=_gh_df_c[gh_country_metric], y=_gh_df_c["Country"], orientation="h",
                marker_color=_gh_bc,
                text=_gh_df_c[gh_country_metric].apply(
                    lambda v: f"${v:,.0f}" if gh_country_metric=="EBITDA"
                    else (f"{v:.1f}%" if "%" in gh_country_metric
                    else (f"{v:.1f} yrs" if v<900 else "N/A"))),
                textposition="outside"))
            if gh_country_metric=="Energy % of Revenue":
                _gh_fig_c.add_vline(x=40, line_dash="dash", line_color="rgba(255,193,61,0.6)",
                                    annotation_text="40% threshold", annotation_font_color="#ffc13d")
            _gh_fig_c.update_layout(
                plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                font_color="#161a16", height=max(500,len(_gh_df_c)*22),
                margin=dict(l=10,r=100,t=20,b=20),
                xaxis=dict(showgrid=False,zeroline=False),
                yaxis=dict(showgrid=False,autorange="reversed"))
            style_fig(_gh_fig_c)
            st.plotly_chart(_gh_fig_c, use_container_width=True)
            _gh_viable_c = (_gh_df_c["Energy % of Revenue"]<40).sum()
            st.caption(f"**{_gh_viable_c} of {len(_gh_df_c)} countries** show energy below 40% of revenue for this crop and setup.")

    with gh_comp_tab2:
        _gh_crop_pool = list(GREENHOUSE_CROPS.keys()) if gh_crop_source=="Greenhouse" else list(POLYTUNNEL_CROPS.keys())
        gh_crop_metric = st.selectbox(
            "Rank by",
            ["EBITDA","EBITDA Margin (%)","Energy % of Revenue","Payback (years)"],
            key="gh_crop_metric_select",
        )
        gh_crop_results = []
        for _crn in _gh_crop_pool:
            _cri = dict(gh_inputs); _cri["crop"] = _crn; _cri["harvest_mode"] = "Single"; _cri["price_override"] = 0
            try:
                _crr = calculate_greenhouse(_cri)
                _ep2 = _crr["annual_energy_cost"]/_crr["annual_revenue"]*100 if _crr["annual_revenue"]>0 else 999
                gh_crop_results.append({
                    "Crop": _crn, "EBITDA": _crr["ebitda"],
                    "EBITDA Margin (%)": _crr["ebitda_margin"]*100,
                    "Energy % of Revenue": _ep2,
                    "Payback (years)": _crr["payback_years"] if _crr["payback_years"] else 999,
                })
            except Exception:
                continue
        if gh_crop_results:
            _gh_df_cr = pd.DataFrame(gh_crop_results)
            _gh_asc2  = gh_crop_metric in ("Energy % of Revenue","Payback (years)")
            _gh_df_cr = _gh_df_cr.sort_values(gh_crop_metric, ascending=_gh_asc2).reset_index(drop=True)
            def _gh_crbar(row, m):
                if m=="EBITDA": return "rgba(0,229,160,0.75)" if row["EBITDA"]>=0 else "rgba(255,77,77,0.75)"
                if m=="Energy % of Revenue": return "rgba(0,229,160,0.75)" if row["Energy % of Revenue"]<40 else "rgba(255,77,77,0.75)"
                if m=="Payback (years)": return "rgba(0,229,160,0.75)" if 0<row["Payback (years)"]<10 else "rgba(255,77,77,0.75)"
                return "rgba(0,229,160,0.75)" if row["EBITDA Margin (%)"]>=0 else "rgba(255,77,77,0.75)"
            _gh_brc = [_gh_crbar(r, gh_crop_metric) for _,r in _gh_df_cr.iterrows()]
            _gh_fig_cr = go.Figure(go.Bar(
                x=_gh_df_cr[gh_crop_metric], y=_gh_df_cr["Crop"], orientation="h",
                marker_color=_gh_brc, textposition="outside"))
            if gh_crop_metric=="Energy % of Revenue":
                _gh_fig_cr.add_vline(x=40, line_dash="dash", line_color="rgba(255,193,61,0.6)",
                                     annotation_text="40% threshold", annotation_font_color="#ffc13d")
            _gh_fig_cr.update_layout(
                plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                font_color="#161a16", height=max(400,len(_gh_df_cr)*22),
                margin=dict(l=10,r=100,t=20,b=20),
                xaxis=dict(showgrid=False,zeroline=False),
                yaxis=dict(showgrid=False,autorange="reversed"))
            style_fig(_gh_fig_cr)
            st.plotly_chart(_gh_fig_cr, use_container_width=True)
            _gh_viable_cr = (_gh_df_cr["Energy % of Revenue"]<40).sum()
            st.caption(f"**{_gh_viable_cr} of {len(_gh_df_cr)} crops** show energy below 40% of revenue. Normalised to Single harvest.")

    # ═══════════════════════════════════════════════════════════════════════════
    # SENSITIVITY ANALYSIS
    # ═══════════════════════════════════════════════════════════════════════════
    st.divider()
    st.subheader("🔬 Sensitivity Analysis")

    def _gh_run_mult(base, kwh_m=1.0, lab_m=1.0, yld_m=1.0, prc_m=1.0):
        import core.greenhouse_data_tables as _ghdt
        import copy as _copy
        _cn = base["country"]
        from core.data_tables import COUNTRIES as _CTRS
        import core.data_tables as _cdt
        _orig_c = _CTRS[_cn]
        _mod_c  = _copy.deepcopy(_orig_c)
        _mod_c["kwh"]    = _orig_c["kwh"]    * kwh_m
        _mod_c["labour"] = _orig_c["labour"] * lab_m
        _mod_i = _copy.deepcopy(base)
        if prc_m != 1.0 or yld_m != 1.0:
            _src = base.get("crop_source","greenhouse")
            _cd  = _ghdt.POLYTUNNEL_CROPS if _src=="polytunnel" else _ghdt.GREENHOUSE_CROPS
            _orig_crop = _cd[base["crop"]]
            _mod_crop  = _copy.deepcopy(_orig_crop)
            _mod_crop["yield"]    = _orig_crop["yield"]    * yld_m
            _mod_crop["yield_h2"] = _orig_crop["yield_h2"] * yld_m
            _mod_crop["yield_h3"] = _orig_crop["yield_h3"] * yld_m
            if base.get("price_override",0)>0:
                _mod_i["price_override"] = base["price_override"] * prc_m
            else:
                _base_price = _orig_crop[f"price_{base['price_scenario']}"]
                _mod_i["price_override"] = _base_price * prc_m
            _cd[base["crop"]] = _mod_crop
        _cdt.COUNTRIES[_cn] = _mod_c
        try:
            if _gh_mix_valid:
                _res = _run_multicrop_generic(_mod_i, _gh_mix,
                                              calculate_greenhouse, _gh_crop_data_dict)
            else:
                _res = calculate_greenhouse(_mod_i)
        finally:
            _cdt.COUNTRIES[_cn] = _orig_c
            _src2 = base.get("crop_source","greenhouse")
            _cd2  = _ghdt.POLYTUNNEL_CROPS if _src2=="polytunnel" else _ghdt.GREENHOUSE_CROPS
            if prc_m != 1.0 or yld_m != 1.0:
                _cd2[base["crop"]] = _orig_crop
        return _res

    st.markdown("#### Tornado Chart — EBITDA Sensitivity")
    st.caption("Each bar shows how EBITDA changes when one variable is stressed while all others stay at base values.")

    _gh_base_ebitda = gh_r["ebitda"]
    _gh_tvars = [
        {"label":"Energy Price",  "pess":_gh_run_mult(gh_inputs,kwh_m=1.50)["ebitda"], "opt":_gh_run_mult(gh_inputs,kwh_m=0.70)["ebitda"],
         "pess_label":"Energy price +50%","opt_label":"Energy price −30%","rationale":"Energy cost volatility is a primary greenhouse risk."},
        {"label":"Selling Price", "pess":_gh_run_mult(gh_inputs,prc_m=0.80)["ebitda"], "opt":_gh_run_mult(gh_inputs,prc_m=1.20)["ebitda"],
         "pess_label":"Selling price −20%","opt_label":"Selling price +20%","rationale":"Wholesale price volatility for fresh produce."},
        {"label":"Yield",         "pess":_gh_run_mult(gh_inputs,yld_m=0.80)["ebitda"], "opt":_gh_run_mult(gh_inputs,yld_m=1.20)["ebitda"],
         "pess_label":"Yield −20%","opt_label":"Yield +20%","rationale":"Actual yields vary with cultivar and growing conditions."},
        {"label":"Labour Cost",   "pess":_gh_run_mult(gh_inputs,lab_m=1.30)["ebitda"], "opt":_gh_run_mult(gh_inputs,lab_m=0.80)["ebitda"],
         "pess_label":"Labour cost +30%","opt_label":"Labour cost −20%","rationale":"Structural wage inflation vs automation gains."},
    ]
    for _tv in _gh_tvars:
        _tv["delta_pess"] = _tv["pess"] - _gh_base_ebitda
        _tv["delta_opt"]  = _tv["opt"]  - _gh_base_ebitda
        _tv["swing"]      = abs(_tv["delta_opt"] - _tv["delta_pess"])
    _gh_tvars.sort(key=lambda x: x["swing"], reverse=True)

    _gh_fig_torn = go.Figure()
    for _tv in _gh_tvars:
        _gh_fig_torn.add_trace(go.Bar(
            name="Pessimistic", y=[_tv["label"]], x=[_tv["delta_pess"]], orientation="h",
            marker_color="rgba(255,77,77,0.75)", showlegend=(_tv==_gh_tvars[0]),
            text=f"${_tv['delta_pess']:,.0f}", textposition="outside",
            customdata=[[_tv["pess_label"],_tv["rationale"]]],
            hovertemplate="<b>%{y}</b><br>EBITDA delta: %{x:$,.0f}<br><i>%{customdata[0]}</i><extra></extra>"))
        _gh_fig_torn.add_trace(go.Bar(
            name="Optimistic", y=[_tv["label"]], x=[_tv["delta_opt"]], orientation="h",
            marker_color="rgba(0,229,160,0.75)", showlegend=(_tv==_gh_tvars[0]),
            text=f"${_tv['delta_opt']:,.0f}", textposition="outside",
            customdata=[[_tv["opt_label"],_tv["rationale"]]],
            hovertemplate="<b>%{y}</b><br>EBITDA delta: %{x:$,.0f}<br><i>%{customdata[0]}</i><extra></extra>"))
    _gh_fig_torn.add_vline(x=0, line_dash="dash", line_color="rgba(255,255,255,0.4)")
    _gh_fig_torn.update_layout(
        barmode="overlay", plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        font_color="#161a16", height=320, margin=dict(l=10,r=80,t=20,b=20),
        xaxis=dict(title="EBITDA delta from base ($)",showgrid=False,zeroline=False),
        yaxis=dict(showgrid=False),
        legend=dict(orientation="h",yanchor="bottom",y=1.02,xanchor="right",x=1))
    style_fig(_gh_fig_torn)
    st.plotly_chart(_gh_fig_torn, use_container_width=True)

    st.divider()
    st.markdown("#### Scenario Comparison")
    st.caption("Define up to 4 named scenarios by adjusting multipliers. The Base Case always shows the current sidebar inputs.")

    if "gh_scenarios" not in st.session_state:
        st.session_state["gh_scenarios"] = []

    with st.expander("➕ Define a new scenario", expanded=len(st.session_state["gh_scenarios"])==0):
        _gh_sc_name = st.text_input("Scenario name", value="Scenario 1", key="gh_sc_name")
        _gh_sc1, _gh_sc2 = st.columns(2)
        with _gh_sc1:
            _gh_sc_energy = st.slider("Energy price multiplier",  0.3, 3.0, 1.0, 0.05, key="gh_sc_energy")
            _gh_sc_yield  = st.slider("Yield multiplier",         0.3, 2.0, 1.0, 0.05, key="gh_sc_yield")
        with _gh_sc2:
            _gh_sc_price  = st.slider("Selling price multiplier", 0.3, 2.0, 1.0, 0.05, key="gh_sc_price")
            _gh_sc_labour = st.slider("Labour cost multiplier",   0.3, 3.0, 1.0, 0.05, key="gh_sc_labour")

        if st.button("💾 Save Scenario", key="gh_save_sc", use_container_width=True):
            if len(st.session_state["gh_scenarios"]) >= 4:
                st.warning("Maximum 4 scenarios reached.")
            else:
                _gh_sc_res = _gh_run_mult(gh_inputs, kwh_m=_gh_sc_energy, lab_m=_gh_sc_labour,
                                          yld_m=_gh_sc_yield, prc_m=_gh_sc_price)
                st.session_state["gh_scenarios"].append({
                    "name": _gh_sc_name, "energy_mult": _gh_sc_energy, "yield_mult": _gh_sc_yield,
                    "price_mult": _gh_sc_price, "labour_mult": _gh_sc_labour, "result": _gh_sc_res})
                st.success(f"Scenario '{_gh_sc_name}' saved.")

        if st.session_state["gh_scenarios"]:
            if st.button("🗑️ Clear all scenarios", key="gh_clear_sc", use_container_width=True):
                st.session_state["gh_scenarios"] = []
                st.rerun()

    if st.session_state["gh_scenarios"]:
        for _idx, _sc in enumerate(st.session_state["gh_scenarios"]):
            _dc1, _dc2 = st.columns([8,1])
            with _dc1:
                st.caption(f"**{_sc['name']}** — energy ×{_sc['energy_mult']} / yield ×{_sc['yield_mult']} / price ×{_sc['price_mult']} / labour ×{_sc['labour_mult']}")
            with _dc2:
                if st.button("🗑️", key=f"gh_del_sc_{_idx}"):
                    st.session_state["gh_scenarios"].pop(_idx)
                    st.rerun()

        _gh_sc_names   = ["Base Case"] + [s["name"] for s in st.session_state["gh_scenarios"]]
        _gh_sc_results = [gh_r] + [s["result"] for s in st.session_state["gh_scenarios"]]

        _gh_fig_comp = go.Figure()
        _gh_comp_metrics = {
            "EBITDA":      [res["ebitda"]            for res in _gh_sc_results],
            "Revenue":     [res["annual_revenue"]     for res in _gh_sc_results],
            "Energy Cost": [res["annual_energy_cost"] for res in _gh_sc_results],
            "Labour Cost": [res["annual_labour_cost"] for res in _gh_sc_results],
        }
        for _i, (_mn, _mv) in enumerate(_gh_comp_metrics.items()):
            _gh_fig_comp.add_trace(go.Bar(name=_mn, x=_gh_sc_names, y=_mv,
                marker_color=["#00e5a0","#ffc13d","#4fc3f7","#ba68c8"][_i],
                text=[f"${v:,.0f}" for v in _mv], textposition="outside"))
        _gh_fig_comp.update_layout(
            barmode="group", plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            font_color="#161a16", height=420, margin=dict(t=30,b=20),
            xaxis=dict(showgrid=False), yaxis=dict(showgrid=False,title="$"),
            legend=dict(orientation="h",yanchor="bottom",y=1.02,xanchor="right",x=1))
        style_fig(_gh_fig_comp)
        st.plotly_chart(_gh_fig_comp, use_container_width=True)

        _gh_sc_rows = []
        for _sn, _sr in zip(_gh_sc_names, _gh_sc_results):
            _ep3 = f"{_sr['annual_energy_cost']/_sr['annual_revenue']*100:.1f}%" if _sr["annual_revenue"]>0 else "N/A"
            _gh_sc_rows.append({
                "Scenario": _sn,
                "Revenue":  f"${_sr['annual_revenue']:,.0f}",
                "Energy Cost": f"${_sr['annual_energy_cost']:,.0f}",
                "Labour Cost": f"${_sr['annual_labour_cost']:,.0f}",
                "EBITDA":   f"${_sr['ebitda']:,.0f}",
                "EBITDA Margin": f"{_sr['ebitda_margin']*100:.1f}%",
                "Payback (yrs)": f"{_sr['payback_years']:.1f}" if _sr["payback_years"] else "N/A",
                "Energy % of Revenue": _ep3,
            })
        _gh_sc_df = pd.DataFrame(_gh_sc_rows)

        def _gh_highlight_ep(row):
            style = severity_cell(row["Energy % of Revenue"], hi=60, mid=40)
            return [style] * len(row)

        st.dataframe(_gh_sc_df.style.apply(_gh_highlight_ep,axis=1),
                     use_container_width=True, hide_index=True)
    else:
        st.info("No scenarios saved yet. Use the form above to define and save your first scenario.")


elif modality in ("🐟 Decoupled Aquaponics", "♻️ Coupled Aquaponics"):

    _aq_mode  = "decoupled" if modality == "🐟 Decoupled Aquaponics" else "coupled"
    _aq_label = modality

    st.markdown(f"### {_aq_label}") # Remove emoji from markdown

    # ── Session state defaults ────────────────────────────────────────────────
    _AQ_DEFAULTS = {
        "aq_country":               "Germany",
        "aq_plant_crop_source":     "Greenhouse",
        "aq_plant_crop":            "Lettuce (Romaine)",
        "aq_plant_footprint":       1000,
        "aq_automation":            "Medium",
        "aq_price_scenario":        "base",
        "aq_harvest_mode":          "Single",
        "aq_packaging_cost":        0.15,
        "aq_loss_rate":             5.0,
        "aq_net_grow_factor":       90.0,
        "aq_walkways_factor":       10.0,
        "aq_water_price":           2.0,
        "aq_kwh_override":          0.0,
        "aq_rent_monthly":          0.0,
        "aq_real_estate_capex":     0.0,
        "aq_depreciation_years":    15,
        "aq_tax_rate":              25.0,
        "aq_ltv":                   60.0,
        "aq_interest_rate":         5.5,
        "aq_loan_term_years":       15,
        "aq_discount_rate":         8.0,
        "aq_species":               "Tilapia (Nile)",
        "aq_tank_volume_m3":        50.0,
        "aq_fish_price_scenario":   "base",
        "aq_target_temp_c":         26.0,
        "aq_fish_depreciation_years": 10,
        "aq_coupled_efficiency":    0.88,
        "aq_multi_crop":            False,
        "aq_crop_mix":              [{"crop": "Lettuce (Romaine)", "pct": 100}],
    }
    for _k, _v in _AQ_DEFAULTS.items():
        if _k not in st.session_state:
            st.session_state[_k] = _v
    if "aq_show_save_form" not in st.session_state:
        st.session_state["aq_show_save_form"] = False

    # ── Sidebar ───────────────────────────────────────────────────────────────
    with st.sidebar:
        st.header("Aquaponics Parameters")

        # Plant parameters
        st.subheader("Plant Side") # Remove emoji from subheader
        aq_country_list = list(COUNTRIES.keys())
        aq_country = st.selectbox("Country", aq_country_list,
            index=aq_country_list.index(st.session_state["aq_country"]) if st.session_state["aq_country"] in aq_country_list else 0,
            key="aq_country")

        # ── Energy & Labour reference rates (energy_labour module) ────────────
        _el_rates  = get_rates_for_country_name(aq_country)
        _el_e      = _el_rates["energy"]
        _el_l      = _el_rates["labour"]
        _model_kwh = COUNTRIES.get(aq_country, {}).get("kwh", 0)
        if _el_rates["iso"]:
            _el_delta = _el_e["industrial"] - _model_kwh
            _el_arrow = "▲" if _el_delta > 0.005 else ("▼" if _el_delta < -0.005 else "≈")
            _el_col   = "#d4a845" if _el_delta > 0.005 else ("#52a066" if _el_delta < -0.005 else "#7a8070")
            _el_live  = " · ⚡ Live" if _el_e.get("live") else ""
            _el_html  = (
                f"<div style='font-size:11px;color:#9ba390;margin:-4px 0 8px 0;"
                f"padding:6px 8px;background:#252a25;border-radius:3px;"
                f"border:1px solid #363c36;border-left:3px solid {_el_col};'>"
                f"<b>Ref. rates ({_el_rates['iso']})</b> &nbsp;&middot;&nbsp; "
                f"Electricity industrial: <b style='color:{_el_col};'>${_el_e['industrial']:.3f}/kWh {_el_arrow}</b> "
                f"<span style='color:#7a8070;'>(model ${_model_kwh:.3f})</span>"
                f"&nbsp;&middot;&nbsp; Labour: <b>${_el_l['industrial_loaded']:.0f}/hr</b>"
                f"{_el_live}</div>"
            )
            st.markdown(_el_html, unsafe_allow_html=True)

        aq_plant_crop_source = st.radio("Crop source", ["Greenhouse", "Polytunnel"],
            index=0 if st.session_state["aq_plant_crop_source"] == "Greenhouse" else 1,
            horizontal=True, key="aq_plant_crop_source")

        if _aq_mode == "coupled":
            _aq_allowed_crops = [c for c, v in CROP_NUTRIENT_DEMAND.items()
                                 if v.get("aquaponics_suitability") in ("high", "medium")]
            _aq_crop_list = [c for c in (GREENHOUSE_CROPS if aq_plant_crop_source == "Greenhouse"
                              else POLYTUNNEL_CROPS) if c in _aq_allowed_crops] # Remove emoji from caption
            if not _aq_crop_list:
                _aq_crop_list = list(GREENHOUSE_CROPS.keys())
            st.caption("♻️ Coupled mode: only aquaponics-compatible crops shown.")
        else:
            _aq_crop_list = list(GREENHOUSE_CROPS.keys() if aq_plant_crop_source == "Greenhouse"
                                 else POLYTUNNEL_CROPS.keys())

        # ── Multi-crop toggle ─────────────────────────────────────────────────
        aq_multi_crop_mode = st.toggle("Multi-crop mode (plant side)", value=False,
                                        key="aq_multi_crop")

        if not aq_multi_crop_mode:
            _aq_crop_def  = st.session_state["aq_plant_crop"]
            aq_plant_crop = st.selectbox("Plant crop", _aq_crop_list,
                index=_aq_crop_list.index(_aq_crop_def) if _aq_crop_def in _aq_crop_list else 0,
                key="aq_plant_crop")
        else:
            aq_plant_crop = _aq_crop_list[0]  # placeholder
            st.markdown("**Plant crop allocation** (must sum to 100%)")
            if "aq_crop_mix" not in st.session_state:
                st.session_state["aq_crop_mix"] = [{"crop": _aq_crop_list[0], "pct": 100}]
            _aq_mix = st.session_state["aq_crop_mix"]
            _aq_mix = [row for row in _aq_mix if row["crop"] in _aq_crop_list]
            if not _aq_mix:
                _aq_mix = [{"crop": _aq_crop_list[0], "pct": 100}]
            _aq_to_remove = None
            for _aci, _arow in enumerate(_aq_mix):
                _ac1, _ac2, _ac3 = st.columns([4, 2, 1])
                with _ac1:
                    _aq_mix[_aci]["crop"] = st.selectbox(
                        f"Crop {_aci+1}", _aq_crop_list,
                        index=_aq_crop_list.index(_arow["crop"]) if _arow["crop"] in _aq_crop_list else 0,
                        key=f"aq_mix_crop_{_aci}")
                with _ac2:
                    _aq_mix[_aci]["pct"] = st.number_input(
                        "%", min_value=1, max_value=100, value=int(_arow["pct"]),
                        step=1, key=f"aq_mix_pct_{_aci}")
                with _ac3:
                    if len(_aq_mix) > 1 and st.button("✕", key=f"aq_mix_del_{_aci}"):
                        _aq_to_remove = _aci
            if _aq_to_remove is not None:
                _aq_mix.pop(_aq_to_remove)
                st.session_state["aq_crop_mix"] = _aq_mix
                st.rerun()
            _aq_total_pct = sum(r["pct"] for r in _aq_mix)
            st.caption(f"Total allocated: **{_aq_total_pct}%**")
            if _aq_total_pct != 100:
                st.warning(f"⚠️ Must sum to 100%. Currently {_aq_total_pct}%.")
            if len(_aq_mix) < 6:
                if st.button("➕ Add crop", key="aq_mix_add"):
                    _aq_mix.append({"crop": _aq_crop_list[0], "pct": 1})
                    st.session_state["aq_crop_mix"] = _aq_mix
                    st.rerun()
            st.session_state["aq_crop_mix"] = _aq_mix

        aq_plant_footprint = st.number_input("Plant footprint (m²)",
            value=st.session_state["aq_plant_footprint"], step=100, min_value=50,
            key="aq_plant_footprint")

        _aq_cd = GREENHOUSE_CROPS.get(aq_plant_crop, POLYTUNNEL_CROPS.get(aq_plant_crop, {}))
        if _aq_cd.get("days_between", 0) > 0:
            _aq_hm_list = ["Single", "2 Harvests", "3 Harvests"]
            _aq_hm_def  = st.session_state["aq_harvest_mode"]
            aq_harvest_mode = st.selectbox("Harvest Mode", _aq_hm_list,
                index=_aq_hm_list.index(_aq_hm_def) if _aq_hm_def in _aq_hm_list else 0,
                key="aq_harvest_mode")
        else:
            st.selectbox("Harvest Mode", ["Single"], disabled=True)
            aq_harvest_mode = "Single"
            st.session_state["aq_harvest_mode"] = "Single"

        st.divider()

        # Fish parameters
        st.subheader("Fish Side") # Remove emoji from subheader
        _aq_species_list = list(FISH_SPECIES.keys())
        aq_species = st.selectbox("Fish species", _aq_species_list,
            index=_aq_species_list.index(st.session_state["aq_species"]) if st.session_state["aq_species"] in _aq_species_list else 0,
            key="aq_species")

        if aq_species == "Atlantic Salmon":
            if _aq_mode == "coupled":
                st.error("Salmon incompatible with coupled aquaponics (cold water ≤14°C vs shared loop).") # Remove emoji from error message
            else:
                st.warning("⚠️ Salmon needs cold water (8–14°C). High heating costs in temperate climates.")

        aq_tank_volume = st.number_input("Tank volume (m³)",
            value=st.session_state["aq_tank_volume_m3"], step=10.0, min_value=5.0,
            key="aq_tank_volume_m3")

        aq_system_scale = "Commercial-scale (>100m³)" if aq_tank_volume >= 100 else "Small-scale (<100m³)"
        st.session_state["aq_system_scale"] = aq_system_scale
        st.caption( # Remove emoji from caption
            f"⚙️ System scale: **{aq_system_scale}** (auto-selected based on tank volume). "
            f"{'Commercial scale applies lower CAPEX rates per m³ due to economies of scale.' if aq_tank_volume >= 100 else 'Small-scale applies higher CAPEX rates per m³. Cross the 100 m³ threshold to unlock commercial rates.'}"
        )

        aq_target_temp = st.number_input("Fish target temp (°C)",
            value=st.session_state["aq_target_temp_c"], step=1.0, min_value=4.0, max_value=35.0,
            key="aq_target_temp_c",
            help=f"Ambient in {aq_country}: {COUNTRY_AMBIENT_TEMP.get(aq_country, 15.0):.1f}°C")

        _aq_fps_list = ["base", "low", "high"]
        aq_fish_price_scenario = st.selectbox("Fish price scenario", _aq_fps_list,
            index=_aq_fps_list.index(st.session_state["aq_fish_price_scenario"]),
            key="aq_fish_price_scenario")

        st.divider()

        # Shared parameters
        st.subheader("Shared") # Remove emoji from subheader
        aq_automation = st.selectbox("Automation", ["None","Low","Medium","High"],
            index=["None","Low","Medium","High"].index(st.session_state["aq_automation"]),
            key="aq_automation")
        aq_price_scenario = st.selectbox("Plant price scenario", ["base","low","high"],
            index=["base","low","high"].index(st.session_state["aq_price_scenario"]),
            key="aq_price_scenario")

        if _aq_mode == "coupled":
            aq_coupled_efficiency = st.slider("Coupled yield efficiency",
                0.70, 1.00, float(st.session_state["aq_coupled_efficiency"]), 0.01,
                key="aq_coupled_efficiency",
                help="Yield reduction vs greenhouse due to pH/EC compromise. Default 0.88.")
        else:
            aq_coupled_efficiency = 0.88

        st.divider()
        st.subheader("Advanced") # Remove emoji from subheader
        aq_packaging_cost    = st.number_input("Packaging ($/kg)", value=st.session_state["aq_packaging_cost"], step=0.01, min_value=0.0, key="aq_packaging_cost")
        aq_loss_rate         = st.number_input("Plant loss rate (%)", value=st.session_state["aq_loss_rate"], step=0.5, min_value=0.0, max_value=100.0, key="aq_loss_rate")
        aq_net_grow_factor   = st.number_input("Net grow factor (%)", value=st.session_state["aq_net_grow_factor"], step=1.0, min_value=1.0, max_value=100.0, key="aq_net_grow_factor")
        aq_walkways_factor   = st.number_input("Walkways factor (%)", value=st.session_state["aq_walkways_factor"], step=1.0, min_value=0.0, max_value=50.0, key="aq_walkways_factor")
        aq_water_price       = st.number_input("Water price ($/m³)", value=st.session_state["aq_water_price"], step=0.1, min_value=0.0, key="aq_water_price")
        _aq_kwh_default      = COUNTRIES.get(aq_country, {}).get("kwh", 0.0)
        aq_kwh_override      = st.number_input(
            "Electricity price ($/kWh)",
            value=float(st.session_state.get("aq_kwh_override") or _aq_kwh_default),
            step=0.005, min_value=0.001, format="%.4f", key="aq_kwh_override",
            help=(
                f"Country default: ${_aq_kwh_default:.4f}/kWh. "
                "Override with your actual site tariff. "
                "Industrial rate is typically 30–60% lower than residential."
            )
        )
        aq_rent_monthly      = st.number_input("Monthly rent ($)", value=st.session_state["aq_rent_monthly"], step=100.0, min_value=0.0, key="aq_rent_monthly")
        aq_real_estate_capex = st.number_input("Real estate CAPEX ($)", value=st.session_state["aq_real_estate_capex"], step=10000.0, min_value=0.0, key="aq_real_estate_capex")

        st.divider()
        st.subheader("Financial Structure") # Remove emoji from subheader
        aq_dep_years  = st.number_input("Plant depreciation (yrs)", value=st.session_state["aq_depreciation_years"], step=1, min_value=1, key="aq_depreciation_years")
        aq_fish_dep   = st.number_input("Fish depreciation (yrs)", value=st.session_state["aq_fish_depreciation_years"], step=1, min_value=1, key="aq_fish_depreciation_years")
        aq_tax_rate   = st.number_input("Tax rate (%)", value=st.session_state["aq_tax_rate"], step=1.0, min_value=0.0, max_value=100.0, key="aq_tax_rate")
        aq_ltv        = st.number_input("LTV (%)", value=st.session_state["aq_ltv"], step=5.0, min_value=0.0, max_value=100.0, key="aq_ltv")
        aq_interest   = st.number_input("Interest rate (%)", value=st.session_state["aq_interest_rate"], step=0.1, min_value=0.0, key="aq_interest_rate")
        aq_loan_term  = st.number_input("Loan term (yrs)", value=st.session_state["aq_loan_term_years"], step=1, min_value=1, key="aq_loan_term_years")
        aq_discount   = st.number_input("Discount rate (%)", value=st.session_state["aq_discount_rate"], step=0.5, min_value=0.0, key="aq_discount_rate")

    _aq_multi_crop_mode = st.session_state.get("aq_multi_crop", False)
    _aq_mix             = st.session_state.get("aq_crop_mix", [])
    _aq_valid_dict_early = POLYTUNNEL_CROPS if aq_plant_crop_source == "Polytunnel" else GREENHOUSE_CROPS
    _aq_mix = [row for row in _aq_mix if row["crop"] in _aq_valid_dict_early]
    _aq_mix_total       = sum(row["pct"] for row in _aq_mix)
    _aq_mix_valid       = _aq_multi_crop_mode and len(_aq_mix) > 0 and _aq_mix_total == 100

    # ── RUN CALCULATION ───────────────────────────────────────────────────────
    if _aq_mode == "coupled" and aq_species == "Atlantic Salmon":
        st.error("Cannot run: Atlantic Salmon is incompatible with coupled aquaponics. " # Remove emoji from error message
                 "Select a different species or switch to Decoupled mode.")
        st.stop()

    aq_inputs = {
        "aquaponics_mode":          _aq_mode,
        "country":                  aq_country,
        "plant_crop":               aq_plant_crop,
        "plant_crop_source":        aq_plant_crop_source.lower(),
        "plant_footprint":          aq_plant_footprint,
        "automation":               aq_automation,
        "price_scenario":           aq_price_scenario,
        "plant_price_override":     0.0,
        "packaging_cost":           aq_packaging_cost,
        "loss_rate":                aq_loss_rate,
        "net_grow_factor":          aq_net_grow_factor,
        "walkways_factor":          aq_walkways_factor,
        "water_price":              aq_water_price,
        "rent_monthly":             aq_rent_monthly,
        "real_estate_capex":        aq_real_estate_capex,
        "harvest_mode":             aq_harvest_mode,
        "depreciation_years":       aq_dep_years,
        "tax_rate":                 aq_tax_rate,
        "ltv":                      aq_ltv,
        "interest_rate":            aq_interest,
        "loan_term_years":          aq_loan_term,
        "discount_rate":            aq_discount,
        "ambient_temp_annual":    st.session_state.get("active_farm", {}).get("ambient_temp_annual"),
        "mean_annual_dli":        st.session_state.get("active_farm", {}).get("mean_annual_dli"),
        "species":                  aq_species,
        "tank_volume_m3":           aq_tank_volume,
        "system_scale":             aq_system_scale,
        "fish_price_scenario":      aq_fish_price_scenario,
        "price_override_fish":      0.0,
        "target_temp_c":            aq_target_temp,
        "fish_depreciation_years":  aq_fish_dep,
        "coupled_efficiency_factor": aq_coupled_efficiency,
        "crop_mix_json":            json.dumps(_aq_mix) if _aq_mix_valid else None,
    }

    if _aq_multi_crop_mode and not _aq_mix_valid:
        st.warning("⚠️ Fix plant crop allocation (must sum to 100%) before results are shown.")
        st.stop()

    # Define plant-side crop dictionary for sensitivity analysis
    _aq_plant_dict = POLYTUNNEL_CROPS if aq_inputs.get("plant_crop_source") == "polytunnel" else GREENHOUSE_CROPS

    # Define plant-side inputs for sensitivity analysis
    # Apply electricity price override
    _aq_kwh_original = COUNTRIES[aq_country]["kwh"]
    if abs(aq_kwh_override - _aq_kwh_original) > 0.0001:
        COUNTRIES[aq_country]["kwh"] = aq_kwh_override

    _aq_plant_sens_inputs = {
        "country": aq_inputs["country"], "crop": aq_inputs["plant_crop"],
        "crop_source": aq_inputs.get("plant_crop_source","greenhouse"),
        "footprint": aq_inputs["plant_footprint"],
        "automation": aq_inputs["automation"],
        "price_scenario": aq_inputs["price_scenario"], "price_override": 0.0,
        "packaging_cost": aq_inputs["packaging_cost"],
        "loss_rate": aq_inputs["loss_rate"],
        "net_grow_factor": aq_inputs["net_grow_factor"],
        "walkways_factor": aq_inputs["walkways_factor"],
        "water_price": aq_inputs["water_price"],
        "rent_monthly": aq_inputs["rent_monthly"],
        "real_estate_capex": aq_inputs["real_estate_capex"],
        "harvest_mode": aq_inputs["harvest_mode"],
        "depreciation_years": aq_inputs["depreciation_years"],
        "tax_rate": aq_inputs["tax_rate"], "ltv": aq_inputs["ltv"],
        "interest_rate": aq_inputs["interest_rate"],
        "loan_term_years": aq_inputs["loan_term_years"],
        "discount_rate": aq_inputs["discount_rate"],
    }

    # Define sensitivity run helper function
    def _aq_run_mult(base_plant_inputs, kwh_m=1.0, lab_m=1.0, yld_m=1.0, prc_m=1.0):
        import core.greenhouse_data_tables as _ghdt
        import copy as _copy
        import core.data_tables as _cdt
        _cn = base_plant_inputs["country"]
        _orig_c = _cdt.COUNTRIES[_cn]
        _mod_c  = _copy.deepcopy(_orig_c)
        _mod_c["kwh"]    = _orig_c["kwh"]    * kwh_m
        _mod_c["labour"] = _orig_c["labour"] * lab_m
        _mod_i = _copy.deepcopy(base_plant_inputs)
        _src   = base_plant_inputs.get("crop_source","greenhouse")
        _cd    = _ghdt.POLYTUNNEL_CROPS if _src=="polytunnel" else _ghdt.GREENHOUSE_CROPS
        if prc_m != 1.0 or yld_m != 1.0:
            _orig_crop = _cd[base_plant_inputs["crop"]]
            _mod_crop  = _copy.deepcopy(_orig_crop)
            _mod_crop["yield"]    = _orig_crop["yield"]    * yld_m
            _mod_crop["yield_h2"] = _orig_crop["yield_h2"] * yld_m
            _mod_crop["yield_h3"] = _orig_crop["yield_h3"] * yld_m
            if base_plant_inputs.get("price_override",0)>0:
                _mod_i["price_override"] = base_plant_inputs["price_override"] * prc_m
            else:
                _base_price = _orig_crop[f"price_{base_plant_inputs['price_scenario']}"]
                _mod_i["price_override"] = _base_price * prc_m
            _cd[base_plant_inputs["crop"]] = _mod_crop
        _cdt.COUNTRIES[_cn] = _mod_c
        try:
            if _aq_mix_valid:
                _res = _run_multicrop_generic(_mod_i, _aq_mix,
                                              calculate_greenhouse, _aq_plant_dict)
            else:
                _res = calculate_greenhouse(_mod_i)
        finally:
            _cdt.COUNTRIES[_cn] = _orig_c
            if prc_m != 1.0 or yld_m != 1.0:
                _cd[base_plant_inputs["crop"]] = _orig_crop
        return _res

    if _aq_mix_valid:
        # Run plant side as multi-crop, fish side as single species
        _aq_plant_dict = POLYTUNNEL_CROPS if aq_inputs.get("plant_crop_source") == "polytunnel" else GREENHOUSE_CROPS
        # Build plant-only inputs for multi-crop engine
        _aq_plant_base = {
            "country":            aq_inputs["country"],
            "crop":               _aq_mix[0]["crop"],
            "crop_source":        aq_inputs.get("plant_crop_source", "greenhouse"),
            "footprint":          aq_inputs["plant_footprint"],
            "automation":         aq_inputs["automation"],
            "price_scenario":     aq_inputs["price_scenario"],
            "plant_price_override": 0.0,
            "packaging_cost":     aq_inputs["packaging_cost"],
            "loss_rate":          aq_inputs["loss_rate"],
            "net_grow_factor":    aq_inputs["net_grow_factor"],
            "walkways_factor":    aq_inputs["walkways_factor"],
            "water_price":        aq_inputs["water_price"],
            "rent_monthly":       aq_inputs["rent_monthly"],
            "real_estate_capex":  aq_inputs["real_estate_capex"],
            "harvest_mode":       aq_inputs["harvest_mode"],
            "depreciation_years": aq_inputs["depreciation_years"],
            "tax_rate":           aq_inputs["tax_rate"],
            "ltv":                aq_inputs["ltv"],
            "interest_rate":      aq_inputs["interest_rate"],
            "loan_term_years":    aq_inputs["loan_term_years"],
            "discount_rate":      aq_inputs["discount_rate"],
            "ambient_temp_annual": aq_inputs.get("ambient_temp_annual"),
            "mean_annual_dli":    aq_inputs.get("mean_annual_dli"),
        }
        _aq_plant_r = _run_multicrop_generic(
            _aq_plant_base, _aq_mix, calculate_greenhouse, _aq_plant_dict)
        # Run fish side normally via calculate_aquaponics, extract fish result
        _aq_single_r = calculate_aquaponics(aq_inputs)
        _fr_multi     = _aq_single_r["fish"]
        # Combine
        aq_r = dict(_aq_single_r)
        aq_r["plant"]             = _aq_plant_r
        aq_r["combined_revenue"]  = _aq_plant_r["annual_revenue"]  + _fr_multi["annual_fish_revenue"]
        aq_r["combined_ebitda"]   = _aq_plant_r["ebitda"]          + _fr_multi["fish_ebitda"]
        aq_r["combined_capex"]    = _aq_plant_r["total_capex"]      + _fr_multi["total_fish_capex"] + aq_r["integration_capex"]
        aq_r["combined_ebitda_margin"] = (aq_r["combined_ebitda"] / aq_r["combined_revenue"]
                                           if aq_r["combined_revenue"] > 0 else 0.0)
        aq_r["_is_multicrop"]     = True
        aq_r["_crop_results"]     = _aq_plant_r.get("_crop_results", [])
    else:
        aq_r = calculate_aquaponics(aq_inputs)
    COUNTRIES[aq_country]["kwh"] = _aq_kwh_original  # always restore

    _pr  = aq_r["plant"]
    _fr  = aq_r["fish"]

    # ── WARNINGS ──────────────────────────────────────────────────────────────
    # ── Climate profile display ─────────────────────────────────────────────────
    _aq_farm_active = st.session_state.get("active_farm")
    if _aq_farm_active and _aq_farm_active.get("ambient_temp_annual"):
        st.caption(
            f"🌤️ **Climate profile active** — "
            f"Ambient temperature: {_aq_farm_active['ambient_temp_annual']:.1f}°C "
            f"(used for fish heating calculation)"
            + (f" · Mean annual DLI: {_aq_farm_active['mean_annual_dli']:.1f} mol/m²/day"
               if _aq_farm_active.get("mean_annual_dli") else "")
        )

    # ── Data Sources panel ───────────────────────────────────────────────────
    _aq_has_climate  = bool(st.session_state.get("active_farm", {}).get("ambient_temp_annual"))
    _aq_active_data  = st.session_state.get("active_farm") or {}
    with st.expander("ℹ️ Data sources & calculation transparency", expanded=False):
        _aqi1, _aqi2 = st.columns(2)
        with _aqi1:
            st.markdown("**📡 Automatic — from Open-Meteo Archive API**")
            if _aq_has_climate:
                _aq_temp2 = _aq_active_data.get("ambient_temp_annual", 0)
                _aq_dli2  = _aq_active_data.get("mean_annual_dli")
                _aq_spec  = aq_inputs.get("fish_species", "Tilapia (Nile)")
                _aq_tgt   = aq_inputs.get("fish_target_temp", 25)
                _aq_dt    = max(0, _aq_tgt - _aq_temp2)
                st.markdown(
                    f"- **Ambient temperature: {_aq_temp2:.1f}°C** "
                    f"— fish tank heating ΔT = max(0, {_aq_tgt}°C target − {_aq_temp2:.1f}°C ambient) = **{_aq_dt:.1f}°C**. "
                    f"Heating energy = 10 W/m³ × tank volume × 8,760 hrs ÷ 1,000 × (ΔT ÷ 15). "
                    f"See Assumptions §12.2.\n"
                    + (f"- **Mean annual DLI: {_aq_dli2:.1f} mol/m²/day** "
                       f"— drives plant zone supplemental lighting energy. See Assumptions §11.2.\n"
                       if _aq_dli2 else "- Mean annual DLI: not available for this farm.\n")
                    + "- Source: Open-Meteo 10-year historical archive. Stored in Supabase at farm save time."
                )
            else:
                st.markdown(
                    "- ⚠️ **No climate data available** for this farm.\n"
                    "- Set coordinates in the **Farm Intelligence Map**, then re-save the farm profile.\n"
                    "- Until then, fish heating uses country-level ambient temperature fallback (see Assumptions §14)."
                )
        with _aqi2:
            st.markdown("**🎛️ Manual inputs — set in this calculator**")
            st.markdown(
                "- Fish species → target water temperature, O₂ demand, water exchange rate, grow cycle\n"
                "- Tank volume, system mode (coupled / decoupled)\n"
                "- Coupled: yield multiplier (default 0.88), near-zero nutrient cost (5%)\n"
                "- Decoupled: 60% nutrient offset from fish effluent\n"
                "- Country → electricity price. See Assumptions §12 and §13 for full detail."
            )
        st.caption(
            "ℹ️ Fish heating: 10 W/m³ × tank_vol × 8,760 hrs ÷ 1,000 × (ΔT ÷ 15). "
            "Fish aeration: base kWh/kg × O₂ scale factor × annual kg fish. "
            "Full model in Assumptions §12."
        )


    # ── Energy & Labour calibration callout ──────────────────────────────────
    _el_r2    = get_rates_for_country_name(aq_inputs["country"])
    _el_e2    = _el_r2["energy"]
    _el_l2    = _el_r2["labour"]
    _mkwh2    = COUNTRIES.get(aq_inputs["country"], {}).get("kwh", 0)
    _mlabour2 = COUNTRIES.get(aq_inputs["country"], {}).get("labour", 0)
    if _el_r2["iso"]:
        _e_diff   = abs(_el_e2["industrial"] - _mkwh2)
        _l_diff   = abs(_el_l2["industrial_loaded"] - _mlabour2)
        _e_flag   = _e_diff > 0.01
        _l_flag   = _l_diff > 3.0
        _exp_label = (
            "⚠️ Verify your input assumptions"
            if (_e_flag or _l_flag) else
            "✅ Input assumptions cross-check"
        )
        with st.expander(_exp_label, expanded=(_e_flag or _l_flag)):
            _rc1, _rc2 = st.columns(2)
            with _rc1:
                st.markdown("**⚡ Electricity**")
                _e_arrow = "higher" if _el_e2["industrial"] > _mkwh2 else "lower"
                _e_pct   = abs(_el_e2["industrial"] - _mkwh2) / _mkwh2 * 100 if _mkwh2 else 0
                if _e_flag:
                    st.warning(
                        f"Reference industrial rate: **${_el_e2['industrial']:.3f}/kWh** "
                        f"({_e_pct:.0f}% {_e_arrow} than model's ${_mkwh2:.3f}/kWh). "
                        f"If your actual tariff differs, update the electricity price "
                        f"in the country table or use a price override."
                    )
                else:
                    st.success(
                        f"Model electricity price (${_mkwh2:.3f}/kWh) aligns with "
                        f"reference industrial rate (${_el_e2['industrial']:.3f}/kWh)."
                    )
                if _el_e2.get("live"):
                    st.caption(f"⚡ {_el_e2['live_note']}")
                else:
                    st.caption(f"Source: {_el_e2['source']} — static 2023–24 data")
            with _rc2:
                st.markdown("**👷 Labour**")
                _l_arrow = "higher" if _el_l2["industrial_loaded"] > _mlabour2 else "lower"
                _l_pct   = abs(_el_l2["industrial_loaded"] - _mlabour2) / _mlabour2 * 100 if _mlabour2 else 0
                if _l_flag:
                    st.warning(
                        f"Reference fully-loaded industrial rate: **${_el_l2['industrial_loaded']:.0f}/hr** "
                        f"({_l_pct:.0f}% {_l_arrow} than model's ${_mlabour2:.0f}/hr). "
                        f"Overhead multiplier applied: {_el_l2['overhead_pct']} "
                        f"(base ${_el_l2['industrial_base']:.0f}/hr × {_el_l2['overhead']:.2f})."
                    )
                else:
                    st.success(
                        f"Model labour rate (${_mlabour2:.0f}/hr) aligns with "
                        f"reference (${_el_l2['industrial_loaded']:.0f}/hr, overhead {_el_l2['overhead_pct']})."
                    )
                st.caption(f"Source: {_el_l2['source']}")


    # ── Energy & Labour calibration callout ──────────────────────────────────
    _el_r2    = get_rates_for_country_name(aq_inputs["country"])
    _el_e2    = _el_r2["energy"]
    _el_l2    = _el_r2["labour"]
    _mkwh2    = COUNTRIES.get(aq_inputs["country"], {}).get("kwh", 0)
    _mlabour2 = COUNTRIES.get(aq_inputs["country"], {}).get("labour", 0)
    if _el_r2["iso"]:
        _e_flag  = abs(_el_e2["industrial"] - _mkwh2) > 0.01
        _l_flag  = abs(_el_l2["industrial_loaded"] - _mlabour2) > 3.0
        _exp_lbl = "⚠️ Verify your input assumptions" if (_e_flag or _l_flag) else "✅ Input assumptions cross-check"
        with st.expander(_exp_lbl, expanded=(_e_flag or _l_flag)):
            _rc1, _rc2 = st.columns(2)
            with _rc1:
                st.markdown("**⚡ Electricity**")
                _e_dir = "higher" if _el_e2["industrial"] > _mkwh2 else "lower"
                _e_pct = abs(_el_e2["industrial"] - _mkwh2) / _mkwh2 * 100 if _mkwh2 else 0 # Keep warning emoji
                if _e_flag:
                    st.warning(
                        f"Reference industrial rate: **${_el_e2['industrial']:.3f}/kWh** "
                        f"({_e_pct:.0f}% {_e_dir} than model’s ${_mkwh2:.3f}/kWh). "
                        f"Verify country default or use a site-specific override if your tariff differs."
                    )
                else:
                    st.success(f"Model electricity (${_mkwh2:.3f}/kWh) aligns with reference industrial rate (${_el_e2['industrial']:.3f}/kWh).")
                st.caption(f"Source: {_el_e2['source']}" + (f" · {_el_e2['live_note']}" if _el_e2.get("live") else "")) # Remove emoji from caption
            with _rc2:
                st.markdown("**👷 Labour**")
                _l_dir = "higher" if _el_l2["industrial_loaded"] > _mlabour2 else "lower"
                _l_pct = abs(_el_l2["industrial_loaded"] - _mlabour2) / _mlabour2 * 100 if _mlabour2 else 0
                if _l_flag:
                    st.warning(
                        f"Reference fully-loaded industrial: **${_el_l2['industrial_loaded']:.0f}/hr** "
                        f"({_l_pct:.0f}% {_l_dir} than model’s ${_mlabour2:.0f}/hr). "
                        f"Overhead {_el_l2['overhead_pct']} applied (base ${_el_l2['industrial_base']:.0f}/hr)."
                    ) # Keep warning emoji
                else:
                    st.success(f"Model labour (${_mlabour2:.0f}/hr) aligns with reference (${_el_l2['industrial_loaded']:.0f}/hr, overhead {_el_l2['overhead_pct']}).")
                st.caption(f"Source: {_el_l2['source']}")

    if aq_r.get("salmon_warning"):
        st.warning(aq_r["salmon_warning"])
    if aq_r.get("ratio_warning"):
        st.warning(aq_r["ratio_warning"]) # Keep warning emoji
    if _aq_mode == "decoupled" and aq_r["nutrient_offset_saving"] > 0:
        st.success( # Keep emoji in success message
            f"🌿 Nutrient offset saving: **${aq_r['nutrient_offset_saving']:,.0f}/year** — "
            f"{aq_r['annual_n_output_g']/1000:.1f} kg N/yr from fish effluent "
            f"({COUPLING_PARAMS['decoupled_nutrient_offset_fraction']['base']*100:.0f}% offset applied)"
        )

    # ── Multi-crop plant breakdown ────────────────────────────────────────────
    if aq_r.get("_is_multicrop") and aq_r.get("_crop_results"):
        st.divider()
        st.subheader("Per-Crop Plant Breakdown") # Remove emoji from subheader
        _aq_mc_rows = []
        for _mc in aq_r["_crop_results"]:
            _aq_mc_rows.append({
                "Crop":           _mc["crop"],
                "Area %":         f"{_mc['pct']:.0f}%",
                "Annual kg":      f"{_mc['total_annual_kg']:,.0f}",
                "Price ($/kg)":   f"${_mc['effective_price']:.2f}",
                "Revenue":        f"${_mc['annual_revenue']:,.0f}",
                "Variable Cost":  f"${_mc['annual_variable_cost']:,.0f}",
                "Labour":         f"${_mc['annual_labour_cost']:,.0f}",
                "EBITDA contrib": f"${_mc['ebitda']:,.0f}",
            })
        st.dataframe(pd.DataFrame(_aq_mc_rows), use_container_width=True, hide_index=True)
        st.caption("Fish side is always single-species. Plant energy and CAPEX shared across the mix.")
        st.divider()
    # ── PDF BUTTON ────────────────────────────────────────────────────────────
    def generate_aq_pdf_report(aq_inputs: dict, aq_r: dict) -> bytes:
        _fn = st.session_state.get("active_farm", {}).get("name", "")
        _mc = "aqc" if aq_inputs.get("aquaponics_mode","").lower() in ("coupled","aqc") else "aqd"
        def _aq_sens(kwh_m=1.0, lab_m=1.0, yld_m=1.0, prc_m=1.0):
            return _aq_run_mult(_aq_plant_sens_inputs, kwh_m=kwh_m, lab_m=lab_m,
                                yld_m=yld_m, prc_m=prc_m)
        return _build_feasibility_pdf(aq_r, aq_inputs, _mc, farm_name=_fn,
                                      run_sens_fn=_aq_sens,
                                      aq_plant_sens_inputs=_aq_plant_sens_inputs)

    aq_pdf_col1, aq_pdf_col2 = st.columns([5, 1])
    with aq_pdf_col2:
        if st.button("📄 Download PDF Report", key="aq_pdf_btn", use_container_width=True): # Keep emoji in button
            with st.spinner("Generating PDF..."):
                _aq_pdf_bytes = generate_aq_pdf_report(aq_inputs, aq_r)
                # Correct naming: use primary crop from mix if multi-crop is valid
                _aq_rep_crop = _aq_mix[0]["crop"] if _aq_mix_valid and _aq_mix else aq_inputs.get('plant_crop','')
                _aq_rep_crop_safe = _aq_rep_crop.replace(' ','_').replace('/','').replace('(','').replace(')','')
                _aq_filename = (
                    f"AQ_Report_{_aq_rep_crop_safe}"
                    f"_{aq_inputs.get('species','').replace(' ','_')}"
                    f"_{aq_inputs.get('country','')}_{date.today().strftime('%Y%m%d')}.pdf"
                )
                st.download_button(label="⬇️ Save PDF", data=_aq_pdf_bytes,
                                   file_name=_aq_filename, mime="application/pdf",
                                   use_container_width=True, key="aq_pdf_dl")
    st.divider()

    # ── COMBINED KEY METRICS ──────────────────────────────────────────────────
    st.subheader("Combined System Metrics")
    # Break-even yield for plant side
    _aq_loss_r   = aq_inputs["loss_rate"] / 100
    _aq_be_denom = (_pr["effective_price"] * (1 - _aq_loss_r) *
                    _pr["cycles_per_year"] * _pr["effective_grow_area"])
    _aq_be_yield = _pr["total_annual_costs"] / _aq_be_denom if _aq_be_denom > 0 else None
    _aq_plant_dict_early = POLYTUNNEL_CROPS if aq_inputs.get("plant_crop_source","greenhouse")=="polytunnel" else GREENHOUSE_CROPS
    _aq_proj_yield = _aq_plant_dict_early.get(aq_plant_crop, {}).get("yield", 0) if not _aq_mix_valid else 0
    _aq_yield_gap_str = (f"{(_aq_proj_yield - _aq_be_yield) / _aq_be_yield * 100:+.1f}%"
                         if _aq_be_yield and _aq_proj_yield else "N/A")
    # DSCR combined
    _aq_combined_ds = _pr.get("annual_debt_service", 0) + _fr.get("annual_debt_service", 0)
    _aq_combined_dscr = aq_r["combined_ebitda"] / _aq_combined_ds if _aq_combined_ds > 0 else None
    am1, am2, am3, am4, am5, am6, am7 = st.columns(7)
    am1.metric("Combined Revenue",  f"${aq_r['combined_revenue']:,.0f}")
    am2.metric("Combined EBITDA",   f"${aq_r['combined_ebitda']:,.0f}")
    am3.metric("EBITDA Margin",     f"{aq_r['combined_ebitda_margin']*100:.1f}%")
    am4.metric("Combined CAPEX",    f"${aq_r['combined_capex']:,.0f}")
    am5.metric("Plant Revenue",     f"${_pr['annual_revenue']:,.0f}")
    am6.metric("Fish Revenue",      f"${_fr['annual_fish_revenue']:,.0f}")
    am7.metric("Plant Break-even",
               f"{_aq_be_yield:.2f} kg/m²/cycle" if _aq_be_yield else "N/A",
               delta=_aq_yield_gap_str if _aq_be_yield and not _aq_mix_valid else None,
               delta_color="normal",
               help="Minimum plant yield to cover all plant costs.")
    if _aq_combined_dscr is not None and _aq_combined_dscr < 1.0:
        st.warning( # Keep warning emoji
            f"⚠️ **Combined debt service coverage is low (DSCR = {_aq_combined_dscr:.2f}x).** "
            f"Total annual debt repayment exceeds combined EBITDA. "
            f"Consider reducing LTV or extending loan terms."
        )

    st.divider()

    # ── SIDE-BY-SIDE P&L ──────────────────────────────────────────────────────
    st.subheader("Plant vs Fish P&L")
    _aq_pc, _aq_fc = st.columns(2) # Remove emoji from markdown

    with _aq_pc:
        st.markdown(f"**🌿 Plant — {aq_plant_crop}**")
        if _aq_mode == "decoupled":
            st.caption(f"Nutrient saving included: ${aq_r['nutrient_offset_saving']:,.0f}/yr")
        else:
            st.caption(f"Coupled efficiency: {aq_coupled_efficiency:.0%} of greenhouse yield")
        st.dataframe(pd.DataFrame({
            "Item": ["Revenue","Energy","Variable","Water","Labour","Maintenance","Rent","EBITDA","EBITDA Margin"],
            "$/year": [f"${_pr['annual_revenue']:,.0f}", f"${_pr['annual_energy_cost']:,.0f}",
                       f"${_pr['annual_variable_cost']:,.0f}", f"${_pr['annual_water_cost']:,.0f}",
                       f"${_pr['annual_labour_cost']:,.0f}", f"${_pr['annual_maintenance']:,.0f}",
                       f"${_pr['annual_rent']:,.0f}", f"${_pr['ebitda']:,.0f}",
                       f"{_pr['ebitda_margin']*100:.1f}%"],
        }), use_container_width=True, hide_index=True)

    with _aq_fc:
        st.markdown(f"**🐟 Fish — {aq_species}**")
        st.caption(f"{_fr['annual_kg_fish']:,.0f} kg/yr · {_fr['cycles_per_year']} cycle(s) · ΔT={_fr['delta_t']:.0f}°C") # Remove emoji from markdown
        st.dataframe(pd.DataFrame({
            "Item": ["Revenue","Feed","Fingerlings","Energy","Water","Labour","Maintenance","EBITDA","EBITDA Margin"],
            "$/year": [f"${_fr['annual_fish_revenue']:,.0f}", f"${_fr['annual_feed_cost']:,.0f}",
                       f"${_fr['annual_fingerling_cost']:,.0f}", f"${_fr['annual_fish_energy_cost']:,.0f}",
                       f"${_fr['annual_water_cost']:,.0f}", f"${_fr['annual_fish_labour_cost']:,.0f}",
                       f"${_fr['annual_fish_maintenance']:,.0f}", f"${_fr['fish_ebitda']:,.0f}",
                       f"{_fr['fish_ebitda_margin']*100:.1f}%"],
        }), use_container_width=True, hide_index=True)

    st.divider()

    # ── COMBINED EBITDA BRIDGE ────────────────────────────────────────────────
    st.subheader("Combined EBITDA Bridge")
    _aq_bl = ["Plant Revenue","Fish Revenue","Plant Costs","Fish Costs","Nutrient Saving","Combined EBITDA"] # No emojis in _aq_bl
    _aq_bv = [_pr["annual_revenue"], _fr["annual_fish_revenue"],
               -_pr["total_annual_costs"], -_fr["total_fish_costs"],
               aq_r["nutrient_offset_saving"], aq_r["combined_ebitda"]]
    _aq_bc = ["rgba(0,229,160,0.85)","rgba(0,229,160,0.85)","rgba(255,77,77,0.6)","rgba(255,77,77,0.6)",
               "rgba(79,195,247,0.85)",
               "rgba(0,229,160,0.85)" if aq_r["combined_ebitda"] >= 0 else "rgba(255,77,77,0.85)"]
    _aq_fig_b = go.Figure(go.Bar(x=_aq_bl, y=_aq_bv, marker_color=_aq_bc,
        text=[f"${abs(v):,.0f}" for v in _aq_bv], textposition="outside"))
    _aq_fig_b.update_layout(plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        font_color="#161a16", showlegend=False, height=380, margin=dict(t=30,b=20),
        yaxis=dict(showgrid=False), xaxis=dict(showgrid=False))
    style_fig(_aq_fig_b)
    st.plotly_chart(_aq_fig_b, use_container_width=True)

    st.divider()

    # ── CAPEX ─────────────────────────────────────────────────────────────────
    st.subheader("CAPEX Breakdown")
    _aq_cc1, _aq_cc2 = st.columns(2) # Remove emoji from markdown
    with _aq_cc1:
        st.markdown(f"**🌿 Plant CAPEX — ${_pr['total_capex']:,.0f}**")
        _pcf = go.Figure(go.Pie(
            labels=["Structure","Climate","Irrigation","Lighting","Automation","Real Estate"],
            values=[_pr["structure_capex"],_pr["climate_capex"],_pr["irrigation_capex"],
                    _pr["lighting_capex"],_pr["automation_capex"],_pr["real_estate_capex"]],
            hole=0.45, marker_colors=["#00e5a0","#26c6da","#66bb6a","#ffa726","#ab47bc","#8d6e63"]))
        _pcf.update_layout(plot_bgcolor="#ffffff",paper_bgcolor="#ffffff",
                           font_color="#161a16",height=300,margin=dict(t=10,b=10))
        style_fig(_pcf)
        st.plotly_chart(_pcf, use_container_width=True)
    with _aq_cc2:
        st.markdown(f"**🐟 Fish CAPEX — ${_fr['total_fish_capex']:,.0f} + Integration ${aq_r['integration_capex']:,.0f}**") # Remove emoji from markdown
        _fcf = go.Figure(go.Pie(
            labels=["Tanks","Filtration","Aeration","Monitoring","Plumbing"],
            values=[_fr["tank_capex"],_fr["filtration_capex"],_fr["aeration_capex"],
                    _fr["monitoring_capex"],_fr["plumbing_capex"]],
            hole=0.45, marker_colors=["#4fc3f7","#29b6f6","#0288d1","#01579b","#80d8ff"]))
        _fcf.update_layout(plot_bgcolor="#ffffff",paper_bgcolor="#ffffff",
                           font_color="#161a16",height=300,margin=dict(t=10,b=10))
        style_fig(_fcf)
        st.plotly_chart(_fcf, use_container_width=True)

    st.divider()

    # ── Annual cost breakdown (plant side) ───────────────────────────────────
    st.subheader("Plant Annual Cost Breakdown")
    _aq_cost_fig = go.Figure(go.Pie( # Remove emoji from markdown
        labels=["Energy", "Variable", "Water", "Labour", "Maintenance", "Rent"],
        values=[_pr["annual_energy_cost"], _pr["annual_variable_cost"],
                _pr["annual_water_cost"], _pr["annual_labour_cost"],
                _pr["annual_maintenance"], _pr["annual_rent"]],
        hole=0.45,
        marker_colors=["#ff4d4d", "#ffc13d", "#00e5a0", "#4fc3f7", "#ba68c8", "#ef9a9a"]))
    _aq_cost_fig.update_layout(
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        font_color="#161a16", height=320, margin=dict(t=10, b=10))
    style_fig(_aq_cost_fig)
    st.plotly_chart(_aq_cost_fig, use_container_width=True)

    st.divider()

    # ── DCF ───────────────────────────────────────────────────────────────────
    st.subheader("Cumulative NPV — 10-year DCF")
    _dcf_col1, _dcf_col2 = st.columns(2)

    with _dcf_col1: # Remove emoji from markdown
        st.markdown("**🌿 Plant Side**")
        _pdcf = go.Figure()
        _pdcf.add_trace(go.Scatter(
            x=[d["year"] for d in _pr["dcf_cashflows"]],
            y=[d["cumulative_npv"] for d in _pr["dcf_cashflows"]],
            mode="lines+markers", line=dict(color="#00e5a0", width=2),
            fill="tozeroy", fillcolor="rgba(0,229,160,0.1)"))
        _pdcf.add_hline(y=0, line_dash="dash", line_color="rgba(255,255,255,0.3)")
        _pdcf.update_layout(plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            font_color="#161a16", height=280,
            xaxis=dict(title="Year", showgrid=False),
            yaxis=dict(title="Cumulative NPV ($)", showgrid=False),
            margin=dict(t=10, b=10))
        style_fig(_pdcf)
        st.plotly_chart(_pdcf, use_container_width=True)

    with _dcf_col2: # Remove emoji from markdown
        st.markdown("**🐟 Fish Side**")
        _fdcf = go.Figure()
        _fdcf.add_trace(go.Scatter(
            x=[d["year"] for d in _fr["dcf_cashflows"]],
            y=[d["cumulative_npv"] for d in _fr["dcf_cashflows"]],
            mode="lines+markers", line=dict(color="#4fc3f7", width=2),
            fill="tozeroy", fillcolor="rgba(79,195,247,0.1)"))
        _fdcf.add_hline(y=0, line_dash="dash", line_color="rgba(255,255,255,0.3)")
        _fdcf.update_layout(plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            font_color="#161a16", height=280,
            xaxis=dict(title="Year", showgrid=False),
            yaxis=dict(title="Cumulative NPV ($)", showgrid=False),
            margin=dict(t=10, b=10))
        style_fig(_fdcf)
        st.plotly_chart(_fdcf, use_container_width=True)

    st.divider()

    # ── FULL RESULTS ──────────────────────────────────────────────────────────
    st.subheader("Full Results")
    _frt1, _frt2 = st.tabs(["🌿 Plant Detail", "🐟 Fish Detail"]) # Keep emojis in tab labels
    with _frt1:
        _aq_p_df = pd.DataFrame({
            "Metric": ["Effective Grow Area (m²)","Gross Area (m²)","Structure Type",
                       "Cycles/Year","Harvest Mode","Annual kg","Price ($/kg)",
                       "Revenue","Energy","Variable","Water","Labour","Maintenance","Rent",
                       "Total Costs","EBITDA","EBITDA Margin","Plant CAPEX","Annual kWh"],
            "Value": [f"{_pr['effective_grow_area']:,.0f}", f"{_pr['gross_area']:,.0f}",
                      _pr["structure_type"], _pr["cycles_per_year"], _pr["harvest_mode"],
                      f"{_pr['total_annual_kg']:,.0f}", f"${_pr['effective_price']:.2f}",
                      f"${_pr['annual_revenue']:,.0f}", f"${_pr['annual_energy_cost']:,.0f}",
                      f"${_pr['annual_variable_cost']:,.0f}", f"${_pr['annual_water_cost']:,.0f}",
                      f"${_pr['annual_labour_cost']:,.0f}", f"${_pr['annual_maintenance']:,.0f}",
                      f"${_pr['annual_rent']:,.0f}", f"${_pr['total_annual_costs']:,.0f}",
                      f"${_pr['ebitda']:,.0f}", f"{_pr['ebitda_margin']*100:.1f}%",
                      f"${_pr['total_capex']:,.0f}", f"{_pr['annual_kwh']:,.0f}"],
        })
        st.dataframe(
            _aq_p_df.style.apply(lambda r: [MATCH if r.name % 2 == 0 else ""] * len(r), axis=1),
            use_container_width=True, hide_index=True,
        )
    with _frt2:
        _aq_f_df = pd.DataFrame({
            "Metric": ["Species","Tank Volume (m³)","System Scale","Harvest Biomass (kg)",
                       "Cycles/Year","Annual kg Fish","Price ($/kg)",
                       "Revenue","Feed","Fingerlings","Energy","Water","Labour","Maintenance",
                       "Total Costs","EBITDA","EBITDA Margin","Fish CAPEX",
                       "Annual kWh","Heating kWh","Ambient Temp (°C)","Target Temp (°C)",
                       "Annual N Output (kg)"],
            "Value": [_fr["species"], f"{_fr['tank_volume_m3']:.0f}", _fr["system_scale"],
                      f"{_fr['harvest_biomass_kg']:,.0f}", _fr["cycles_per_year"],
                      f"{_fr['annual_kg_fish']:,.0f}", f"${_fr['effective_fish_price']:.2f}",
                      f"${_fr['annual_fish_revenue']:,.0f}", f"${_fr['annual_feed_cost']:,.0f}",
                      f"${_fr['annual_fingerling_cost']:,.0f}", f"${_fr['annual_fish_energy_cost']:,.0f}",
                      f"${_fr['annual_water_cost']:,.0f}", f"${_fr['annual_fish_labour_cost']:,.0f}",
                      f"${_fr['annual_fish_maintenance']:,.0f}", f"${_fr['total_fish_costs']:,.0f}",
                      f"${_fr['fish_ebitda']:,.0f}", f"{_fr['fish_ebitda_margin']*100:.1f}%",
                      f"${_fr['total_fish_capex']:,.0f}", f"{_fr['annual_fish_kwh']:,.0f}",
                      f"{_fr['heating_kwh']:,.0f}", f"{_fr['ambient_temp_c']:.1f}",
                      f"{_fr['target_temp_c']:.1f}", f"{_fr['annual_n_output_g']/1000:.1f}"],
        })
        st.dataframe(
            _aq_f_df.style.apply(lambda r: [MATCH if r.name % 2 == 0 else ""] * len(r), axis=1),
            use_container_width=True, hide_index=True,
        )

    st.divider()

    # ═══════════════════════════════════════════════════════════════════════════
    # VIABILITY COMPARISON (plant side)
    # ═══════════════════════════════════════════════════════════════════════════
    st.subheader("🌍 Plant Side Viability Comparison")
    st.caption("Plant side only — uses current plant inputs. Only the dimension being compared changes.")

    _aq_comp_tab1, _aq_comp_tab2 = st.tabs(["Compare Countries", "Compare Crops"])

    with _aq_comp_tab1:
        _aq_country_metric = st.selectbox("Rank by",
            ["EBITDA", "Energy % of Revenue", "Payback (years)", "EBITDA Margin (%)"],
            key="aq_country_metric")
        _aq_country_results = []
        for _cn in COUNTRIES.keys():
            _ci = dict(aq_inputs)
            _ci["country"] = _cn
            _ci["plant_crop"] = aq_plant_crop
            try:
                _cr = calculate_greenhouse({
                    "country": _cn, "crop": aq_inputs["plant_crop"],
                    "crop_source": aq_inputs.get("plant_crop_source","greenhouse"),
                    "footprint": aq_inputs["plant_footprint"],
                    "automation": aq_inputs["automation"],
                    "price_scenario": aq_inputs["price_scenario"],
                    "plant_price_override": 0.0, "price_override": 0.0,
                    "packaging_cost": aq_inputs["packaging_cost"],
                    "loss_rate": aq_inputs["loss_rate"],
                    "net_grow_factor": aq_inputs["net_grow_factor"],
                    "walkways_factor": aq_inputs["walkways_factor"],
                    "water_price": aq_inputs["water_price"],
                    "rent_monthly": aq_inputs["rent_monthly"],
                    "real_estate_capex": aq_inputs["real_estate_capex"],
                    "harvest_mode": aq_inputs["harvest_mode"],
                    "depreciation_years": aq_inputs["depreciation_years"],
                    "tax_rate": aq_inputs["tax_rate"], "ltv": aq_inputs["ltv"],
                    "interest_rate": aq_inputs["interest_rate"],
                    "loan_term_years": aq_inputs["loan_term_years"],
                    "discount_rate": aq_inputs["discount_rate"],
                })
                _ep = _cr["annual_energy_cost"]/_cr["annual_revenue"]*100 if _cr["annual_revenue"]>0 else 999
                _aq_country_results.append({
                    "Country": _cn, "EBITDA": _cr["ebitda"],
                    "Energy % of Revenue": _ep,
                    "Payback (years)": _cr["payback_years"] if _cr["payback_years"] else 999,
                    "EBITDA Margin (%)": _cr["ebitda_margin"]*100,
                })
            except Exception:
                continue
        if _aq_country_results:
            _aq_df_c = pd.DataFrame(_aq_country_results)
            _aq_asc  = _aq_country_metric in ("Energy % of Revenue","Payback (years)")
            _aq_df_c = _aq_df_c.sort_values(_aq_country_metric, ascending=_aq_asc).reset_index(drop=True)
            _aq_fig_c = go.Figure(go.Bar(
                x=_aq_df_c[_aq_country_metric], y=_aq_df_c["Country"], orientation="h",
                marker_color=["rgba(0,229,160,0.75)" if (
                    (_aq_country_metric=="EBITDA" and v>=0) or
                    (_aq_country_metric=="Energy % of Revenue" and v<40) or
                    (_aq_country_metric=="Payback (years)" and 0<v<10) or
                    (_aq_country_metric=="EBITDA Margin (%)" and v>=0)
                ) else "rgba(255,77,77,0.75)" for v in _aq_df_c[_aq_country_metric]],
                text=_aq_df_c[_aq_country_metric].apply(
                    lambda v: f"${v:,.0f}" if _aq_country_metric=="EBITDA"
                    else (f"{v:.1f}%" if "%" in _aq_country_metric
                    else (f"{v:.1f} yrs" if v<900 else "N/A"))),
                textposition="outside"))
            if _aq_country_metric == "Energy % of Revenue":
                _aq_fig_c.add_vline(x=40, line_dash="dash", line_color="rgba(255,193,61,0.6)",
                                    annotation_text="40% threshold", annotation_font_color="#ffc13d")
            _aq_fig_c.update_layout(
                plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                font_color="#161a16", height=max(500, len(_aq_df_c)*22),
                margin=dict(l=10,r=100,t=20,b=20),
                xaxis=dict(showgrid=False,zeroline=False),
                yaxis=dict(showgrid=False,autorange="reversed"))
            style_fig(_aq_fig_c)
            st.plotly_chart(_aq_fig_c, use_container_width=True)

    with _aq_comp_tab2:
        _aq_plant_dict_comp = POLYTUNNEL_CROPS if aq_inputs.get("plant_crop_source","greenhouse")=="polytunnel" else GREENHOUSE_CROPS
        _aq_crop_metric = st.selectbox("Rank by",
            ["EBITDA","EBITDA Margin (%)","Energy % of Revenue","Payback (years)"],
            key="aq_crop_metric")
        _aq_crop_results = []
        for _crn in _aq_plant_dict_comp.keys():
            try:
                _crr = calculate_greenhouse({
                    "country": aq_inputs["country"], "crop": _crn,
                    "crop_source": aq_inputs.get("plant_crop_source","greenhouse"),
                    "footprint": aq_inputs["plant_footprint"],
                    "automation": aq_inputs["automation"],
                    "price_scenario": aq_inputs["price_scenario"], "price_override": 0.0,
                    "packaging_cost": aq_inputs["packaging_cost"],
                    "loss_rate": aq_inputs["loss_rate"],
                    "net_grow_factor": aq_inputs["net_grow_factor"],
                    "walkways_factor": aq_inputs["walkways_factor"],
                    "water_price": aq_inputs["water_price"],
                    "rent_monthly": 0.0, "real_estate_capex": 0.0,
                    "harvest_mode": "Single",
                    "depreciation_years": aq_inputs["depreciation_years"],
                    "tax_rate": aq_inputs["tax_rate"], "ltv": aq_inputs["ltv"],
                    "interest_rate": aq_inputs["interest_rate"],
                    "loan_term_years": aq_inputs["loan_term_years"],
                    "discount_rate": aq_inputs["discount_rate"],
                })
                _ep2 = _crr["annual_energy_cost"]/_crr["annual_revenue"]*100 if _crr["annual_revenue"]>0 else 999
                _aq_crop_results.append({
                    "Crop": _crn, "EBITDA": _crr["ebitda"],
                    "EBITDA Margin (%)": _crr["ebitda_margin"]*100,
                    "Energy % of Revenue": _ep2,
                    "Payback (years)": _crr["payback_years"] if _crr["payback_years"] else 999,
                })
            except Exception:
                continue
        if _aq_crop_results:
            _aq_df_cr = pd.DataFrame(_aq_crop_results)
            _aq_asc2  = _aq_crop_metric in ("Energy % of Revenue","Payback (years)")
            _aq_df_cr = _aq_df_cr.sort_values(_aq_crop_metric, ascending=_aq_asc2).reset_index(drop=True)
            _aq_fig_cr = go.Figure(go.Bar(
                x=_aq_df_cr[_aq_crop_metric], y=_aq_df_cr["Crop"], orientation="h",
                marker_color=["rgba(0,229,160,0.75)" if (
                    (_aq_crop_metric=="EBITDA" and v>=0) or
                    (_aq_crop_metric=="Energy % of Revenue" and v<40) or
                    (_aq_crop_metric=="Payback (years)" and 0<v<10) or
                    (_aq_crop_metric=="EBITDA Margin (%)" and v>=0)
                ) else "rgba(255,77,77,0.75)" for v in _aq_df_cr[_aq_crop_metric]],
                text=_aq_df_cr[_aq_crop_metric].apply(
                    lambda v: f"${v:,.0f}" if _aq_crop_metric=="EBITDA"
                    else (f"{v:.1f}%" if "%" in _aq_crop_metric
                    else (f"{v:.1f} yrs" if v<900 else "N/A"))),
                textposition="outside"))
            _aq_fig_cr.update_layout(
                plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                font_color="#161a16", height=max(400, len(_aq_df_cr)*22),
                margin=dict(l=10,r=100,t=20,b=20),
                xaxis=dict(showgrid=False,zeroline=False),
                yaxis=dict(showgrid=False,autorange="reversed"))
            style_fig(_aq_fig_cr)
            st.plotly_chart(_aq_fig_cr, use_container_width=True)

    st.divider()

    # ═══════════════════════════════════════════════════════════════════════════
    # SENSITIVITY ANALYSIS (plant side)
    # ═══════════════════════════════════════════════════════════════════════════
    st.subheader("🔬 Plant Side Sensitivity Analysis")
    st.markdown("#### Tornado Chart — Plant EBITDA Sensitivity")
    st.caption("Each bar shows how plant EBITDA changes when one variable is stressed. Fish side held constant.")

    _aq_base_ebitda = _pr["ebitda"]
    _aq_tvars = [
        {"label":"Energy Price",  "pess":_aq_run_mult(_aq_plant_sens_inputs,kwh_m=1.50)["ebitda"], "opt":_aq_run_mult(_aq_plant_sens_inputs,kwh_m=0.70)["ebitda"],
         "pess_label":"Energy price +50%","opt_label":"Energy price −30%"},
        {"label":"Selling Price", "pess":_aq_run_mult(_aq_plant_sens_inputs,prc_m=0.80)["ebitda"], "opt":_aq_run_mult(_aq_plant_sens_inputs,prc_m=1.20)["ebitda"],
         "pess_label":"Selling price −20%","opt_label":"Selling price +20%"},
        {"label":"Yield",         "pess":_aq_run_mult(_aq_plant_sens_inputs,yld_m=0.80)["ebitda"], "opt":_aq_run_mult(_aq_plant_sens_inputs,yld_m=1.20)["ebitda"],
         "pess_label":"Yield −20%","opt_label":"Yield +20%"},
        {"label":"Labour Cost",   "pess":_aq_run_mult(_aq_plant_sens_inputs,lab_m=1.30)["ebitda"], "opt":_aq_run_mult(_aq_plant_sens_inputs,lab_m=0.80)["ebitda"],
         "pess_label":"Labour cost +30%","opt_label":"Labour cost −20%"},
    ]
    for _tv in _aq_tvars:
        _tv["delta_pess"] = _tv["pess"] - _aq_base_ebitda
        _tv["delta_opt"]  = _tv["opt"]  - _aq_base_ebitda
        _tv["swing"]      = abs(_tv["delta_opt"] - _tv["delta_pess"])
    _aq_tvars.sort(key=lambda x: x["swing"], reverse=True)

    _aq_fig_torn = go.Figure()
    for _tv in _aq_tvars:
        _aq_fig_torn.add_trace(go.Bar(
            name="Pessimistic", y=[_tv["label"]], x=[_tv["delta_pess"]], orientation="h",
            marker_color="rgba(255,77,77,0.75)", showlegend=(_tv==_aq_tvars[0]),
            text=f"${_tv['delta_pess']:,.0f}", textposition="outside"))
        _aq_fig_torn.add_trace(go.Bar(
            name="Optimistic", y=[_tv["label"]], x=[_tv["delta_opt"]], orientation="h",
            marker_color="rgba(0,229,160,0.75)", showlegend=(_tv==_aq_tvars[0]),
            text=f"${_tv['delta_opt']:,.0f}", textposition="outside"))
    _aq_fig_torn.add_vline(x=0, line_dash="dash", line_color="rgba(255,255,255,0.4)")
    _aq_fig_torn.update_layout(
        barmode="overlay", plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        font_color="#161a16", height=320, margin=dict(l=10,r=80,t=20,b=20),
        xaxis=dict(title="Plant EBITDA delta ($)",showgrid=False,zeroline=False),
        yaxis=dict(showgrid=False),
        legend=dict(orientation="h",yanchor="bottom",y=1.02,xanchor="right",x=1))
    style_fig(_aq_fig_torn)
    st.plotly_chart(_aq_fig_torn, use_container_width=True)

    st.divider()
    st.markdown("#### Scenario Comparison")
    st.caption("Stress the plant side. Fish side is always recalculated at base parameters.")

    if "aq_scenarios" not in st.session_state:
        st.session_state["aq_scenarios"] = []

    with st.expander("➕ Define a new scenario", expanded=len(st.session_state["aq_scenarios"])==0):
        _aq_sc_name   = st.text_input("Scenario name", value="Scenario 1", key="aq_sc_name")
        _aq_sc1, _aq_sc2 = st.columns(2)
        with _aq_sc1:
            _aq_sc_energy = st.slider("Energy price multiplier",  0.3, 3.0, 1.0, 0.05, key="aq_sc_energy")
            _aq_sc_yield  = st.slider("Yield multiplier",         0.3, 2.0, 1.0, 0.05, key="aq_sc_yield")
        with _aq_sc2:
            _aq_sc_price  = st.slider("Selling price multiplier", 0.3, 2.0, 1.0, 0.05, key="aq_sc_price")
            _aq_sc_labour = st.slider("Labour cost multiplier",   0.3, 3.0, 1.0, 0.05, key="aq_sc_labour")
        if st.button("💾 Save Scenario", key="aq_save_sc", use_container_width=True):
            if len(st.session_state["aq_scenarios"]) >= 4:
                st.warning("Maximum 4 scenarios reached.")
            else:
                _aq_sc_plant_r = _aq_run_mult(_aq_plant_sens_inputs,
                    kwh_m=_aq_sc_energy, lab_m=_aq_sc_labour,
                    yld_m=_aq_sc_yield, prc_m=_aq_sc_price)
                st.session_state["aq_scenarios"].append({
                    "name": _aq_sc_name, "energy_mult": _aq_sc_energy,
                    "yield_mult": _aq_sc_yield, "price_mult": _aq_sc_price,
                    "labour_mult": _aq_sc_labour, "plant_result": _aq_sc_plant_r})
                st.success(f"Scenario '{_aq_sc_name}' saved.")
        if st.session_state["aq_scenarios"]:
            if st.button("🗑️ Clear all scenarios", key="aq_clear_sc", use_container_width=True):
                st.session_state["aq_scenarios"] = []
                st.rerun()

    if st.session_state["aq_scenarios"]:
        for _idx2, _sc2 in enumerate(st.session_state["aq_scenarios"]):
            _dc1, _dc2 = st.columns([8,1])
            with _dc1:
                st.caption(f"**{_sc2['name']}** — energy ×{_sc2['energy_mult']} / yield ×{_sc2['yield_mult']} / price ×{_sc2['price_mult']} / labour ×{_sc2['labour_mult']}")
            with _dc2:
                if st.button("🗑️", key=f"aq_del_sc_{_idx2}"):
                    st.session_state["aq_scenarios"].pop(_idx2)
                    st.rerun()
        _aq_sc_names   = ["Base Case"] + [s["name"] for s in st.session_state["aq_scenarios"]]
        _aq_sc_presults = [_pr] + [s["plant_result"] for s in st.session_state["aq_scenarios"]]
        _aq_fig_comp = go.Figure()
        for _i3, (_mn3, _mv3) in enumerate({
            "Plant EBITDA":  [res["ebitda"]            for res in _aq_sc_presults],
            "Revenue":       [res["annual_revenue"]     for res in _aq_sc_presults],
            "Energy Cost":   [res["annual_energy_cost"] for res in _aq_sc_presults],
            "Labour Cost":   [res["annual_labour_cost"] for res in _aq_sc_presults],
        }.items()):
            _aq_fig_comp.add_trace(go.Bar(name=_mn3, x=_aq_sc_names, y=_mv3,
                marker_color=["#00e5a0","#ffc13d","#4fc3f7","#ba68c8"][_i3],
                text=[f"${v:,.0f}" for v in _mv3], textposition="outside"))
        _aq_fig_comp.update_layout(
            barmode="group", plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            font_color="#161a16", height=420, margin=dict(t=30,b=20),
            xaxis=dict(showgrid=False), yaxis=dict(showgrid=False,title="$"),
            legend=dict(orientation="h",yanchor="bottom",y=1.02,xanchor="right",x=1))
        style_fig(_aq_fig_comp)
        st.plotly_chart(_aq_fig_comp, use_container_width=True)

        _aq_sc_rows = []
        for _sn3, _sr3 in zip(_aq_sc_names, _aq_sc_presults):
            _ep4 = f"{_sr3['annual_energy_cost']/_sr3['annual_revenue']*100:.1f}%" if _sr3["annual_revenue"]>0 else "N/A"
            _aq_sc_rows.append({
                "Scenario": _sn3,
                "Revenue":  f"${_sr3['annual_revenue']:,.0f}",
                "Energy":   f"${_sr3['annual_energy_cost']:,.0f}",
                "Labour":   f"${_sr3['annual_labour_cost']:,.0f}",
                "Plant EBITDA": f"${_sr3['ebitda']:,.0f}",
                "Margin":   f"{_sr3['ebitda_margin']*100:.1f}%",
                "Energy % Rev": _ep4,
            })
        _aq_sc_df = pd.DataFrame(_aq_sc_rows)
        def _aq_highlight_ep(row):
            style = severity_cell(row["Energy % Rev"], hi=60, mid=40)
            return [style] * len(row)
        st.dataframe(_aq_sc_df.style.apply(_aq_highlight_ep,axis=1),
                     use_container_width=True, hide_index=True)
    else:
        st.info("No scenarios saved yet. Define and save your first scenario above.")

    st.divider()

    # ── SAVE AS FARM PROFILE ──────────────────────────────────────────────────
    aq_save_col1, aq_save_col2 = st.columns([5, 1])
    with aq_save_col2:
        if st.button("💾 Save as Farm Profile", key="aq_save_btn", use_container_width=True): # Keep emoji in button
            st.session_state["aq_show_save_form"] = True

    if st.session_state["aq_show_save_form"]:
        with st.container(border=True):
            _aq_active   = st.session_state.get("active_farm")
            _save_lat = st.session_state.get("shared_lat")
            _save_lon = st.session_state.get("shared_lng")
            _climate_data = {}
            if _save_lat and _save_lon:
                with st.spinner("🌤️ Fetching climate profile for this location…"):
                    try: # Keep emoji in spinner
                        from core.climate import fetch_climate_profile
                        _climate_data = fetch_climate_profile(_save_lat, _save_lon)
                    except Exception:
                        _climate_data = {}
            _aq_modality = f"aquaponics_{_aq_mode}"
            _aq_meta     = json.dumps({
                "species":                 aq_species,
                "tank_volume_m3":          aq_tank_volume,
                "system_scale":            aq_system_scale,
                "target_temp_c":           aq_target_temp,
                "fish_depreciation_years": aq_fish_dep,
                "plant_crop_source":       aq_plant_crop_source.lower(),
            })
            _aq_snap = json.dumps({
                "plant": {k: v for k, v in _pr.items()},
                "fish":  {k: v for k, v in _fr.items()},
                "combined_revenue":  aq_r["combined_revenue"],
                "combined_ebitda":   aq_r["combined_ebitda"],
                "combined_capex":    aq_r["combined_capex"],
            })
            _aq_payload = {
                "country":           aq_country,
                "crop":              (_aq_crop_mix[0]["crop"] if _aq_mix_valid and _aq_crop_mix else aq_plant_crop),
                "crop_source":       aq_plant_crop_source.lower(),
                "footprint":         aq_plant_footprint,
                "automation":        aq_automation,
                "price_scenario":    aq_price_scenario,
                "price_override":    0.0,
                "packaging_cost":    aq_packaging_cost,
                "loss_rate":         aq_loss_rate,
                "net_grow_factor":   aq_net_grow_factor,
                "walkways_factor":   aq_walkways_factor,
                "water_price":       aq_water_price,
                "rent_monthly":      aq_rent_monthly,
                "real_estate_capex": aq_real_estate_capex,
                "harvest_mode":      aq_harvest_mode,
                "depreciation_years": aq_dep_years,
                "tax_rate":          aq_tax_rate,
                "ltv":               aq_ltv,
                "interest_rate":     aq_interest,
                "loan_term_years":   aq_loan_term,
                "lat":               st.session_state.get("shared_lat"),
                "lon":               st.session_state.get("shared_lng"),
                "ambient_temp_annual": _climate_data.get("ambient_temp_annual"),
                "mean_annual_dli":     _climate_data.get("mean_annual_dli"),
                "agriculture_type":  _aq_modality,
                "modality":          _aq_modality,
                "discount_rate":     aq_discount,
                "metadata":          _aq_meta,
                "model_snapshot":    _aq_snap,
                "model_updated_at":  date.today().isoformat(),
                "crop_mix_json":     json.dumps(_aq_crop_mix) if _aq_mix_valid else None,
                "notes":             None,
            }
            if _aq_active:
                st.markdown(f"**Update** existing farm **{_aq_active['name']}**, or save as a new profile.")
                _au1, _au2, _au3 = st.columns([2, 2, 1])
                with _au1:
                    if st.button("✅ Update existing farm", use_container_width=True, key="aq_update_btn"):
                        try: # Keep emoji in success message
                            supabase.table("farms").update(_aq_payload).eq("id", _aq_active["id"]).execute()
                            st.session_state["active_farm"] = {**_aq_active, **_aq_payload}
                            st.success(f"✅ Farm **{_aq_active['name']}** updated.")
                            if _climate_data.get("mean_annual_dli"):
                                st.caption(f"🌤️ Climate profile saved — Mean DLI: {_climate_data['mean_annual_dli']:.1f} mol/m²/day · Ambient temp: {_climate_data['ambient_temp_annual']:.1f}°C")
                            st.session_state["aq_show_save_form"] = False
                            st.rerun()
                        except Exception as e:
                            st.error(f"Update failed: {e}")
                with _au2:
                    aq_farm_name = st.text_input("New farm name", key="aq_farm_name_input", placeholder="Enter name for new profile")
                    if st.button("➕ Save as new farm", use_container_width=True, key="aq_saveas_btn"):
                        if not aq_farm_name.strip():
                            st.error("Please enter a name for the new farm profile.") # Keep emoji in success message
                        else:
                            try:
                                supabase.table("farms").insert({**_aq_payload, "name": aq_farm_name.strip(), "owner_id": current_user()}).execute()
                                st.success(f"✅ New farm profile '{aq_farm_name.strip()}' saved.")
                                if _climate_data.get("mean_annual_dli"):
                                    st.caption(f"🌤️ Climate profile saved — Mean DLI: {_climate_data['mean_annual_dli']:.1f} mol/m²/day · Ambient temp: {_climate_data['ambient_temp_annual']:.1f}°C")
                                st.session_state["aq_show_save_form"] = False
                                st.rerun()
                            except Exception as e:
                                st.error(f"Could not save: {e}")
                with _au3:
                    if st.button("✖ Cancel", use_container_width=True, key="aq_cancel_save"):
                        st.session_state["aq_show_save_form"] = False
                        st.rerun()
            else:
                st.markdown("**Save current aquaponics configuration as a Farm Profile**")
                st.caption("Saves all parameters so you can track harvests in the Harvest Tracker.")
                aq_farm_name = st.text_input("Farm name", key="aq_farm_name_input")
                _an1, _an2 = st.columns([3, 1])
                with _an1:
                    if st.button("✅ Confirm Save", key="aq_confirm_save", use_container_width=True):
                        if not aq_farm_name.strip():
                            st.error("Please enter a farm name.")
                        else:
                            try:
                                supabase.table("farms").insert({**_aq_payload, "name": aq_farm_name.strip(), "owner_id": current_user()}).execute()
                                st.success(f"Farm profile '{aq_farm_name.strip()}' saved.")
                                if _climate_data.get("mean_annual_dli"):
                                    st.caption(f"🌤️ Climate profile saved — Mean DLI: {_climate_data['mean_annual_dli']:.1f} mol/m²/day · Ambient temp: {_climate_data['ambient_temp_annual']:.1f}°C")
                                st.session_state["aq_show_save_form"] = False
                                st.rerun()
                            except Exception as e:
                                st.error(f"Could not save: {e}")
                with _an2:
                    if st.button("✖ Cancel", key="aq_cancel_new", use_container_width=True):
                        st.session_state["aq_show_save_form"] = False
                        st.rerun()
